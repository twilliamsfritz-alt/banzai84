from __future__ import annotations

import io
import json
import os
import random
import sqlite3
import zipfile

try:
    from dotenv import load_dotenv
except Exception:
    load_dotenv = None

try:
    import requests
except Exception:
    requests = None

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    import stripe
except Exception:
    stripe = None
from contextlib import closing
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, redirect, render_template, render_template_string, request, send_file, send_from_directory, session
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "banzai.db"

if load_dotenv is not None:
    load_dotenv(BASE_DIR / ".env")

SECRET_KEY = os.environ.get("BANZAI_SECRET_KEY", "change-me-in-production")
APP_VERSION = "5.0.0-integration-ready"
APP_NAME = "Banzai"

# ── Role permission map ────────────────────────────────────────────────────
# owner  — dueño del negocio: acceso total
# seller — vendedor: puede hacer demos, ver conversaciones, NO ve costos ni finanzas
# viewer — solo lectura: ve dashboard y reportes
# demo   — acceso temporal con vencimiento, solo para demos a clientes
ROLE_PERMISSIONS = {
    "owner": {
        "see_costs": True, "see_finances": True, "export": True,
        "manage_products": True, "manage_users": True, "manage_settings": True,
        "manage_billing": True, "run_demo": True, "see_deals": True,
        "see_invoices": True, "see_agents": True, "send_messages": True,
        "watermark": False,
    },
    "seller": {
        "see_costs": False, "see_finances": False, "export": False,
        "manage_products": False, "manage_users": False, "manage_settings": False,
        "manage_billing": False, "run_demo": True, "see_deals": True,
        "see_invoices": False, "see_agents": False, "send_messages": True,
        "watermark": True,
    },
    "viewer": {
        "see_costs": False, "see_finances": True, "export": False,
        "manage_products": False, "manage_users": False, "manage_settings": False,
        "manage_billing": False, "run_demo": False, "see_deals": True,
        "see_invoices": True, "see_agents": True, "send_messages": False,
        "watermark": False,
    },
    "demo": {
        "see_costs": False, "see_finances": False, "export": False,
        "manage_products": False, "manage_users": False, "manage_settings": False,
        "manage_billing": False, "run_demo": True, "see_deals": True,
        "see_invoices": False, "see_agents": False, "send_messages": True,
        "watermark": True,
    },
}


def user_can(user: dict, permission: str) -> bool:
    role = user.get("role", "viewer")
    base = ROLE_PERMISSIONS.get(role, ROLE_PERMISSIONS["viewer"])
    perm = base.get(permission, False)
    override_map = {
        "see_costs": "can_see_costs",
        "see_finances": "can_see_finances",
        "export": "can_export",
        "manage_products": "can_manage_products",
        "manage_users": "can_manage_users",
    }
    if permission in override_map:
        val = user.get(override_map[permission])
        if val is not None:
            return bool(val)
    return perm


def require_permission(user: dict, permission: str):
    if not user_can(user, permission):
        raise PermissionError(f"Permission denied: {permission}")


# ── Subscription Plans ────────────────────────────────────────────────────
PLAN_FEATURES = {
    "trial":      {"name": "Starter Gratuito",  "price": 0,   "currency": "USD", "features": ["agent_center","pipeline","contacts","deals"],                                                                                                 "max_users": 2},
    "starter":    {"name": "Starter",            "price": 99,  "currency": "USD", "features": ["agent_center","pipeline","contacts","deals"],                                                                                                 "max_users": 2},
    "basic":      {"name": "Basic",              "price": 149, "currency": "USD", "features": ["agent_center","pipeline","contacts","deals","broadcast"],                                                                                      "max_users": 3},
    "pro":        {"name": "Pro",                "price": 249, "currency": "USD", "features": ["agent_center","pipeline","contacts","deals","broadcast","inventory"],                                                                          "max_users": 5},
    "growth":     {"name": "Growth",             "price": 399, "currency": "USD", "features": ["agent_center","pipeline","contacts","deals","broadcast","inventory","finance","automation","surveys","export","tasks"],                         "max_users": 10},
    "enterprise": {"name": "Enterprise",         "price": 799, "currency": "USD", "features": ["agent_center","pipeline","contacts","deals","broadcast","inventory","finance","automation","surveys","export","tasks","advisor","white_label"],  "max_users": 9999},
}

FEATURE_LABELS = {
    "agent_center": "Agent Center (Sales AI)", "pipeline": "Pipeline Kanban",
    "contacts": "Contactos 360", "deals": "Deals", "broadcast": "Broadcast",
    "inventory": "Inventario", "finance": "Finanzas y P&L", "automation": "Automatizaciones",
    "surveys": "Encuestas NPS", "export": "Export Excel", "tasks": "Tareas",
    "advisor": "Asesor IA", "white_label": "White Label",
}


def get_workspace_plan(workspace_id):
    try:
        with closing(get_db()) as conn:
            row = conn.execute("SELECT plan, trial_ends_at FROM workspaces WHERE id=?", (workspace_id,)).fetchone()
            if not row:
                return "trial"
            plan = row["plan"] if row["plan"] else "trial"
            if plan == "trial" and row["trial_ends_at"]:
                try:
                    import datetime as _dtm
                    if _dtm.datetime.utcnow() > _dtm.datetime.fromisoformat(row["trial_ends_at"]):
                        return "expired"
                except Exception:
                    pass
            return plan
    except Exception:
        return "trial"


def workspace_has_feature(workspace_id, feature):
    plan = get_workspace_plan(workspace_id)
    if plan == "expired":
        return False
    return feature in PLAN_FEATURES.get(plan, PLAN_FEATURES["trial"]).get("features", [])


def plan_blocked_response(workspace_id, feature):
    plan = get_workspace_plan(workspace_id)
    plan_info = PLAN_FEATURES.get(plan, {})
    needed = "growth"
    for p, v in PLAN_FEATURES.items():
        if feature in v.get("features", []):
            needed = p
            break
    needed_info = PLAN_FEATURES.get(needed, {})
    return jsonify({
        "ok": False, "plan_required": True,
        "current_plan": plan, "current_plan_name": plan_info.get("name", plan),
        "required_plan": needed, "required_plan_name": needed_info.get("name", needed),
        "required_price": needed_info.get("price", 0),
        "feature": feature, "feature_label": FEATURE_LABELS.get(feature, feature),
        "error": "Esta funcion requiere el plan " + needed_info.get("name", needed) + " ($" + str(needed_info.get("price", 0)) + "/mes)"
    }), 403



def log_access(user_id: int, workspace_id: int, action: str,
               resource: str = None, result: str = "ok", detail: str = None):
    try:
        now = datetime.utcnow().isoformat()
        ip = request.remote_addr or ""
        ua = request.headers.get("User-Agent", "")[:200]
        with closing(get_db()) as conn:
            conn.execute(
                "INSERT INTO access_log (user_id,workspace_id,action,resource,ip,user_agent,result,detail,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
                (user_id, workspace_id, action, resource, ip, ua, result, detail, now)
            )
            conn.commit()
    except Exception:
        pass


def get_release_name(version: str = APP_VERSION) -> str:
    """Get human release name for a version. Falls back to version string."""
    try:
        with closing(get_db()) as conn:
            row = conn.execute("SELECT release_name FROM release_names WHERE version = ?", (version,)).fetchone()
            return row["release_name"] if row else version
    except Exception:
        return version
APP_EDITION = "Integration-Ready Official Candidate"
APP_URL = os.environ.get("APP_URL", "http://127.0.0.1:5000")

app = Flask(__name__, template_folder="templates", static_folder="static")
app.secret_key = SECRET_KEY
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("FLASK_ENV", "production") == "production"
app.config["MAX_CONTENT_LENGTH"] = 12 * 1024 * 1024  # 12MB max upload size


# No demo workspaces — first run triggers setup wizard
WORKSPACES = []


SCHEMA = """
CREATE TABLE IF NOT EXISTS workspaces (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    slug TEXT UNIQUE NOT NULL,
    name TEXT NOT NULL,
    region TEXT NOT NULL,
    currency TEXT NOT NULL,
    language TEXT NOT NULL,
    tone TEXT NOT NULL,
    plan TEXT NOT NULL DEFAULT 'trial',
    trial_ends_at TEXT,
    plan_expires_at TEXT,
    whatsapp_number TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    email TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    name TEXT NOT NULL,
    role TEXT NOT NULL DEFAULT 'owner',
    -- roles: owner | seller | viewer | demo
    -- owner:  full access, billing, user management
    -- seller: demo access, send messages, view conversations — NO products costs, NO finances, NO settings
    -- viewer: read-only dashboard and reports — NO editing anything
    -- demo:   can only run demos, expires after N days
    active INTEGER NOT NULL DEFAULT 1,
    demo_expires_at TEXT,
    -- seller-specific limits
    can_see_costs INTEGER NOT NULL DEFAULT 0,
    can_see_finances INTEGER NOT NULL DEFAULT 0,
    can_export INTEGER NOT NULL DEFAULT 0,
    can_manage_products INTEGER NOT NULL DEFAULT 0,
    can_manage_users INTEGER NOT NULL DEFAULT 0,
    watermark_demos INTEGER NOT NULL DEFAULT 1,
    -- security audit
    last_login_at TEXT,
    last_login_ip TEXT,
    failed_login_count INTEGER NOT NULL DEFAULT 0,
    locked_until TEXT,
    invited_by INTEGER,
    invite_token TEXT UNIQUE,
    invite_used INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS access_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    workspace_id INTEGER NOT NULL,
    action TEXT NOT NULL,
    resource TEXT,
    ip TEXT,
    user_agent TEXT,
    result TEXT NOT NULL DEFAULT 'ok',
    detail TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS workspace_profiles (
    workspace_id INTEGER PRIMARY KEY,
    response_style TEXT NOT NULL,
    personality_notes TEXT NOT NULL,
    forbidden_tone TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    sku TEXT NOT NULL,
    name TEXT NOT NULL,
    category TEXT NOT NULL,
    cost REAL NOT NULL,
    price REAL NOT NULL,
    stock INTEGER NOT NULL DEFAULT 0,
    stock_min INTEGER NOT NULL DEFAULT 0,
    unit TEXT NOT NULL DEFAULT 'unit',
    barcode TEXT,
    competitor_price REAL NOT NULL DEFAULT 0,
    demand_score INTEGER NOT NULL DEFAULT 60,
    active INTEGER NOT NULL DEFAULT 1,
    notes TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS stock_movements (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    product_id INTEGER NOT NULL,
    delta INTEGER NOT NULL,
    stock_after INTEGER NOT NULL,
    movement_type TEXT NOT NULL DEFAULT 'manual',
    note TEXT,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id),
    FOREIGN KEY (product_id) REFERENCES products(id)
);

CREATE TABLE IF NOT EXISTS knowledge_articles (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    title TEXT NOT NULL,
    content TEXT NOT NULL,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS source_documents (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    title TEXT NOT NULL,
    domain TEXT NOT NULL,
    source_type TEXT NOT NULL,
    content TEXT NOT NULL,
    excerpt TEXT NOT NULL,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS templates (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    category TEXT NOT NULL,
    body TEXT NOT NULL,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS conversations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    customer_name TEXT NOT NULL,
    customer_phone TEXT,
    channel TEXT NOT NULL,
    status TEXT NOT NULL,
    country TEXT NOT NULL,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS messages (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    conversation_id INTEGER NOT NULL,
    role TEXT NOT NULL,
    text TEXT NOT NULL,
    created_at TEXT NOT NULL,
    FOREIGN KEY (conversation_id) REFERENCES conversations(id)
);

CREATE TABLE IF NOT EXISTS tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    title TEXT NOT NULL,
    area TEXT NOT NULL,
    owner TEXT NOT NULL,
    status TEXT NOT NULL,
    priority TEXT NOT NULL,
    impact INTEGER NOT NULL,
    due_label TEXT NOT NULL,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS ledger_entries (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    entry_type TEXT NOT NULL,
    concept TEXT NOT NULL,
    category TEXT NOT NULL,
    amount REAL NOT NULL,
    currency TEXT NOT NULL,
    state TEXT NOT NULL,
    due_date TEXT NOT NULL,
    source TEXT NOT NULL DEFAULT 'manual',
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS traces (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    flow TEXT NOT NULL,
    customer TEXT NOT NULL,
    status TEXT NOT NULL,
    detail TEXT NOT NULL,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS app_meta (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS advisor_insights_cache (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL UNIQUE,
    insights_json TEXT NOT NULL,
    context_json TEXT NOT NULL,
    source TEXT NOT NULL,
    generated_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS industry_playbooks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL DEFAULT 0,
    -- 0 = global (all workspaces), specific id = per-workspace override
    slug TEXT NOT NULL,
    name TEXT NOT NULL,
    data_json TEXT NOT NULL,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    UNIQUE(workspace_id, slug)
);

CREATE TABLE IF NOT EXISTS billing_plans (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL DEFAULT 0,
    name TEXT NOT NULL,
    description TEXT,
    price REAL NOT NULL DEFAULT 0,
    currency TEXT NOT NULL DEFAULT 'USD',
    interval TEXT NOT NULL DEFAULT 'month',
    stripe_price_id TEXT,
    mp_preference_id TEXT,
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS billing_payments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    plan_id INTEGER,
    customer_name TEXT NOT NULL,
    customer_email TEXT NOT NULL,
    amount REAL NOT NULL,
    currency TEXT NOT NULL,
    method TEXT NOT NULL DEFAULT 'manual',
    status TEXT NOT NULL DEFAULT 'pending',
    reference TEXT,
    notes TEXT,
    banzai_invoice_number TEXT,
    paid_at TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS vendor_invoices (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    vendor_name TEXT,
    invoice_number TEXT,
    invoice_date TEXT,
    due_date TEXT,
    amount REAL,
    currency TEXT DEFAULT 'USD',
    status TEXT NOT NULL DEFAULT 'pending',
    source_email TEXT,
    source_subject TEXT,
    raw_text TEXT,
    file_name TEXT,
    file_base64 TEXT,
    file_mime_type TEXT DEFAULT 'application/pdf',
    ledger_entry_id INTEGER,
    extracted_confidence REAL DEFAULT 0,
    needs_review INTEGER DEFAULT 0,
    processed_at TEXT NOT NULL,
    paid_at TEXT,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS email_inbox_config (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL UNIQUE,
    imap_host TEXT,
    imap_port INTEGER DEFAULT 993,
    email_address TEXT,
    email_password TEXT,
    last_checked_at TEXT,
    active INTEGER DEFAULT 1,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS release_names (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    version TEXT NOT NULL UNIQUE,
    release_name TEXT NOT NULL,
    description TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS feedback_surveys (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    token TEXT NOT NULL UNIQUE,
    customer_name TEXT NOT NULL,
    customer_email TEXT NOT NULL,
    sent_at TEXT NOT NULL,
    responded_at TEXT,
    nps_score INTEGER,
    response_text TEXT,
    suggested_update TEXT,
    status TEXT NOT NULL DEFAULT 'sent',
    update_status TEXT NOT NULL DEFAULT 'pending',
    update_activated_at TEXT
);

CREATE TABLE IF NOT EXISTS pipelines (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    description TEXT,
    stages_json TEXT NOT NULL DEFAULT '["Nuevo","Contactado","Demo","Propuesta","Negociación","Cerrado","Perdido"]',
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS pipeline_cards (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    pipeline_id INTEGER NOT NULL,
    conversation_id INTEGER,
    customer_name TEXT NOT NULL,
    deal_value REAL NOT NULL DEFAULT 0,
    currency TEXT NOT NULL DEFAULT 'USD',
    stage TEXT NOT NULL DEFAULT 'Nuevo',
    assigned_to INTEGER,
    notes TEXT,
    expected_close TEXT,
    probability INTEGER NOT NULL DEFAULT 50,
    tags TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS automations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    trigger_type TEXT NOT NULL,
    trigger_config TEXT NOT NULL DEFAULT '{}',
    action_type TEXT NOT NULL,
    action_config TEXT NOT NULL DEFAULT '{}',
    active INTEGER NOT NULL DEFAULT 1,
    run_count INTEGER NOT NULL DEFAULT 0,
    last_run_at TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS automation_runs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    automation_id INTEGER NOT NULL,
    workspace_id INTEGER NOT NULL,
    trigger_data TEXT,
    result TEXT NOT NULL DEFAULT 'ok',
    detail TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS goals (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    metric TEXT NOT NULL,
    target_value REAL NOT NULL,
    current_value REAL NOT NULL DEFAULT 0,
    period TEXT NOT NULL DEFAULT 'month',
    start_date TEXT NOT NULL,
    end_date TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'active',
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS contacts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    email TEXT,
    phone TEXT,
    company TEXT,
    role TEXT,
    tags TEXT,
    notes TEXT,
    nps_score INTEGER,
    total_revenue REAL NOT NULL DEFAULT 0,
    conversation_count INTEGER NOT NULL DEFAULT 0,
    last_contact_at TEXT,
    source TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS setup (
    id INTEGER PRIMARY KEY CHECK (id = 1),
    completed INTEGER NOT NULL DEFAULT 0,
    workspace_name TEXT,
    owner_name TEXT,
    owner_email TEXT,
    region TEXT,
    currency TEXT,
    language TEXT,
    completed_at TEXT
);

-- ── Negotiation & autonomous deal engine ─────────────────────────────────
CREATE TABLE IF NOT EXISTS deals (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    conversation_id INTEGER,
    customer_name TEXT NOT NULL,
    channel TEXT NOT NULL DEFAULT 'webchat',
    status TEXT NOT NULL DEFAULT 'negotiating',
    items_json TEXT NOT NULL DEFAULT '[]',
    original_total REAL NOT NULL DEFAULT 0,
    negotiated_total REAL NOT NULL DEFAULT 0,
    discount_pct REAL NOT NULL DEFAULT 0,
    margin_floor REAL NOT NULL DEFAULT 0,
    currency TEXT NOT NULL DEFAULT 'USD',
    closed_at TEXT,
    expires_at TEXT,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

CREATE TABLE IF NOT EXISTS negotiation_margins (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    product_id INTEGER,
    min_margin_pct REAL NOT NULL DEFAULT 0.25,
    max_discount_pct REAL NOT NULL DEFAULT 0.15,
    auto_approve_below_pct REAL NOT NULL DEFAULT 0.05,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

-- ── Invoices & fiscal records ─────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS invoices (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    deal_id INTEGER,
    ledger_entry_id INTEGER,
    invoice_number TEXT NOT NULL,
    customer_name TEXT NOT NULL,
    customer_tax_id TEXT,
    items_json TEXT NOT NULL DEFAULT '[]',
    subtotal REAL NOT NULL DEFAULT 0,
    tax_pct REAL NOT NULL DEFAULT 0,
    tax_amount REAL NOT NULL DEFAULT 0,
    total REAL NOT NULL DEFAULT 0,
    currency TEXT NOT NULL,
    country_code TEXT NOT NULL DEFAULT 'US',
    fiscal_regime TEXT,
    status TEXT NOT NULL DEFAULT 'draft',
    stripe_payment_intent TEXT,
    issued_at TEXT,
    paid_at TEXT,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

-- ── Multi-agent audit trail ──────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS agent_events (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    agent TEXT NOT NULL,
    action TEXT NOT NULL,
    entity_type TEXT,
    entity_id INTEGER,
    input_summary TEXT,
    output_summary TEXT,
    confidence REAL,
    requires_human_review INTEGER NOT NULL DEFAULT 0,
    reviewed_at TEXT,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

-- ── Value-based commission billing ──────────────────────────────────────
CREATE TABLE IF NOT EXISTS commission_events (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workspace_id INTEGER NOT NULL,
    deal_id INTEGER,
    invoice_id INTEGER,
    event_type TEXT NOT NULL,
    gross_value REAL NOT NULL DEFAULT 0,
    commission_pct REAL NOT NULL DEFAULT 0,
    commission_amount REAL NOT NULL DEFAULT 0,
    currency TEXT NOT NULL,
    billed INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);

-- ── Tax / fiscal profile per workspace ──────────────────────────────────
CREATE TABLE IF NOT EXISTS tax_profiles (
    workspace_id INTEGER PRIMARY KEY,
    country_code TEXT NOT NULL DEFAULT 'US',
    tax_authority TEXT,
    tax_id TEXT,
    default_tax_pct REAL NOT NULL DEFAULT 0,
    fiscal_regime TEXT,
    invoice_prefix TEXT NOT NULL DEFAULT 'INV',
    next_invoice_seq INTEGER NOT NULL DEFAULT 1,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (workspace_id) REFERENCES workspaces(id)
);
"""


# ── Database connection ───────────────────────────────────────────────────

def get_db():
    database_url = os.environ.get("DATABASE_URL", "")
    if database_url and "postgres" in database_url:
        import psycopg2
        import psycopg2.extras
        url = database_url.replace("postgres://", "postgresql://", 1)
        conn = psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
        conn.autocommit = False
        return _PgConn(conn)
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


class _PgConn:
    def __init__(self, conn):
        self._conn = conn

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            try: self._conn.rollback()
            except Exception: pass
        self.close()

    def close(self):
        try: self._conn.close()
        except Exception: pass

    def commit(self):
        self._conn.commit()

    def execute(self, sql, params=()):
        s = sql.replace("?", "%s")
        s = s.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
        if "INSERT OR IGNORE INTO" in s:
            s = s.replace("INSERT OR IGNORE INTO", "INSERT INTO")
            if "ON CONFLICT" not in s:
                s = s.rstrip() + " ON CONFLICT DO NOTHING"
        if "INSERT OR REPLACE INTO" in s:
            s = s.replace("INSERT OR REPLACE INTO", "INSERT INTO")
            if "ON CONFLICT" not in s:
                s = s.rstrip() + " ON CONFLICT DO NOTHING"
        cur = self._conn.cursor()
        try:
            cur.execute(s, params if params else ())
        except Exception:
            try: self._conn.rollback()
            except Exception: pass
            raise
        return _PgCursor(cur)

    def executescript(self, sql):
        s = sql.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
        s = s.replace("INSERT OR IGNORE INTO", "INSERT INTO")
        s = s.replace("INSERT OR REPLACE INTO", "INSERT INTO")
        cur = self._conn.cursor()
        for stmt in s.split(";"):
            stmt = stmt.strip()
            if len(stmt) < 5:
                continue
            try:
                cur.execute(stmt)
                self._conn.commit()
            except Exception:
                try: self._conn.rollback()
                except Exception: pass


class _PgCursor:
    def __init__(self, cur):
        self._cur = cur

    def fetchone(self):
        row = self._cur.fetchone()
        if row is None:
            return None
        return _Row(dict(row) if isinstance(row, dict) else dict(zip([d[0] for d in self._cur.description], row)))

    def fetchall(self):
        rows = self._cur.fetchall()
        if not rows:
            return []
        if isinstance(rows[0], dict):
            return [_Row(r) for r in rows]
        cols = [d[0] for d in self._cur.description]
        return [_Row(dict(zip(cols, r))) for r in rows]

    @property
    def lastrowid(self):
        try:
            self._cur.execute("SELECT lastval()")
            r = self._cur.fetchone()
            return list(r.values())[0] if isinstance(r, dict) else r[0]
        except Exception:
            return None

    def __iter__(self):
        return iter(self.fetchall())


class _Row(dict):
    def __getitem__(self, key):
        if isinstance(key, int):
            return list(self.values())[key]
        try:
            return super().__getitem__(key)
        except KeyError:
            for k in self.keys():
                if k.lower() == str(key).lower():
                    return super().__getitem__(k)
            raise

    def get(self, key, default=None):
        try:
            return self[key]
        except (KeyError, IndexError):
            return default



def is_setup_complete() -> bool:
    """Returns True if the first-run setup wizard has been completed."""
    try:
        with closing(get_db()) as conn:
            row = conn.execute("SELECT completed FROM setup WHERE id = 1").fetchone()
            return bool(row and row["completed"])
    except Exception:
        return False


def init_db(force_reset: bool = False) -> None:
    database_url = os.environ.get("DATABASE_URL", "")
    is_postgres = database_url and database_url.startswith("postgres")
    
    if force_reset and not is_postgres and DB_PATH.exists():
        DB_PATH.unlink()

    with closing(get_db()) as conn:
        if is_postgres:
            # For PostgreSQL, run each statement individually
            import psycopg2
            schema_pg = SCHEMA.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
            schema_pg = schema_pg.replace("INSERT OR IGNORE INTO", "INSERT INTO")
            schema_pg = schema_pg.replace("INSERT OR REPLACE INTO", "INSERT INTO")
            cur = conn._conn.cursor()
            for stmt in schema_pg.split(";"):
                stmt = stmt.strip()
                if stmt and len(stmt) > 5:
                    try:
                        cur.execute(stmt)
                        conn._conn.commit()
                    except Exception as e:
                        conn._conn.rollback()
                        # Table already exists is fine
                        if "already exists" not in str(e):
                            pass
            # Insert version
            try:
                cur.execute("INSERT INTO app_meta (key, value) VALUES (%s, %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
                           ("schema_version", APP_VERSION))
                conn._conn.commit()
            except Exception:
                conn._conn.rollback()
            try:
                cur.execute("SELECT COUNT(*) FROM workspaces")
                count = cur.fetchone()[0]
                if count:
                    return
            except Exception:
                pass
        else:
            conn.executescript(SCHEMA)
        try:
            conn.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_vendor_invoice_dedup ON vendor_invoices(workspace_id, vendor_name, invoice_number) WHERE vendor_name IS NOT NULL AND invoice_number IS NOT NULL"
            )
        except Exception:
            pass

        # Performance indexes for common lookups (safe no-ops if they already exist)
        for idx_sql in [
            "CREATE INDEX IF NOT EXISTS idx_users_email ON users(email)",
            "CREATE INDEX IF NOT EXISTS idx_users_workspace ON users(workspace_id)",
            "CREATE INDEX IF NOT EXISTS idx_conversations_workspace ON conversations(workspace_id, status)",
            "CREATE INDEX IF NOT EXISTS idx_conversations_phone ON conversations(workspace_id, customer_phone)",
            "CREATE INDEX IF NOT EXISTS idx_messages_conversation ON messages(conversation_id)",
            "CREATE INDEX IF NOT EXISTS idx_deals_workspace ON deals(workspace_id, status)",
            "CREATE INDEX IF NOT EXISTS idx_products_workspace ON products(workspace_id)",
            "CREATE INDEX IF NOT EXISTS idx_ledger_workspace ON ledger_entries(workspace_id, entry_type)",
            "CREATE INDEX IF NOT EXISTS idx_vendor_invoices_workspace ON vendor_invoices(workspace_id, status)",
            "CREATE INDEX IF NOT EXISTS idx_vendor_invoices_vendor ON vendor_invoices(workspace_id, vendor_name)",
        ]:
            try:
                conn.execute(idx_sql)
            except Exception:
                pass

        try:
            conn.execute("INSERT OR REPLACE INTO app_meta (key, value) VALUES (?, ?)", ("schema_version", APP_VERSION))
        except Exception:
            pass
            conn.commit()
            existing = conn.execute("SELECT COUNT(*) AS count FROM workspaces").fetchone()["count"]
            if existing:
                return  # DB already initialized



def auto_setup_if_needed():
    # Migrate plan columns if needed
    try:
        with closing(get_db()) as conn:
            for col_def in [
                "ALTER TABLE workspaces ADD COLUMN plan TEXT DEFAULT 'trial'",
                "ALTER TABLE workspaces ADD COLUMN trial_ends_at TEXT",
                "ALTER TABLE workspaces ADD COLUMN plan_expires_at TEXT",
                "ALTER TABLE conversations ADD COLUMN customer_phone TEXT",
                "ALTER TABLE vendor_invoices ADD COLUMN file_base64 TEXT",
                "ALTER TABLE vendor_invoices ADD COLUMN file_mime_type TEXT DEFAULT 'application/pdf'",
                "ALTER TABLE workspaces ADD COLUMN whatsapp_number TEXT",
                "ALTER TABLE ledger_entries ADD COLUMN source TEXT DEFAULT 'manual'",
            ]:
                try:
                    conn.execute(col_def)
                    conn.commit()
                except Exception:
                    pass
    except Exception:
        pass
    """Auto-create admin account on first boot if ADMIN_EMAIL env var is set."""
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@banzai84.com")
    admin_pass = os.environ.get("ADMIN_PASSWORD", "Admin2026")
    admin_name = os.environ.get("ADMIN_NAME", "Admin")
    workspace_name = os.environ.get("WORKSPACE_NAME", "Banzai84")
    
    try:
        with closing(get_db()) as conn:
            # Check if setup already done
            try:
                row = conn.execute("SELECT completed FROM setup WHERE id=1").fetchone()
                if row and (row["completed"] if isinstance(row, dict) else row[0]):
                    return  # Already set up
            except Exception:
                pass
            
            now = datetime.utcnow().isoformat()
            import re as _re
            slug = _re.sub(r"[^a-z0-9]+", "-", workspace_name.lower()).strip("-") or "workspace"
            
            # Create workspace
            try:
                conn.execute(
                    "INSERT INTO workspaces (slug,name,region,currency,language,tone,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING"
                    if os.environ.get("DATABASE_URL","").startswith("postgres")
                    else "INSERT OR IGNORE INTO workspaces (slug,name,region,currency,language,tone,created_at) VALUES (?,?,?,?,?,?,?)",
                    (slug, workspace_name, "AR", "ARS", "es", "warm_professional", now)
                )
                conn.commit()
            except Exception:
                pass
            
            # Get workspace id
            ws = conn.execute("SELECT id FROM workspaces LIMIT 1").fetchone()
            if not ws:
                return
            wid = ws["id"] if isinstance(ws, dict) else ws[0]
            
            # Create admin user
            try:
                ph = generate_password_hash(admin_pass)
                conn.execute(
                    "INSERT INTO users (workspace_id,email,password_hash,name,role,active,created_at) VALUES (%s,%s,%s,%s,%s,1,%s) ON CONFLICT DO NOTHING"
                    if os.environ.get("DATABASE_URL","").startswith("postgres")
                    else "INSERT OR IGNORE INTO users (workspace_id,email,password_hash,name,role,active,created_at) VALUES (?,?,?,?,?,1,?)",
                    (wid, admin_email, ph, admin_name, "owner", now)
                )
                conn.commit()
            except Exception:
                pass
            
            # Mark setup complete
            try:
                if os.environ.get("DATABASE_URL","").startswith("postgres"):
                    conn.execute(
                        "INSERT INTO setup (id,completed,workspace_name,owner_name,owner_email,region,currency,language,completed_at) VALUES (1,1,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT(id) DO UPDATE SET completed=1,completed_at=EXCLUDED.completed_at",
                        (workspace_name, admin_name, admin_email, "AR", "ARS", "es", now)
                    )
                else:
                    conn.execute(
                        "INSERT OR REPLACE INTO setup (id,completed,workspace_name,owner_name,owner_email,region,currency,language,completed_at) VALUES (1,1,?,?,?,?,?,?,?)",
                        (workspace_name, admin_name, admin_email, "AR", "ARS", "es", now)
                    )
                conn.commit()
            except Exception:
                pass
    except Exception as e:
        print(f"Auto-setup error (non-fatal): {e}")


def current_user() -> dict[str, Any] | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    with closing(get_db()) as conn:
        row = conn.execute(
            """
            SELECT users.id, users.email, users.name, users.role, workspaces.id AS workspace_id,
                   workspaces.slug AS workspace_slug, workspaces.name AS workspace_name,
                   workspaces.currency, workspaces.language, workspaces.tone
            FROM users JOIN workspaces ON users.workspace_id = workspaces.id
            WHERE users.id = ?
            """,
            (user_id,),
        ).fetchone()
        return dict(row) if row else None


def require_auth() -> dict[str, Any]:
    user = current_user()
    if not user:
        raise PermissionError("Authentication required")
    return user


def json_error(message: str, status: int = 400):
    return jsonify({"ok": False, "error": message}), status


def format_money(value: float, currency: str) -> str:
    symbols = {"USD": "$", "ARS": "$", "BRL": "R$", "EUR": "€", "GBP": "£"}
    return f"{symbols.get(currency, '$')}{value:,.0f}"


def fetch_workspace_sources(conn: sqlite3.Connection, workspace_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        "SELECT id, title, domain, source_type, content, excerpt, created_at FROM source_documents WHERE workspace_id = ? ORDER BY id DESC",
        (workspace_id,),
    ).fetchall()
    return [dict(r) for r in rows]


def source_snippets_for_text(text: str, sources: list[dict[str, Any]]) -> list[dict[str, Any]]:
    text_l = text.lower()
    keyword_map = {
        "sales": ["price", "precio", "quote", "budget", "presupuesto", "sell", "venta", "buyer", "lead"],
        "admin": ["task", "tarea", "owner", "assign", "agenda", "schedule", "process", "workflow"],
        "finance": ["pay", "payment", "invoice", "cash", "cobro", "pago", "finance", "ledger"],
        "pricing": ["margin", "price", "pricing", "competitor", "discount", "markup", "precio", "margen"],
        "operations": ["delivery", "dispatch", "ship", "envio", "entrega", "operations"],
    }
    scored = []
    for source in sources:
        domain = source.get("domain", "")
        score = 0
        for kw in keyword_map.get(domain, []):
            if kw in text_l:
                score += 1
        if score:
            scored.append((score, source))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [item[1] for item in scored[:2]]


def integration_status() -> dict[str, Any]:
    openai_ready = bool(os.environ.get("OPENAI_API_KEY")) and OpenAI is not None
    whatsapp_ready = all([
        os.environ.get("WHATSAPP_VERIFY_TOKEN"),
        os.environ.get("WHATSAPP_ACCESS_TOKEN"),
        os.environ.get("WHATSAPP_PHONE_NUMBER_ID"),
    ]) and requests is not None
    stripe_ready = bool(os.environ.get("STRIPE_SECRET_KEY")) and stripe is not None
    if stripe_ready:
        stripe.api_key = os.environ.get("STRIPE_SECRET_KEY")
    return {
        "openai": {"configured": openai_ready, "library_installed": OpenAI is not None, "model": os.environ.get("OPENAI_MODEL", "gpt-4o-mini")},
        "whatsapp": {
            "configured": whatsapp_ready,
            "library_installed": requests is not None,
            "phone_number_id_present": bool(os.environ.get("WHATSAPP_PHONE_NUMBER_ID")),
            "waba_id_present": bool(os.environ.get("WHATSAPP_WABA_ID")),
        },
        "stripe": {
            "configured": stripe_ready,
            "library_installed": stripe is not None,
            "publishable_key_present": bool(os.environ.get("STRIPE_PUBLISHABLE_KEY")),
            "webhook_secret_present": bool(os.environ.get("STRIPE_WEBHOOK_SECRET")),
        },
        "app_url": APP_URL,
    }


# ── Industry knowledge base for the Genius Sales Agent ──────────────────────
# ── 30-industry playbook (editable from DB — DB overrides this at runtime) ──
INDUSTRY_PLAYBOOKS = {
    "distribuidora": {"tactics":["volumen","exclusividad de zona","plazo de pago 30/60 días","entrega express","combo productos"],"objections":{"caro":"¿Cuánto pagás hoy por unidad? Te muestro el ahorro por volumen.","lo pienso":"Te separo el stock hoy, el pedido lo confirmás mañana.","tengo proveedor":"¿Qué plazo te da? Nosotros damos 30 días sin interés."},"upsell":"Con 3 cajas del combo el precio por unidad baja un 12%.","kpis":["rotación de stock","margen por SKU","días de inventario"],"keywords":["distribu","mayoris","wholesale","almacen","reparto","mayoreo"]},
    "clinica": {"tactics":["agenda sin esperas","resultados medibles","confidencialidad","seguimiento post-sesión","turnos online"],"objections":{"caro":"¿Cuánto vale resolver esto en 4 sesiones vs arrastrarlo 2 años?","no tengo tiempo":"Tenemos turnos a las 7am y por videollamada."},"upsell":"El pack trimestral tiene 15% de descuento y mejora los resultados.","kpis":["retención de pacientes","ticket promedio","cancelaciones"],"keywords":["clínica","clinica","médico","medico","salud","health","terapia","doctor","paciente","consultorio"]},
    "retail": {"tactics":["urgencia de stock limitado","bundle complementarios","programa de fidelidad","envío gratis por monto mínimo"],"objections":{"caro":"Tenemos financiación en 6 cuotas sin interés.","lo veo online más barato":"¿Incluye garantía local y soporte? El nuestro sí."},"upsell":"El 70% de los clientes también lleva el accesorio X. ¿Lo agregamos?","kpis":["ticket promedio","conversión","devoluciones"],"keywords":["tienda","store","retail","comercio","local","boutique","shop","kiosco"]},
    "agencia": {"tactics":["caso de éxito similar","ROI medible","onboarding rápido","reportes transparentes"],"objections":{"caro":"¿Cuánto cuesta no tenerlo? Calculemos el costo de oportunidad.","ya tenemos uno":"¿Qué resultados está dando? Comparemos métricas."},"upsell":"El plan anual incluye estrategia mensual y ahorra 2 meses vs mensual.","kpis":["CAC","LTV","churn mensual","NPS"],"keywords":["agencia","agency","marketing","publicidad","diseño","consultor","creativa"]},
    "construccion": {"tactics":["precio por proyecto completo","garantía de materiales","entrega por etapas","stock reservado para la obra"],"objections":{"caro":"¿Comparaste calidad? El material B dura 3x más.","demora":"Tenemos stock garantizado para toda la obra."},"upsell":"Si cerrás el pedido completo hoy te incluyo el flete sin costo.","kpis":["margen por obra","tiempo de cobro","costo de materiales / presupuesto"],"keywords":["construc","obra","albañil","albañileria","contratista","cuadrilla","metro cuadrado"]},
    "gastro": {"tactics":["precio por volumen semanal","exclusividad de producto","entrega antes de las 7am","calidad certificada"],"objections":{"caro":"¿Cuánto tirás por merma con el proveedor actual? El nuestro viene con fecha garantizada.","tengo proveedor":"¿Te cobran el flete? Nosotros lo incluimos."},"upsell":"El combo panadería + lácteos tiene descuento de temporada esta semana.","kpis":["costo MP / venta","merma","rotación de carta"],"keywords":["restaurant","restaurante","café","cafe","gastronomía","gastronomia","cocina","food","bar","panadería","pizzería","sushi","catering"]},
    "tecnologia": {"tactics":["prueba gratuita 14 días","integración con herramientas existentes","soporte incluido","escalabilidad"],"objections":{"caro":"¿Cuánto cuesta el tiempo de tu equipo haciendo eso manualmente?","complejo":"Onboarding en 30 minutos, luego se maneja solo."},"upsell":"El plan Enterprise incluye API y usuarios ilimitados por solo 40% más.","kpis":["MRR","churn","NPS","tiempo hasta primer valor"],"keywords":["software","tech","app","sistema","digital","saas","código","codigo","startup","plataforma","web","ecommerce"]},
    "inmobiliaria": {"tactics":["urgencia de mercado","comparativa de zona","financiación bancaria","potencial de valorización"],"objections":{"caro":"En esta zona el m² subió 18% en 12 meses. Hoy es el precio más bajo que vas a ver.","lo pienso":"Hay 2 interesados. ¿Reservamos con señal mínima?"},"upsell":"Por $X más podés asegurar el estacionamiento que valoriza la propiedad.","kpis":["tiempo promedio de venta","precio vs tasación","comisión por operación"],"keywords":["inmobil","propiedad","alquiler","venta","m²","metros","depto","casa","lote","terreno","ph","oficina comercial"]},
    "educacion": {"tactics":["transformación medible","flexibilidad horaria","comunidad de alumnos","certificación avalada"],"objections":{"caro":"¿Cuánto vale el salario que vas a poder pedir con este título?","no tengo tiempo":"Son 3 horas semanales. Lo hacés a tu ritmo."},"upsell":"El pack con mentoría individual tiene 3x más tasa de empleo en 6 meses.","kpis":["tasa de finalización","NPS","empleabilidad post-curso"],"keywords":["escuela","academia","curso","capacitación","universidad","aprendizaje","training","colegio","instituto","online learning"]},
    "logistica": {"tactics":["precio por volumen de envíos","rastreo en tiempo real","seguro de carga incluido","entrega same-day"],"objections":{"caro":"¿Cuánto perdés en reclamos con el servicio actual? Nuestro índice de pérdida es 0.3%.","lento":"Nuestro tiempo promedio de entrega es 24h, con SLA garantizado."},"upsell":"Con el plan mensual cerrado ahorras 20% vs tarifa spot.","kpis":["costo por envío","tasa de entrega a tiempo","reclamos / total envíos"],"keywords":["logística","logistica","envío","envio","transporte","flete","courier","delivery","depósito","deposito","almacenaje","cadena de suministro"]},
    "seguros": {"tactics":["cobertura real vs precio","comparativa de pólizas","proceso de siniestro simple","atención 24/7"],"objections":{"caro":"¿Cuánto perdés si no tenés cobertura el día que más la necesitás?","ya tengo":"¿Cubrís pérdida total? ¿Con qué franquicia? Comparemos."},"upsell":"El combo auto + hogar tiene 18% de descuento vs por separado.","kpis":["siniestralidad","tasa de renovación","tiempo de respuesta en siniestro"],"keywords":["seguro","seguros","póliza","poliza","cobertura","aseguradora","siniestro","prima","vida","auto","hogar","mala praxis"]},
    "eventos": {"tactics":["cierre anticipado con descuento","paquete todo incluido","referencias de eventos previos","flexibilidad ante imprevistos"],"objections":{"caro":"El servicio fragmentado sale 30% más caro en total.","es mucho":"¿Cuántos invitados tenés? Ajustamos el paquete exacto."},"upsell":"Por $X más incluimos fotografía profesional y livestream.","kpis":["costo por invitado","NPS post-evento","tasa de repetición"],"keywords":["evento","eventos","catering","boda","casamiento","fiesta","cumpleaños","corporativo","congreso","conferencia","foto","video"]},
    "salud_bienestar": {"tactics":["resultados visibles en 30 días","programa personalizado","app de seguimiento","comunidad de apoyo"],"objections":{"caro":"¿Cuánto gastás en remedios por no tener hábitos saludables?","no funciona":"Te mostramos 3 casos similares al tuyo con resultados documentados."},"upsell":"El paquete anual incluye análisis de laboratorio trimestrales.","kpis":["adherencia al programa","NPS","retención mensual"],"keywords":["gym","gimnasio","fitness","nutrición","nutricion","dieta","bienestar","wellness","yoga","pilates","spa","meditación","psicólogo","psicologia"]},
    "automotriz": {"tactics":["test drive sin compromiso","financiación propia","cero km garantizado","plan de mantenimiento incluido"],"objections":{"caro":"¿Comparaste el costo total de propiedad a 5 años? Este modelo ahorra $X en combustible.","lo pienso":"El precio está sujeto a disponibilidad de color. Hoy tenemos el que querés."},"upsell":"El pack de seguros + service los primeros 2 años sale 40% más barato cerrado acá.","kpis":["tiempo en patio","margen por unidad","tasa de financiación interna"],"keywords":["auto","coche","vehículo","vehiculo","moto","camión","camion","taller","mecánica","mecanica","repuesto","concesionaria","0km","usado"]},
    "turismo": {"tactics":["precio dinámico","paquete todo incluido","flexibilidad de cambio","experiencias únicas"],"objections":{"caro":"¿Viste lo que incluye? Alojamiento, traslados y actividades. Por separado sale un 40% más.","lo busco solo":"¿Tenés tiempo para coordinar 6 reservas distintas? Nosotros lo hacemos."},"upsell":"Por $X más subimos a habitación con vista al mar.","kpis":["ocupación","RevPAR","tasa de cancelación","reviews"],"keywords":["turismo","viaje","hotel","vuelo","excursión","excursion","tour","hostel","airbnb","crucero","agencia de viajes","vacaciones"]},
    "moda_indumentaria": {"tactics":["tendencia de temporada","edición limitada","programa de fidelidad","talle reservado 48h"],"objections":{"caro":"Es una prenda que vas a usar 5 años vs una de fast fashion que dura 5 lavados.","lo pienso":"Solo nos queda 1 en tu talle. La segunda colorway ya se agotó."},"upsell":"El look completo con el accesorio sale $X menos que por separado.","kpis":["rotación por temporada","devoluciones","ticket promedio"],"keywords":["ropa","indumentaria","moda","fashion","vestido","remera","calzado","zapato","zapatilla","accesorio","bolso","joyería","joyeria","bijou"]},
    "mascotas": {"tactics":["salud preventiva = ahorro","club de beneficios","suscripción mensual de alimento","atención de emergencia 24h"],"objections":{"caro":"Una consulta de emergencia cuesta 10x más que la prevención anual.","no lo necesita":"Un chequeo anual detecta problemas antes de que sean graves."},"upsell":"El plan anual incluye vacunas, baño y chip de identificación.","kpis":["ticket promedio","retención anual","servicios adicionales por visita"],"keywords":["mascota","perro","gato","veterinaria","veterinario","pet","alimento balanceado","antiparasitario","grooming","guardería canina","tienda de mascotas"]},
    "legal": {"tactics":["primera consulta gratuita","tarifa fija por caso","experiencia en tu tipo de conflicto","resultados documentados"],"objections":{"caro":"¿Cuánto te cuesta perder el juicio por ir sin representación?","no sé si necesito":"15 minutos de consulta gratuita y te digo si tenés caso."},"upsell":"El abono mensual para empresas evita contingencias antes de que aparezcan.","kpis":["casos ganados / total","tiempo promedio de resolución","reclamaciones por honorarios"],"keywords":["abogado","legal","derecho","juicio","contrato","laboral","societario","penal","notaría","escribanía","patentes","marca"]},
    "contabilidad": {"tactics":["deadline fiscal protegido","digitalización de papeles","reportes mensuales claros","asesor dedicado"],"objections":{"caro":"Una multa por mora sale más cara que un año de honorarios.","lo hago yo":"¿Cuántas horas semanales? ¿Estás actualizado en los últimos cambios impositivos?"},"upsell":"El plan integral incluye liquidación de sueldos y DDJJ.","kpis":["clientes activos","rentabilidad por cliente","errores en declaraciones"],"keywords":["contador","contabilidad","impuestos","afip","iva","ganancias","balances","sueldos","monotributo","facturación","auditoría","auditoria"]},
    "limpieza": {"tactics":["contrato mensual con descuento","personal verificado","productos incluidos","servicio de emergencia 24h"],"objections":{"caro":"¿Cuánto cuesta el tiempo de tu equipo haciendo limpieza en vez de su trabajo principal?","ya tenemos":"¿Tienen seguro de responsabilidad civil? Nosotros sí."},"upsell":"El servicio profundo semestral alarga la vida de las instalaciones.","kpis":["retención de contratos","quejas por visita","costo por m²"],"keywords":["limpieza","cleaning","servicio de limpieza","mucama","higiene","desinfección","desinfeccion","mantenimiento","portería","porteria"]},
    "marketing_digital": {"tactics":["caso de éxito del mismo rubro","resultados en 30 días","reporte semanal transparente","gestión integral"],"objections":{"caro":"¿Cuánto vale un lead de calidad para tu negocio?","no sé si funciona":"Te garantizamos X cantidad de leads en el primer mes o te devolvemos el 50%."},"upsell":"Con el plan full incluimos email marketing y automatización de CRM.","kpis":["CPL","ROAS","engagement rate","conversión"],"keywords":["marketing digital","redes sociales","community manager","seo","sem","google ads","facebook ads","instagram","influencer","email marketing","contenido"]},
    "energia_solar": {"tactics":["ROI en 4 años","ahorro mensual visible desde el primer mes","subsidios disponibles","instalación en 2 días"],"objections":{"caro":"Calculamos el payback: en 48 meses recuperás la inversión, luego energía gratis.","no sé si funciona":"Tenemos monitoreo en tiempo real para que veas lo que generás cada hora."},"upsell":"Con batería de respaldo seguís con energía aunque haya corte de red.","kpis":["kWh generados / mes","ROI","tiempo de instalación"],"keywords":["solar","energía solar","energia solar","paneles","fotovoltaico","renovable","sustentable","ahorro energético","batería solar"]},
    "agro": {"tactics":["precio por volumen de cosecha","financiación en pesos o dólares","asesoramiento técnico incluido","entrega en campo"],"objections":{"caro":"El rinde extra por hectárea que genera nuestro insumo paga el costo en la primera cosecha.","ya tengo proveedor":"¿Te ofrece análisis de suelo y seguimiento de campaña? Nosotros sí."},"upsell":"Con el paquete completo fertilizante + herbicida + asesor técnico bajás el costo por qq.","kpis":["rendimiento por hectárea","costo por qq producido","días de crédito"],"keywords":["campo","agro","agrícola","agricola","semilla","fertilizante","herbicida","soja","maíz","trigo","ganadería","ganaderia","tambero","agrovet","cosecha"]},
    "salud_mental": {"tactics":["confidencialidad total","formato flexible presencial/online","plan de tratamiento claro","primeros resultados en 4 sesiones"],"objections":{"caro":"¿Cuánto cuesta en productividad y relaciones no trabajar esto?","no creo en eso":"No es sobre creer, es sobre resultados medibles. Primera sesión sin compromiso."},"upsell":"La terapia combinada individual + grupal acelera el proceso y tiene descuento.","kpis":["adherencia al tratamiento","NPS","altas por mejoría"],"keywords":["psicólogo","psicologo","psicología","psicologia","terapeuta","terapia","salud mental","ansiedad","depresión","depresion","bienestar emocional","coaching","mindfulness"]},
    "comercio_electronico": {"tactics":["envío gratis por umbral de compra","garantía de devolución 30 días","checkout en 1 click","descuento por primera compra"],"objections":{"caro":"¿Incluye el envío? Con nosotros el total es $X vs $X+envío.","no confío en comprar online":"Nuestro proceso de devolución es gratuito y en 48h. Sin preguntas."},"upsell":"Comprando el pack de 3 el costo unitario baja 22%.","kpis":["tasa de conversión","abandono de carrito","LTV","ROAS"],"keywords":["ecommerce","e-commerce","tienda online","shopify","woocommerce","mercadolibre","carrito","checkout","envío gratis","marketplace"]},
    "rrhh_consultora": {"tactics":["reducción de tiempo de búsqueda","base de candidatos pre-filtrada","garantía de reemplazo 90 días","reportes de proceso"],"objections":{"caro":"¿Cuánto cuesta tener un puesto vacante 3 meses? Calculemos.","lo hacemos internamente":"¿Tienen acceso a candidatos pasivos? Nosotros sí — el 60% de nuestros perfiles no busca activamente."},"upsell":"El servicio de onboarding asistido reduce la rotación temprana en un 40%.","kpis":["time-to-hire","costo por contratación","retención a 12 meses"],"keywords":["rrhh","recursos humanos","headhunting","búsqueda laboral","talento","selección","contratación","empresa de personal","staff","outsourcing laboral"]},
    "financiero": {"tactics":["tasa diferencial por volumen","proceso 100% digital","aprobación en 24h","sin penalidad por pago anticipado"],"objections":{"caro":"Comparado con el banco, nuestra tasa efectiva es X% menor con aprobación 10x más rápida.","no califico":"Tenemos productos para todo perfil. ¿Cuánto necesitás y en cuánto tiempo?"},"upsell":"El seguro de cuota protege tus cuotas ante imprevistos sin costo adicional el primer año.","kpis":["morosidad","ticket promedio","tasa de aprobación","costo de fondeo"],"keywords":["préstamo","prestamo","crédito","credito","banco","fintech","inversión","inversion","ahorro","billetera","transferencia","cuota","financiamiento","microcrédito"]},
    "spa_belleza": {"tactics":["primer servicio con descuento","pack de sesiones con ahorro","membresía mensual","agenda online"],"objections":{"caro":"Por lo que cuesta un café diario tenés tu cuidado mensual.","no tengo tiempo":"Sesiones de 30 min, sin turno previo los miércoles."},"upsell":"El tratamiento facial + masaje completo tiene 25% de descuento en pack.","kpis":["ticket promedio","retención mensual","servicios adicionales por visita"],"keywords":["spa","belleza","peluquería","peluqueria","estética","estetica","manicura","pedicura","masaje","depilación","depilacion","tratamiento facial","uñas","cabello"]},
    "deportivo": {"tactics":["prueba gratuita","flexibilidad de horario","comunidad activa","progreso visible en 30 días"],"objections":{"caro":"¿Cuánto gastás en medicación o días de baja por no hacer actividad?","no tengo tiempo":"Tenemos clases de 45 min de lunes a sábado en 6 horarios distintos."},"upsell":"El plan anual sale equivalente a 10 meses y te deja el 20% adelantado para equipamiento.","kpis":["retención mensual","nuevos socios / mes","NPS","ocupación por clase"],"keywords":["club","deporte","fútbol","futbol","tenis","natación","natacion","paddle","crossfit","box","atletismo","básquet","basquet","rugby","entrenamiento"]},
    "ferreteria": {"tactics":["descuento por volumen de compra","demostraciones de productos en el local","asesoramiento tecnico sobre el uso correcto","combos de herramienta + accesorios","fidelizacion con cuenta corriente para clientes frecuentes","financiacion en cuotas para compras grandes","reserva de stock para obras en curso","atencion prioritaria a maestros mayores de obra"],"objections":{"caro":"Comparalo con la marca que usabas antes. Esta dura el doble y la garantia te cubre.","lo consigo mas barato":"Puede ser, pero ahi te dan factura, asesoramiento y garantia? Aca si.","no tengo efectivo":"Tenemos cuenta corriente para clientes frecuentes y tarjeta hasta en 6 cuotas.","lo pienso":"Te lo separo hasta manana sin compromiso. Asi no te quedas sin stock.","necesito factura":"Por supuesto, facturamos todo. Solo paso tus datos.","me fias":"Para clientes nuevos no, pero arrancamos cuenta corriente despues de la segunda compra.","viene con garantia":"Si, todo lo que vendemos tiene garantia de fabrica. Te doy el comprobante.","no se si me sirve":"Contame para que lo necesitas y te armo lo justo, sin que gastes de mas."},"upsell":"Si llevas la herramienta, te conviene el set completo con los accesorios — sale mas barato junto.","kpis":["ticket promedio","rotacion de stock","clientes con cuenta corriente activa","margen por categoria"],"keywords":["ferreteria","ferretera","corralon","tornillo","tornillos","clavo","clavos","pintura","herramienta","herramientas","cemento","hierro","caño","caños","electricidad","plomeria","sanitario","grifería","cerradura","candado","taladro","amoladora","soldadora","construccion menor","bulonería"]},
    "default": {"tactics":["valor sobre precio","urgencia real","caso de éxito","próximo paso concreto"],"objections":{"caro":"¿Qué comparás exactamente? Veamos qué incluye cada opción.","lo pienso":"¿Qué información te falta para decidir hoy?"},"upsell":"¿Qué otro problema querés resolver mientras estamos?","kpis":["conversión","ticket promedio","retención"],"keywords":[]},
}

def _load_db_industries(workspace_id: int) -> dict:
    """Load custom industries from DB, overriding static playbook."""
    try:
        with closing(get_db()) as conn:
            rows = conn.execute(
                "SELECT slug, data_json FROM industry_playbooks WHERE workspace_id IN (0, ?) ORDER BY workspace_id ASC",
                (workspace_id,),
            ).fetchall()
            custom = {}
            for r in rows:
                try:
                    custom[r["slug"]] = json.loads(r["data_json"])
                except Exception:
                    pass
            return custom
    except Exception:
        return {}


def get_playbooks(workspace_id: int) -> dict:
    """Merge static playbooks with DB overrides. DB wins."""
    merged = dict(INDUSTRY_PLAYBOOKS)
    merged.update(_load_db_industries(workspace_id))
    return merged


def detect_industry(text: str, products: list, knowledge: list, playbooks: dict | None = None) -> str:
    """Detect industry from context clues."""
    combined = (text + " ".join(p.get("category","") for p in products) + " ".join(a.get("title","") for a in knowledge)).lower()
    pb = playbooks or INDUSTRY_PLAYBOOKS
    for slug, data in pb.items():
        if slug == "default":
            continue
        keywords = data.get("keywords", [])
        if any(k.lower() in combined for k in keywords):
            return slug
    return "default"


def generate_ai_reply(workspace_id: int, text: str, language: str, currency: str) -> dict[str, Any]:
    status = integration_status()
    fallback = natural_reply(workspace_id, text, language, currency)
    if not status["openai"]["configured"]:
        return {"reply": fallback, "provider": "local_fallback", "used_openai": False}

    try:
        with closing(get_db()) as conn:
            products = [dict(r) for r in conn.execute(
                "SELECT sku, name, price, stock, category, cost FROM products WHERE workspace_id = ? ORDER BY demand_score DESC",
                (workspace_id,),
            ).fetchall()]
            knowledge = [dict(r) for r in conn.execute(
                "SELECT title, content FROM knowledge_articles WHERE workspace_id = ?",
                (workspace_id,),
            ).fetchall()]
            sources = fetch_workspace_sources(conn, workspace_id)
            profile = get_workspace_profile(conn, workspace_id)
            # Recent deals for context
            recent_deals = [dict(r) for r in conn.execute(
                "SELECT customer_name, negotiated_total, discount_pct, status FROM deals WHERE workspace_id = ? ORDER BY id DESC LIMIT 5",
                (workspace_id,),
            ).fetchall()]

        playbooks = get_playbooks(workspace_id)
        industry = detect_industry(text, products, knowledge, playbooks)
        playbook = playbooks.get(industry, playbooks.get("default", INDUSTRY_PLAYBOOKS["default"]))
        client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
        model = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
        source_snippets = source_snippets_for_text(text, sources)

        lang_instruction = {
            "es": "Responde siempre en español. Usa tuteo natural y tono comercial directo.",
            "pt": "Responda sempre em português brasileiro. Tom comercial direto.",
            "en": "Always reply in English. Direct, commercially sharp tone.",
        }.get(language, f"Always reply in language: {language}.")

        if products:
            products_ctx = json.dumps([{"name":p["name"],"price":p["price"],"stock":p["stock"],"sku":p["sku"]} for p in products[:15]], ensure_ascii=False)
        else:
            products_ctx = "NINGUN PRODUCTO CARGADO TODAVIA EN EL INVENTARIO - no inventes productos, precios ni stock. Decile al cliente que confirmás disponibilidad y precio, y pedile los datos del pedido para cotizar."
        recent_ctx   = json.dumps(recent_deals[:3], ensure_ascii=False)
        knowledge_ctx = json.dumps([a["title"] for a in knowledge[:10]], ensure_ascii=False)
        tactics_str   = ", ".join(playbook["tactics"])
        kpis_str      = ", ".join(playbook["kpis"])
        objections_str = json.dumps(playbook["objections"], ensure_ascii=False, indent=2)

        system_prompt = (
            "Sos el Agente de Ventas Genio de Banzai — el cerebro comercial más efectivo jamás construido. "
            "Combinás psicología de ventas, conocimiento profundo de negocios, y experiencia en más de 30 industrias.\n\n"

            "IDENTIDAD Y ESTILO\n"
            f"- Estilo de respuesta: {profile.get('response_style', 'Cálido y Profesional')}\n"
            f"- Personalidad: {profile.get('personality_notes', 'Agudo, humano, comercialmente consciente')}\n"
            f"- Nunca sonar como: {profile.get('forbidden_tone', 'robótico, sobre-guionado')}\n"
            f"- {lang_instruction}\n\n"

            f"INDUSTRIA DETECTADA: {industry.upper()}\n"
            f"Tácticas de esta industria: {tactics_str}\n"
            f"KPIs que le importan: {kpis_str}\n"
            f"Ángulo de upsell: {playbook['upsell']}\n\n"

            "MANEJO DE OBJECIONES (usá naturalmente, no textual):\n"
            f"{objections_str}\n\n"

            "PRINCIPIOS DE VENTAS GENIUS (siempre activos):\n"
            "1. Cada respuesta debe avanzar la venta — siempre terminá con un próximo paso concreto\n"
            "2. Liderá con valor, no con precio — anclá en lo que GANAN, no en lo que cuesta\n"
            "3. Si objetan el precio → reformulá en ROI, nunca en descuento\n"
            "4. Urgencia real — stock limitado, fechas, vigencia de oferta — nunca urgencia falsa\n"
            "5. Una pregunta calificadora por respuesta para entender la necesidad real\n"
            "6. Espejá la energía del cliente — si es breve, sé breve; si pide detalle, dale detalle\n"
            "7. Nunca digas 'no puedo' — decí 'esto es lo que SÍ puedo hacer'\n"
            "8. Después de cada oferta → cerrá con pregunta confirmadora\n"
            "9. Prueba social cuando esté disponible ('La mayoría de clientes en tu situación eligen X porque...')\n"
            "10. Sabés cuándo escalar — deal > 3x el promedio → marcá para revisión humana\n\n"

            "CONOCIMIENTO EXPERTO DE NEGOCIOS (aplicá según contexto):\n"
            "- Finanzas: flujo de caja, márgenes, punto de equilibrio, ROI, payback period\n"
            "- Ventas: SPIN selling, challenger sale, value-based selling, BANT, MEDDIC\n"
            "- Marketing: funnel, CAC, LTV, churn, NPS, cohort analysis\n"
            "- Operaciones: lean, just-in-time, KPIs operativos, automatización de procesos\n"
            "- Legal/Fiscal: IVA, AFIP, facturación electrónica, contratos, compliance\n"
            "- RRHH: estructura de incentivos, comisiones, onboarding, retención de talento\n"
            "- Tecnología: SaaS, APIs, integraciones, escalabilidad, seguridad de datos\n"
            "- Estrategia: Porter, FODA, océano azul, disruption, product-market fit\n\n"

            "CONTEXTO DEL NEGOCIO:\n"
            f"- Productos disponibles (ESTOS SON LOS UNICOS PRODUCTOS REALES QUE EXISTEN): {products_ctx}\n"
            f"- Deals recientes cerrados: {recent_ctx}\n"
            f"- Base de conocimiento: {knowledge_ctx}\n"
            f"- Moneda: {currency}\n\n"

            "REGLAS ESTRICTAS - NUNCA VIOLAR:\n"
            "1. NUNCA inventes productos que no estén en la lista de 'Productos disponibles' de arriba. Si el cliente pide algo que no está en la lista, decile que consultás disponibilidad o que no lo tenés — nunca inventes que sí lo tenés.\n"
            "2. NUNCA ofrezcas descuentos, promociones, flete gratis, cuotas o beneficios que el cliente no mencionó primero y que no estén explícitamente en las tácticas de la industria de arriba. No regales nada por iniciativa propia.\n"
            "3. NUNCA inventes precios. Si un producto no tiene precio en la lista de arriba, decile al cliente que confirmás el precio y no des un número inventado.\n"
            "4. NUNCA confirmes stock que no está en la lista de arriba. Si no tenés el dato de stock, decile que confirmás disponibilidad.\n"
            "5. Si el cliente pregunta algo que no podés responder con la información de arriba, sé honesto: decile que confirmás esa información y volvés con la respuesta — no inventes.\n"
            "6. Solo podés ofrecer el ángulo de upsell configurado arriba, y solo si tiene sentido con lo que el cliente ya pidió — nunca antes de que el cliente muestre intención de compra.\n\n"

            "Respondé en menos de 150 palabras salvo que la situación requiera detalle. "
            "Sé humano, directo y comercialmente efectivo, pero SIEMPRE basado en datos reales del negocio, nunca en invenciones."
        )

        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": text},
            ],
            max_tokens=500,
            temperature=0.75,
        )
        reply_text = response.choices[0].message.content or fallback
        return {"reply": reply_text.strip(), "provider": "openai", "used_openai": True, "industry": industry}
    except Exception as exc:
        return {"reply": fallback, "provider": f"openai_error:{type(exc).__name__}", "used_openai": False}




# ── Vendor Invoice Email Ingestion ─────────────────────────────────────────

def _extract_invoice_data_with_ai(text_content: str, filename: str) -> dict:
    """Use OpenAI to extract structured invoice data from raw text (PDF or email body)."""
    api_key = os.environ.get("OPENAI_API_KEY", "")
    if not api_key or OpenAI is None:
        return {"vendor_name": None, "invoice_number": None, "invoice_date": None,
                "amount": None, "currency": "USD", "is_paid": False, "confidence": 0.0}
    try:
        client = OpenAI(api_key=api_key)
        prompt = (
            "Extrae los datos de esta factura de proveedor. Devolve SOLO un JSON valido con estas claves exactas:\n"
            '{"vendor_name": string o null, "invoice_number": string o null, "invoice_date": "YYYY-MM-DD" o null, '
            '"due_date": "YYYY-MM-DD" o null, "amount": number o null, "currency": "USD"/"ARS"/etc, '
            '"is_paid": true/false (true solo si el texto dice explicitamente pagada/paid/abonada), "confidence": number entre 0 y 1}\n\n'
            f"Archivo: {filename}\n\nContenido:\n{text_content[:6000]}"
        )
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "Sos un extractor de datos de facturas. Respondes solo JSON valido, sin markdown ni texto adicional."},
                {"role": "user", "content": prompt},
            ],
            max_tokens=400,
            temperature=0.0,
        )
        raw = resp.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = "\n".join(raw.split("\n")[1:-1])
        data = json.loads(raw)
        return data
    except Exception as e:
        return {"vendor_name": None, "invoice_number": None, "invoice_date": None,
                "amount": None, "currency": "USD", "is_paid": False, "confidence": 0.0, "error": str(e)}


def _extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    """Extract raw text from PDF bytes using pypdf if available."""
    try:
        import io as _io
        try:
            from pypdf import PdfReader
        except ImportError:
            from PyPDF2 import PdfReader
        reader = PdfReader(_io.BytesIO(pdf_bytes))
        text = ""
        for page in reader.pages[:5]:
            text += page.extract_text() or ""
        return text
    except Exception as e:
        return f"[Could not extract PDF text: {e}]"


def check_vendor_invoice_inbox(workspace_id: int) -> dict:
    """Connect to the configured mailbox via IMAP, find unread invoice attachments, extract and save them."""
    import imaplib
    import email as email_lib
    from email.header import decode_header

    with closing(get_db()) as conn:
        cfg = conn.execute(
            "SELECT * FROM email_inbox_config WHERE workspace_id = ? AND active = 1",
            (workspace_id,)
        ).fetchone()

    if not cfg:
        return {"ok": False, "error": "No hay una casilla de mail configurada para este negocio"}

    cfg = dict(cfg)
    processed = []
    errors = []

    try:
        imap = imaplib.IMAP4_SSL(cfg["imap_host"], cfg.get("imap_port") or 993)
        imap.login(cfg["email_address"], cfg["email_password"])
        imap.select("INBOX")

        status, message_ids = imap.search(None, "UNSEEN")
        if status != "OK":
            imap.logout()
            return {"ok": False, "error": "No se pudo buscar mensajes"}

        ids = message_ids[0].split()
        ids = list(reversed(ids))  # process most recent unread messages first
        now_iso = datetime.utcnow().isoformat()

        for msg_id in ids[:20]:
            try:
                status, msg_data = imap.fetch(msg_id, "(RFC822)")
                if status != "OK":
                    continue
                raw_email = msg_data[0][1]
                msg = email_lib.message_from_bytes(raw_email)

                subject_raw = decode_header(msg.get("Subject", ""))[0]
                subject = subject_raw[0]
                if isinstance(subject, bytes):
                    subject = subject.decode(subject_raw[1] or "utf-8", errors="ignore")

                from_addr = msg.get("From", "")

                found_attachment = False
                for part in msg.walk():
                    content_disposition = str(part.get("Content-Disposition", ""))
                    content_type = part.get_content_type()
                    filename = part.get_filename()

                    if filename and ("attachment" in content_disposition or content_type == "application/pdf"):
                        found_attachment = True
                        file_bytes = part.get_payload(decode=True)

                        if content_type == "application/pdf":
                            text_content = _extract_text_from_pdf_bytes(file_bytes)
                        else:
                            text_content = f"[Archivo adjunto: {filename}, tipo: {content_type}]"

                        file_b64 = None
                        try:
                            import base64 as _b64mod
                            if file_bytes and len(file_bytes) < 8_000_000:  # keep DB row size reasonable, max ~8MB
                                file_b64 = _b64mod.b64encode(file_bytes).decode("ascii")
                        except Exception:
                            file_b64 = None

                        extracted = _extract_invoice_data_with_ai(text_content, filename)

                        with closing(get_db()) as conn:
                            is_paid = extracted.get("is_paid", False)
                            amount = extracted.get("amount")
                            confidence = extracted.get("confidence", 0.0)
                            needs_review = 1 if (confidence < 0.6 or not amount) else 0

                            # Duplicate detection: same vendor + same invoice number already exists
                            inv_num = extracted.get("invoice_number")
                            vendor_nm = extracted.get("vendor_name")
                            if inv_num and vendor_nm:
                                dup = conn.execute(
                                    "SELECT id FROM vendor_invoices WHERE workspace_id = ? AND invoice_number = ? AND vendor_name = ?",
                                    (workspace_id, inv_num, vendor_nm)
                                ).fetchone()
                                if dup:
                                    processed.append({
                                        "invoice_id": dup["id"], "vendor": vendor_nm,
                                        "amount": amount, "is_paid": is_paid,
                                        "duplicate": True, "skipped": True,
                                    })
                                    continue

                            cur = conn.execute(
                                """INSERT INTO vendor_invoices
                                   (workspace_id, vendor_name, invoice_number, invoice_date, due_date,
                                    amount, currency, status, source_email, source_subject,
                                    raw_text, file_name, file_base64, file_mime_type,
                                    extracted_confidence, needs_review, processed_at, paid_at)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                                (workspace_id, extracted.get("vendor_name"), extracted.get("invoice_number"),
                                 extracted.get("invoice_date"), extracted.get("due_date"),
                                 amount, extracted.get("currency", "USD"),
                                 "paid" if is_paid else "pending",
                                 from_addr, subject, text_content[:3000], filename,
                                 file_b64, content_type,
                                 confidence, needs_review, now_iso,
                                 now_iso if is_paid else None)
                            )
                            invoice_id = cur.lastrowid

                            if is_paid and amount:
                                ledger_cur = conn.execute(
                                    """INSERT INTO ledger_entries
                                       (workspace_id, entry_type, concept, category, amount, currency, state, due_date, source, created_at)
                                       VALUES (?,?,?,?,?,?,?,?,'auto',?)""",
                                    (workspace_id, "Expense",
                                     f"Factura proveedor {extracted.get('vendor_name','')} — {extracted.get('invoice_number','')}",
                                     "Proveedores", amount, extracted.get("currency", "USD"),
                                     "Paid", now_iso[:10], now_iso)
                                )
                                conn.execute("UPDATE vendor_invoices SET ledger_entry_id = ? WHERE id = ?",
                                             (ledger_cur.lastrowid, invoice_id))

                            conn.commit()

                        processed.append({
                            "invoice_id": invoice_id, "vendor": extracted.get("vendor_name"),
                            "amount": amount, "is_paid": is_paid, "needs_review": bool(needs_review),
                        })

                if not found_attachment:
                    imap.store(msg_id, "-FLAGS", "\\Seen")

            except Exception as msg_err:
                errors.append(str(msg_err))
                continue

        imap.logout()

        with closing(get_db()) as conn:
            conn.execute("UPDATE email_inbox_config SET last_checked_at = ? WHERE workspace_id = ?",
                         (now_iso, workspace_id))
            conn.commit()

        return {"ok": True, "processed": processed, "errors": errors, "count": len(processed)}

    except Exception as e:
        return {"ok": False, "error": str(e)}


def send_whatsapp_text(to: str, text: str, workspace_id: int | None = None) -> dict[str, Any]:
    """Send WhatsApp message via Twilio. Falls back to Meta direct API if Twilio not configured.
    If workspace_id is given, sends FROM that business's own registered number so each
    business's customers see messages coming from the number they actually wrote to —
    critical for multi-tenant, otherwise every business would share one sender identity."""
    twilio_sid = os.environ.get("TWILIO_ACCOUNT_SID", "")
    twilio_token = os.environ.get("TWILIO_AUTH_TOKEN", "")
    twilio_from = os.environ.get("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886")
    if workspace_id:
        try:
            with closing(get_db()) as _conn_wa:
                _ws_row = _conn_wa.execute("SELECT whatsapp_number FROM workspaces WHERE id=?", (workspace_id,)).fetchone()
                if _ws_row and _ws_row["whatsapp_number"]:
                    twilio_from = f"whatsapp:{_ws_row['whatsapp_number']}"
        except Exception:
            pass
    if twilio_sid and twilio_token:
        to_fmt = to if to.startswith("whatsapp:") else f"whatsapp:{to}"
        url = f"https://api.twilio.com/2010-04-01/Accounts/{twilio_sid}/Messages.json"
        resp = requests.post(
            url,
            data={"To": to_fmt, "From": twilio_from, "Body": text},
            auth=(twilio_sid, twilio_token),
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()
    status = integration_status()
    if not status["whatsapp"]["configured"]:
        raise RuntimeError("WhatsApp integration is not configured")
    url = f"https://graph.facebook.com/v23.0/{os.environ['WHATSAPP_PHONE_NUMBER_ID']}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": to,
        "type": "text",
        "text": {"body": text},
    }
    headers = {
        "Authorization": f"Bearer {os.environ['WHATSAPP_ACCESS_TOKEN']}",
        "Content-Type": "application/json",
    }
    resp = requests.post(url, json=payload, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.json()


def get_workspace_profile(conn: sqlite3.Connection, workspace_id: int) -> dict[str, Any]:
    row = conn.execute(
        "SELECT response_style, personality_notes, forbidden_tone FROM workspace_profiles WHERE workspace_id = ?",
        (workspace_id,),
    ).fetchone()
    return dict(row) if row else {
        "response_style": "Warm Professional",
        "personality_notes": "Sound natural, clear and calm.",
        "forbidden_tone": "Avoid robotic wording.",
    }


def natural_reply(workspace_id: int, text: str, language: str, currency: str) -> str:
    text_l = text.lower()
    with closing(get_db()) as conn:
        products = [dict(r) for r in conn.execute(
            "SELECT sku, name, price, stock FROM products WHERE workspace_id = ? ORDER BY demand_score DESC",
            (workspace_id,),
        ).fetchall()]
        knowledge = [dict(r) for r in conn.execute(
            "SELECT title, content FROM knowledge_articles WHERE workspace_id = ?",
            (workspace_id,),
        ).fetchall()]
        sources = fetch_workspace_sources(conn, workspace_id)
        profile = get_workspace_profile(conn, workspace_id)

    matches = [p for p in products if p["sku"].lower() in text_l or p["name"].split()[0].lower() in text_l]
    delivery = next((k["content"] for k in knowledge if "delivery" in k["title"].lower() or "entrega" in k["title"].lower()), "")
    source_snippets = source_snippets_for_text(text, sources)

    openers = {
        "Warm Professional": {
            "en": [
                "Absolutely — I can help with that.",
                "Of course. Let me sort that out for you.",
                "Sure, I can help you move this forward.",
            ],
            "es": [
                "Claro, te ayudo con eso.",
                "Sí, lo vemos ahora mismo.",
                "Perfecto, te lo resuelvo enseguida.",
            ],
            "pt": [
                "Claro, eu te ajudo com isso.",
                "Perfeito, vamos resolver isso agora.",
                "Sim, consigo te ajudar com isso.",
            ],
        },
        "Premium Concierge": {
            "en": [
                "Of course — happy to help.",
                "Absolutely. I’ll make this as smooth as possible for you.",
                "Certainly. Let’s get this handled properly.",
            ],
            "es": [
                "Por supuesto, con gusto te ayudo.",
                "Claro, lo resolvemos con mucho gusto.",
                "Sí, encantado de ayudarte con esto.",
            ],
            "pt": [
                "Claro, será um prazer te ajudar.",
                "Com certeza, vou te ajudar com isso.",
                "Perfeito, cuidamos disso agora.",
            ],
        },
        "Direct Sales": {
            "en": [
                "Yes — we can do that.",
                "Absolutely. Here’s the quickest way to move forward.",
                "Sure. Let’s keep this simple.",
            ],
            "es": [
                "Sí, lo podemos hacer.",
                "Perfecto, vamos directo a eso.",
                "Claro, te lo dejo resuelto rápido.",
            ],
            "pt": [
                "Sim, conseguimos fazer isso.",
                "Perfeito, vamos direto ao ponto.",
                "Claro, resolvemos isso rapidamente.",
            ],
        },
    }

    style = profile.get("response_style", "Warm Professional")
    opener_bank = openers.get(style, openers["Warm Professional"])
    opener = random.choice(opener_bank.get(language, opener_bank["en"]))

    if matches:
        lines = [f"• {item['name']} = {format_money(item['price'], currency)} (stock: {item['stock']})" for item in matches[:3]]
        if language == "es":
            closer = "Si querés, te preparo el presupuesto ahora mismo y te confirmo el siguiente paso."
            return f"{opener}\n\nEncontré esto para vos:\n" + "\n".join(lines) + f"\n\n{closer}"
        if language == "pt":
            closer = "Se quiser, eu preparo o orçamento agora mesmo e já deixo o próximo passo bem claro."
            return f"{opener}\n\nEncontrei isto:\n" + "\n".join(lines) + f"\n\n{closer}"
        closer = "If you want, I can put the quote together now and leave everything clear so you can decide without wasting time."
        return f"{opener}\n\nHere’s what I found:\n" + "\n".join(lines) + f"\n\n{closer}"

    if any(word in text_l for word in ["delivery", "entrega", "envio", "today", "hoy", "hoje"]):
        extra = delivery or {
            "es": "Podemos revisar la entrega según la zona y el horario de confirmación.",
            "pt": "Podemos revisar a entrega conforme a região e ao horário de confirmação.",
            "en": "We can check delivery based on your area and confirmation time.",
        }.get(language, "We can check delivery based on your area and confirmation time.")
        tail = {
            "es": "Si querés, también te dejo el pedido listo para no perder tiempo.",
            "pt": "Se quiser, já deixo o pedido encaminhado também.",
            "en": "If you want, I can also leave the order ready so everything moves faster.",
        }
        return f"{opener} {extra} {tail.get(language, tail['en'])}"

    if any(word in text_l for word in ["meeting", "reunion", "reunião", "schedule", "agendar"]):
        options = {
            "es": "Si te parece, te propongo el próximo horario disponible y lo dejamos confirmado.",
            "pt": "Se quiser, eu já proponho o próximo horário disponível e deixo encaminhado.",
            "en": "If you’d like, I can suggest the next available slot and get it lined up for you.",
        }
        return f"{opener} {options.get(language, options['en'])}"

    expert_tail = {
        "es": " Además, tengo cargada una biblioteca experta para responder con criterio profesional en ventas, administración, finanzas y precios." if source_snippets else "",
        "pt": " Além disso, tenho uma biblioteca especialista ativa para responder com critério profissional em vendas, administração, finanças e preços." if source_snippets else "",
        "en": " I also have an expert library loaded in this workspace to answer with stronger professional judgment across sales, admin, finance and pricing." if source_snippets else "",
    }

    generic = {
        "es": "Puedo ayudarte con precios, presupuesto, seguimiento, tareas administrativas o el próximo paso que necesites.",
        "pt": "Posso te ajudar com preços, orçamento, acompanhamento, tarefas administrativas ou o próximo passo que você precisar.",
        "en": "I can help with pricing, quotes, follow-up, admin work or simply sort out the next step with you in a clear way.",
    }
    return f"{opener} {generic.get(language, generic['en'])}{expert_tail.get(language, expert_tail['en'])}"


# ══════════════════════════════════════════════════════════════════════════════
# MULTI-AGENT SYSTEM — Sales Agent · Auditor Agent · Accounting Agent
# Orchestrator coordinates them; each logs to agent_events for full audit trail
# ══════════════════════════════════════════════════════════════════════════════

def _log_agent_event(
    conn: sqlite3.Connection,
    workspace_id: int,
    agent: str,
    action: str,
    entity_type: str | None = None,
    entity_id: int | None = None,
    input_summary: str = "",
    output_summary: str = "",
    confidence: float = 1.0,
    requires_human_review: bool = False,
) -> None:
    now = datetime.utcnow().isoformat()
    conn.execute(
        """INSERT INTO agent_events
           (workspace_id, agent, action, entity_type, entity_id,
            input_summary, output_summary, confidence, requires_human_review, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (workspace_id, agent, action, entity_type, entity_id,
         input_summary[:500], output_summary[:500], confidence,
         1 if requires_human_review else 0, now),
    )


def get_negotiation_margins(conn: sqlite3.Connection, workspace_id: int, product_id: int | None = None) -> dict:
    row = conn.execute(
        "SELECT * FROM negotiation_margins WHERE workspace_id = ? AND (product_id = ? OR product_id IS NULL) ORDER BY product_id DESC LIMIT 1",
        (workspace_id, product_id),
    ).fetchone()
    if row:
        return dict(row)
    return {
        "min_margin_pct": 0.25,
        "max_discount_pct": 0.15,
        "auto_approve_below_pct": 0.05,
    }


def get_tax_profile(conn: sqlite3.Connection, workspace_id: int) -> dict:
    row = conn.execute("SELECT * FROM tax_profiles WHERE workspace_id = ?", (workspace_id,)).fetchone()
    if row:
        return dict(row)
    # Derive defaults from workspace
    ws = conn.execute("SELECT region, currency FROM workspaces WHERE id = ?", (workspace_id,)).fetchone()
    if ws:
        region = ws["region"]
        TAX_DEFAULTS = {
            "AR": {"tax_pct": 21.0, "authority": "AFIP", "regime": "AFIP-FCE", "prefix": "FC"},
            "US": {"tax_pct": 0.0,  "authority": "IRS",  "regime": "US-standard", "prefix": "INV"},
            "BR": {"tax_pct": 12.0, "authority": "Receita Federal", "regime": "NF-e", "prefix": "NF"},
        }
        d = TAX_DEFAULTS.get(region, TAX_DEFAULTS["US"])
        return {
            "country_code": region,
            "tax_authority": d["authority"],
            "tax_id": None,
            "default_tax_pct": d["tax_pct"],
            "fiscal_regime": d["regime"],
            "invoice_prefix": d["prefix"],
            "next_invoice_seq": 1,
        }
    return {"country_code": "US", "tax_authority": "IRS", "tax_id": None,
            "default_tax_pct": 0.0, "fiscal_regime": "US-standard",
            "invoice_prefix": "INV", "next_invoice_seq": 1}


def _next_invoice_number(conn: sqlite3.Connection, workspace_id: int) -> str:
    tax = get_tax_profile(conn, workspace_id)
    prefix = tax.get("invoice_prefix", "INV")
    seq = tax.get("next_invoice_seq", 1)
    conn.execute(
        """INSERT INTO tax_profiles (workspace_id, country_code, tax_authority, tax_id,
           default_tax_pct, fiscal_regime, invoice_prefix, next_invoice_seq, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(workspace_id) DO UPDATE SET next_invoice_seq = next_invoice_seq + 1, updated_at = excluded.updated_at""",
        (workspace_id, tax.get("country_code","US"), tax.get("tax_authority",""),
         tax.get("tax_id",""), tax.get("default_tax_pct",0), tax.get("fiscal_regime",""),
         prefix, seq + 1, datetime.utcnow().isoformat()),
    )
    return f"{prefix}-{seq:05d}"


# ── Agent 1: Sales Agent ──────────────────────────────────────────────────────
def sales_agent_negotiate(
    workspace_id: int,
    conversation_id: int | None,
    customer_name: str,
    customer_message: str,
    proposed_items: list[dict],
    language: str,
    currency: str,
) -> dict[str, Any]:
    """
    Analyzes the customer message, checks approved margin bands,
    decides whether to auto-approve a discount or escalate to human,
    and returns a counter-offer or acceptance.
    """
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        margins = get_negotiation_margins(conn, workspace_id)
        products_map = {
            str(r["id"]): dict(r)
            for r in conn.execute("SELECT * FROM products WHERE workspace_id = ?", (workspace_id,)).fetchall()
        }

        # Compute original total and minimum floor
        original_total = 0.0
        floor_total = 0.0
        enriched_items = []
        for item in proposed_items:
            pid = str(item.get("product_id", ""))
            product = products_map.get(pid, {})
            cost = float(product.get("cost", item.get("unit_price", 0)))
            list_price = float(product.get("price", item.get("unit_price", 0)))
            qty = float(item.get("qty", 1))
            min_margin = margins["min_margin_pct"]
            floor_price = cost / max(0.001, (1 - min_margin))
            enriched_items.append({
                **item,
                "list_price": list_price,
                "cost": cost,
                "floor_price": round(floor_price, 2),
                "qty": qty,
            })
            original_total += list_price * qty
            floor_total += floor_price * qty

        # Detect if customer is asking for a discount
        discount_signals = ["descuento", "precio especial", "discount", "cheaper", "better price",
                            "desconto", "melhor preço", "rebate", "deal", "negociar"]
        msg_lower = customer_message.lower()
        wants_discount = any(s in msg_lower for s in discount_signals)

        # Determine counter-offer
        if not wants_discount:
            # No discount requested → offer full price
            negotiated_total = original_total
            discount_pct = 0.0
            decision = "offer_sent"
            agent_note = "Full price offered — no discount signal detected."
            confidence = 0.95
            requires_review = False
        else:
            max_discount = margins["max_discount_pct"]
            auto_floor = margins["auto_approve_below_pct"]
            # Start with a mid-point offer: half the max discount
            proposed_discount = max_discount * 0.5
            negotiated_total = max(floor_total, original_total * (1 - proposed_discount))
            actual_discount = (original_total - negotiated_total) / max(0.001, original_total)
            discount_pct = round(actual_discount * 100, 2)

            if actual_discount <= auto_floor:
                decision = "offer_sent"
                agent_note = f"Auto-approved {discount_pct:.1f}% discount (below auto-approve threshold of {auto_floor*100:.0f}%)."
                confidence = 0.90
                requires_review = False
            elif actual_discount <= max_discount:
                decision = "offer_sent"
                agent_note = f"Offered {discount_pct:.1f}% discount (within approved band of {max_discount*100:.0f}%)."
                confidence = 0.80
                requires_review = False
            else:
                decision = "escalated"
                agent_note = f"Requested discount exceeds max band ({max_discount*100:.0f}%). Escalated to human review."
                confidence = 0.50
                requires_review = True

        # Persist the deal
        expires_at = datetime.utcnow().replace(hour=23, minute=59).isoformat()
        with closing(get_db()) as conn2:
            cur = conn2.execute(
                """INSERT INTO deals
                   (workspace_id, conversation_id, customer_name, channel, status,
                    items_json, original_total, negotiated_total, discount_pct,
                    margin_floor, currency, expires_at, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (workspace_id, conversation_id, customer_name, "webchat",
                 decision, json.dumps(enriched_items),
                 round(original_total, 2), round(negotiated_total, 2),
                 discount_pct, round(floor_total, 2), currency, expires_at, now),
            )
            deal_id = cur.lastrowid

            # Auto-create a Pipeline card so the Sales Agent's work shows up on the visual board
            # without needing anyone to add it manually.
            try:
                pipeline_row = conn2.execute(
                    "SELECT id FROM pipelines WHERE workspace_id=? ORDER BY id LIMIT 1", (workspace_id,)
                ).fetchone()
                if not pipeline_row:
                    pcur = conn2.execute(
                        "INSERT INTO pipelines (workspace_id,name,stages_json,created_at) VALUES (?,?,?,?)",
                        (workspace_id, "Pipeline principal",
                         '["Nuevo","Contactado","Demo","Propuesta","Negociación","Cerrado","Perdido"]', now)
                    )
                    pipeline_id = pcur.lastrowid
                else:
                    pipeline_id = pipeline_row["id"]

                stage_map = {"offer_sent": "Propuesta", "negotiating": "Negociación", "closed": "Cerrado", "rejected": "Perdido"}
                card_stage = stage_map.get(decision, "Contactado")

                existing_card = conn2.execute(
                    "SELECT id FROM pipeline_cards WHERE workspace_id=? AND conversation_id=?",
                    (workspace_id, conversation_id)
                ).fetchone()
                if existing_card:
                    conn2.execute(
                        "UPDATE pipeline_cards SET stage=?, deal_value=?, currency=?, updated_at=? WHERE id=?",
                        (card_stage, round(negotiated_total, 2), currency, now, existing_card["id"])
                    )
                else:
                    conn2.execute(
                        "INSERT INTO pipeline_cards (workspace_id,pipeline_id,conversation_id,customer_name,deal_value,currency,stage,probability,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (workspace_id, pipeline_id, conversation_id, customer_name,
                         round(negotiated_total, 2), currency, card_stage, 50, now, now)
                    )
            except Exception:
                pass

            _log_agent_event(
                conn2, workspace_id, "sales_agent", "negotiate",
                "deal", deal_id,
                f"Customer: {customer_name} | msg: {customer_message[:100]}",
                f"Decision: {decision} | discount: {discount_pct:.1f}% | total: {negotiated_total:.2f}",
                confidence, requires_review,
            )
            conn2.commit()

        # Build human-readable offer text
        discount_str = f" ({discount_pct:.1f}% off)" if discount_pct > 0 else ""
        lines = [f"• {it.get('name', it.get('sku','?'))} x{int(it['qty'])} = {format_money(it['list_price'] * it['qty'], currency)}"
                 for it in enriched_items]
        total_str = format_money(negotiated_total, currency)

        if language == "es":
            if discount_pct > 0:
                offer_text = f"Preparé una propuesta para vos con {discount_pct:.1f}% de descuento especial:\n\n" + "\n".join(lines) + f"\n\n**Total: {total_str}{discount_str}**\n\nEsta oferta es válida por hoy. ¿La confirmamos?"
            else:
                offer_text = f"Acá está tu cotización:\n\n" + "\n".join(lines) + f"\n\n**Total: {total_str}**\n\n¿Avanzamos?"
        elif language == "pt":
            if discount_pct > 0:
                offer_text = f"Preparei uma proposta com {discount_pct:.1f}% de desconto especial:\n\n" + "\n".join(lines) + f"\n\n**Total: {total_str}{discount_str}**\n\nOferta válida por hoje. Confirmamos?"
            else:
                offer_text = f"Aqui está sua cotação:\n\n" + "\n".join(lines) + f"\n\n**Total: {total_str}**\n\nSeguimos?"
        else:
            if discount_pct > 0:
                offer_text = f"Here is a special offer with {discount_pct:.1f}% off:\n\n" + "\n".join(lines) + f"\n\n**Total: {total_str}{discount_str}**\n\nThis offer is valid today. Shall we confirm?"
            else:
                offer_text = f"Here is your quote:\n\n" + "\n".join(lines) + f"\n\n**Total: {total_str}**\n\nShall we move forward?"

        return {
            "deal_id": deal_id,
            "decision": decision,
            "discount_pct": discount_pct,
            "original_total": round(original_total, 2),
            "negotiated_total": round(negotiated_total, 2),
            "currency": currency,
            "offer_text": offer_text,
            "agent_note": agent_note,
            "requires_human_review": requires_review,
        }


# ── Agent 2: Accounting Agent ─────────────────────────────────────────────────
def accounting_agent_close_deal(
    workspace_id: int,
    deal_id: int,
    customer_tax_id: str | None = None,
) -> dict[str, Any]:
    """
    When a deal is accepted:
    1. Issues a fiscal invoice (with correct tax per country).
    2. Posts a ledger entry (Income / Paid).
    3. Records a commission event for value-based billing.
    4. Logs everything to agent_events.
    All in a single atomic transaction.
    """
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        deal = conn.execute("SELECT * FROM deals WHERE id = ? AND workspace_id = ?", (deal_id, workspace_id)).fetchone()
        if not deal:
            raise ValueError(f"Deal {deal_id} not found")
        if deal["status"] not in ("offer_sent", "negotiating"):
            raise ValueError(f"Deal {deal_id} is already {deal['status']}")

        deal = dict(deal)
        tax = get_tax_profile(conn, workspace_id)
        tax_pct = tax.get("default_tax_pct", 0.0)
        subtotal = deal["negotiated_total"]
        tax_amount = round(subtotal * tax_pct / 100, 2)
        total = round(subtotal + tax_amount, 2)
        invoice_number = _next_invoice_number(conn, workspace_id)

        # 1. Issue invoice
        inv_cur = conn.execute(
            """INSERT INTO invoices
               (workspace_id, deal_id, invoice_number, customer_name, customer_tax_id,
                items_json, subtotal, tax_pct, tax_amount, total, currency,
                country_code, fiscal_regime, status, issued_at, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (workspace_id, deal_id, invoice_number,
             deal["customer_name"], customer_tax_id or "",
             deal["items_json"], subtotal, tax_pct, tax_amount, total,
             deal["currency"], tax.get("country_code","US"),
             tax.get("fiscal_regime",""), "issued", now, now),
        )
        invoice_id = inv_cur.lastrowid

        # 2. Ledger entry (Income / Paid — instantaneous)
        ledger_cur = conn.execute(
            """INSERT INTO ledger_entries
               (workspace_id, entry_type, concept, category, amount, currency, state, due_date, source, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'auto', ?)""",
            (workspace_id, "Income",
             f"Invoice {invoice_number} — {deal['customer_name']}",
             "Sales", total, deal["currency"], "Paid", now[:10], now),
        )
        ledger_id = ledger_cur.lastrowid

        # Update invoice with ledger link
        conn.execute("UPDATE invoices SET ledger_entry_id = ? WHERE id = ?", (ledger_id, invoice_id))

        # 3. Mark deal closed
        conn.execute(
            "UPDATE deals SET status = 'closed', closed_at = ? WHERE id = ?",
            (now, deal_id),
        )

        # 3b. Deduct sold quantities from inventory stock automatically
        try:
            deal_items = json.loads(deal.get("items_json") or "[]")
            for item in deal_items:
                pid = item.get("product_id")
                qty_sold = item.get("qty", 0)
                if not pid or not qty_sold:
                    continue
                prod_row = conn.execute("SELECT id, stock FROM products WHERE id = ? AND workspace_id = ?", (pid, workspace_id)).fetchone()
                if not prod_row:
                    continue
                new_stock = max(0, prod_row["stock"] - int(qty_sold))
                conn.execute("UPDATE products SET stock = ?, updated_at = ? WHERE id = ?", (new_stock, now, pid))
                conn.execute(
                    "INSERT INTO stock_movements (workspace_id, product_id, delta, stock_after, movement_type, note, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (workspace_id, pid, -int(qty_sold), new_stock, "sale", f"Venta automatica - Deal {deal_id}", now),
                )
                stock_min_row = conn.execute("SELECT stock_min, name FROM products WHERE id = ?", (pid,)).fetchone()
                if stock_min_row and new_stock <= (stock_min_row["stock_min"] or 0):
                    try:
                        trigger_automations(workspace_id, "low_stock", {
                            "product_id": pid, "product_name": stock_min_row["name"], "stock": new_stock,
                        })
                    except Exception:
                        pass
        except Exception as stock_err:
            print(f"Stock deduction error for deal {deal_id}: {stock_err}")

        # 4. Commission event (value-based: 2% of deal value)
        COMMISSION_PCT = 0.02
        commission_amount = round(total * COMMISSION_PCT, 4)
        conn.execute(
            """INSERT INTO commission_events
               (workspace_id, deal_id, invoice_id, event_type, gross_value,
                commission_pct, commission_amount, currency, billed, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (workspace_id, deal_id, invoice_id, "sale_closed",
             total, COMMISSION_PCT, commission_amount,
             deal["currency"], 0, now),
        )

        # 5. Auditor Agent validates the transaction
        _log_agent_event(
            conn, workspace_id, "auditor_agent", "validate_close",
            "deal", deal_id,
            f"Deal {deal_id} | Invoice {invoice_number} | Total {total} {deal['currency']}",
            f"Invoice issued, ledger posted (entry {ledger_id}), commission {commission_amount:.4f}",
            0.99, False,
        )
        _log_agent_event(
            conn, workspace_id, "accounting_agent", "close_deal",
            "invoice", invoice_id,
            f"Deal {deal_id} | Subtotal {subtotal} | Tax {tax_pct}% = {tax_amount}",
            f"Invoice {invoice_number} issued and ledger entry {ledger_id} created instantly",
            1.0, False,
        )
        conn.commit()

        # Move the linked Pipeline card to "Cerrado" automatically
        try:
            if deal.get("conversation_id"):
                conn.execute(
                    "UPDATE pipeline_cards SET stage='Cerrado', updated_at=? WHERE workspace_id=? AND conversation_id=?",
                    (now, workspace_id, deal["conversation_id"])
                )
        except Exception:
            pass

        # Fire automations for "deal_closed" trigger — passes real customer contact info
        try:
            phone_row = None
            if deal.get("conversation_id"):
                phone_row = conn.execute(
                    "SELECT customer_phone FROM conversations WHERE id = ?", (deal["conversation_id"],)
                ).fetchone()
            trigger_automations(workspace_id, "deal_closed", {
                "deal_id": deal_id,
                "customer_name": deal.get("customer_name", ""),
                "customer_phone": phone_row["customer_phone"] if phone_row else None,
                "amount": total,
            })
        except Exception:
            pass

    return {
        "invoice_id": invoice_id,
        "invoice_number": invoice_number,
        "deal_id": deal_id,
        "ledger_entry_id": ledger_id,
        "subtotal": subtotal,
        "tax_pct": tax_pct,
        "tax_amount": tax_amount,
        "total": total,
        "currency": deal["currency"],
        "fiscal_regime": tax.get("fiscal_regime", ""),
        "commission_amount": commission_amount,
        "status": "closed",
    }


# ── Agent 3: Auditor Agent ────────────────────────────────────────────────────
def auditor_agent_review(workspace_id: int) -> dict[str, Any]:
    """
    Scans recent agent_events, deals, and invoices for anomalies:
    - Deals closed without invoices
    - Commission events not yet billed
    - Invoices with missing tax data
    - Agent events flagged for human review
    Returns a structured audit report.
    """
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        # Deals closed but no invoice
        orphan_deals = [dict(r) for r in conn.execute(
            """SELECT d.id, d.customer_name, d.negotiated_total, d.closed_at
               FROM deals d
               LEFT JOIN invoices i ON i.deal_id = d.id
               WHERE d.workspace_id = ? AND d.status = 'closed' AND i.id IS NULL""",
            (workspace_id,),
        ).fetchall()]

        # Commission not yet billed
        unbilled_commission = conn.execute(
            "SELECT SUM(commission_amount) as total, currency FROM commission_events WHERE workspace_id = ? AND billed = 0 GROUP BY currency",
            (workspace_id,),
        ).fetchall()

        # Agent events requiring human review
        pending_review = [dict(r) for r in conn.execute(
            """SELECT * FROM agent_events
               WHERE workspace_id = ? AND requires_human_review = 1 AND reviewed_at IS NULL
               ORDER BY id DESC LIMIT 20""",
            (workspace_id,),
        ).fetchall()]

        # Recent agent activity summary (last 24h)
        activity = [dict(r) for r in conn.execute(
            """SELECT agent, action, COUNT(*) as count, AVG(confidence) as avg_confidence
               FROM agent_events WHERE workspace_id = ?
               AND created_at >= datetime('now', '-24 hours')
               GROUP BY agent, action ORDER BY count DESC""",
            (workspace_id,),
        ).fetchall()]

        # Invoices with zero tax in taxable regions
        suspicious_invoices = [dict(r) for r in conn.execute(
            """SELECT id, invoice_number, total, country_code, tax_pct
               FROM invoices
               WHERE workspace_id = ? AND tax_pct = 0 AND country_code IN ('AR','BR') AND status != 'cancelled'
               ORDER BY id DESC LIMIT 10""",
            (workspace_id,),
        ).fetchall()]

        _log_agent_event(
            conn, workspace_id, "auditor_agent", "periodic_review",
            None, None,
            "Scheduled audit run",
            f"Orphan deals: {len(orphan_deals)} | Pending review: {len(pending_review)} | Suspicious invoices: {len(suspicious_invoices)}",
            1.0, len(orphan_deals) > 0 or len(suspicious_invoices) > 0,
        )
        conn.commit()

    return {
        "audit_at": now,
        "orphan_deals": orphan_deals,
        "unbilled_commission": [dict(r) for r in unbilled_commission],
        "pending_human_review": pending_review,
        "agent_activity_24h": activity,
        "suspicious_invoices": suspicious_invoices,
        "health": "warning" if (orphan_deals or suspicious_invoices) else "ok",
    }


# ── Orchestrator ──────────────────────────────────────────────────────────────
def orchestrator_process_message(
    workspace_id: int,
    conversation_id: int | None,
    customer_name: str,
    customer_message: str,
    language: str,
    currency: str,
) -> dict[str, Any]:
    """
    Main entry point for autonomous message processing.
    Routes message to the right agent(s) and returns a unified response.
    """
    now = datetime.utcnow().isoformat()
    msg_lower = customer_message.lower()

    # Detect intent
    CLOSE_SIGNALS = ["acepto", "confirmamos", "sí quiero", "lo tomo", "confirm", "yes i accept",
                     "deal", "accepted", "let's do it", "aceito", "confirmado", "vamos"]
    PRICE_SIGNALS = ["precio", "price", "cotización", "quote", "preço", "orçamento",
                     "cuánto", "how much", "quanto"]
    NEGOTIATE_SIGNALS = ["descuento", "discount", "desconto", "rebate", "deal", "cheaper",
                         "negociar", "better price", "melhor preço"]

    wants_close = any(s in msg_lower for s in CLOSE_SIGNALS)
    wants_price = any(s in msg_lower for s in PRICE_SIGNALS)
    wants_negotiate = any(s in msg_lower for s in NEGOTIATE_SIGNALS)

    # Check for open deal for this customer in this conversation
    with closing(get_db()) as conn:
        open_deal = conn.execute(
            """SELECT * FROM deals WHERE workspace_id = ? AND customer_name = ?
               AND status IN ('negotiating','offer_sent') ORDER BY id DESC LIMIT 1""",
            (workspace_id, customer_name),
        ).fetchone()
        open_deal = dict(open_deal) if open_deal else None

        products = [dict(r) for r in conn.execute(
            "SELECT * FROM products WHERE workspace_id = ? ORDER BY demand_score DESC",
            (workspace_id,),
        ).fetchall()]

    # Route to agents
    if wants_close and open_deal:
        # Customer accepts → Accounting Agent closes instantly
        try:
            result = accounting_agent_close_deal(workspace_id, open_deal["id"])
            deal_currency = result["currency"]
            tax_line = f" (incluye {result['tax_pct']:.0f}% impuesto)" if result["tax_pct"] > 0 else ""
            if language == "es":
                reply = (f"¡Perfecto! Cerramos el trato. 🎉\n\n"
                         f"**Factura {result['invoice_number']}** emitida al instante{tax_line}.\n"
                         f"Total cobrado: **{format_money(result['total'], deal_currency)}**\n\n"
                         f"El asiento contable ya está registrado. ¿Algo más?")
            elif language == "pt":
                reply = (f"Perfeito! Negócio fechado. 🎉\n\n"
                         f"**Nota fiscal {result['invoice_number']}** emitida instantaneamente{tax_line}.\n"
                         f"Total: **{format_money(result['total'], deal_currency)}**\n\n"
                         f"O lançamento contábil já foi registrado. Mais alguma coisa?")
            else:
                reply = (f"Deal closed! 🎉\n\n"
                         f"**Invoice {result['invoice_number']}** issued instantly{tax_line}.\n"
                         f"Total: **{format_money(result['total'], deal_currency)}**\n\n"
                         f"Ledger entry posted automatically. Anything else?")
            return {"reply": reply, "agent": "accounting_agent", "action": "deal_closed", "data": result}
        except Exception as exc:
            # Fall through to standard reply
            pass

    if (wants_negotiate or wants_price) and products:
        # Build proposed items from products matching the message or top products
        matched = [p for p in products if p["name"].split()[0].lower() in msg_lower or p["sku"].lower() in msg_lower]
        if not matched:
            matched = products[:2]  # default: top 2 by demand

        proposed_items = [
            {"product_id": str(p["id"]), "sku": p["sku"], "name": p["name"],
             "unit_price": p["price"], "qty": 1}
            for p in matched[:3]
        ]
        result = sales_agent_negotiate(
            workspace_id, conversation_id, customer_name,
            customer_message, proposed_items, language, currency,
        )
        return {"reply": result["offer_text"], "agent": "sales_agent",
                "action": "negotiate", "data": result}

    # Default: use AI reply (OpenAI or local)
    ai_result = generate_ai_reply(workspace_id, customer_message, language, currency)
    return {"reply": ai_result["reply"], "agent": "ai_engine",
            "action": "reply", "data": ai_result}


def pricing_recommendation(product: dict[str, Any], target_margin: float = 0.42) -> dict[str, float]:
    floor = product["cost"] / max(0.01, (1 - target_margin))
    competitor_anchor = product["competitor_price"] * 0.98 if product["competitor_price"] else floor
    demand_boost = 1.06 if product["demand_score"] > 75 else 0.97 if product["demand_score"] < 50 else 1.0
    recommended = max(floor, competitor_anchor) * demand_boost
    current_margin = ((product["price"] - product["cost"]) / product["price"] * 100) if product["price"] else 0
    projected_margin = ((recommended - product["cost"]) / recommended * 100) if recommended else 0
    return {
        "recommended_price": round(recommended, 2),
        "current_margin_pct": round(current_margin, 2),
        "projected_margin_pct": round(projected_margin, 2),
        "delta": round(recommended - product["price"], 2),
    }


def fetch_messages(conn: sqlite3.Connection, conversation_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        "SELECT role, text, created_at FROM messages WHERE conversation_id = ? ORDER BY id ASC",
        (conversation_id,),
    ).fetchall()
    return [dict(r) for r in rows]


def workspace_dashboard(workspace_id: int) -> dict[str, Any]:
    with closing(get_db()) as conn:
        conversations = [dict(r) for r in conn.execute(
            "SELECT * FROM conversations WHERE workspace_id = ? ORDER BY id DESC",
            (workspace_id,),
        ).fetchall()]
        for conv in conversations:
            conv["messages"] = fetch_messages(conn, conv["id"])
        products = [dict(r) for r in conn.execute("SELECT * FROM products WHERE workspace_id = ? ORDER BY id DESC", (workspace_id,)).fetchall()]
        tasks = [dict(r) for r in conn.execute("SELECT * FROM tasks WHERE workspace_id = ? ORDER BY impact DESC, id DESC", (workspace_id,)).fetchall()]
        ledger = [dict(r) for r in conn.execute("SELECT * FROM ledger_entries WHERE workspace_id = ? ORDER BY id DESC", (workspace_id,)).fetchall()]
        knowledge = [dict(r) for r in conn.execute("SELECT * FROM knowledge_articles WHERE workspace_id = ? ORDER BY id DESC", (workspace_id,)).fetchall()]
        sources = fetch_workspace_sources(conn, workspace_id)
        traces = [dict(r) for r in conn.execute("SELECT * FROM traces WHERE workspace_id = ? ORDER BY id DESC LIMIT 20", (workspace_id,)).fetchall()]
        templates = [dict(r) for r in conn.execute("SELECT * FROM templates WHERE workspace_id = ? ORDER BY id DESC", (workspace_id,)).fetchall()]
        profile = get_workspace_profile(conn, workspace_id)

    income = sum(e["amount"] for e in ledger if e["entry_type"] == "Income")
    expenses = sum(e["amount"] for e in ledger if e["entry_type"] == "Expense")
    hot_leads = sum(1 for c in conversations if c["status"] in {"Hot lead", "Qualified"})
    pending_orders = sum(1 for e in ledger if e["entry_type"] == "Income" and e["state"] == "Pending")

    return {
        "conversations": conversations,
        "products": products,
        "tasks": tasks,
        "ledger": ledger,
        "knowledge": knowledge,
        "sources": sources,
        "templates": templates,
        "traces": traces,
        "profile": profile,
        "kpis": {
            "open_chats": len(conversations),
            "hot_leads": hot_leads,
            "pending_income": pending_orders,
            "income": round(income, 2),
            "expenses": round(expenses, 2),
            "net": round(income - expenses, 2),
        },
    }


# ── Hardening: global error handling, validation helpers, security ─────────

import re as _re_mod
import logging as _logging_mod
import time as _time_mod
from collections import defaultdict as _defaultdict

_logger = _logging_mod.getLogger("banzai")
_logger.setLevel(_logging_mod.INFO)


@app.errorhandler(404)
def _handle_404(e):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": "Recurso no encontrado"}), 404
    return e


@app.errorhandler(500)
def _handle_500(e):
    _logger.error(f"Internal error on {request.path}: {e}")
    if request.path.startswith("/api/") or request.path.startswith("/webhook/"):
        return jsonify({"ok": False, "error": "Ocurrió un error interno. Intentá de nuevo en unos segundos."}), 500
    return jsonify({"ok": False, "error": "internal_error"}), 500


@app.errorhandler(Exception)
def _handle_unexpected_exception(e):
    """Catch-all: never let an unhandled exception crash a worker or leak a stack trace to the client."""
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e
    _logger.error(f"Unhandled exception on {request.path}: {type(e).__name__}: {e}")
    if request.path.startswith("/webhook/"):
        return ("", 200)  # webhooks should always ack, never retry-storm the sender
    return jsonify({"ok": False, "error": "Ocurrió un error inesperado. Intentá de nuevo."}), 500


@app.after_request
def _add_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response


def is_valid_email(value: str) -> bool:
    if not value or not isinstance(value, str):
        return False
    return bool(_re_mod.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", value.strip()))


def is_positive_number(value, allow_zero: bool = True) -> bool:
    try:
        n = float(value)
    except (TypeError, ValueError):
        return False
    if allow_zero:
        return n >= 0
    return n > 0


def sanitize_text(value: str, max_len: int = 500) -> str:
    if not value or not isinstance(value, str):
        return ""
    return value.strip()[:max_len]


# ── Simple in-memory rate limiter for sensitive endpoints (login, setup) ───
_rate_limit_buckets = _defaultdict(list)

def rate_limited(key: str, max_attempts: int = 8, window_seconds: int = 60) -> bool:
    """Returns True if the caller should be blocked (too many attempts)."""
    now = _time_mod.time()
    bucket = _rate_limit_buckets[key]
    bucket[:] = [t for t in bucket if now - t < window_seconds]
    if len(bucket) >= max_attempts:
        return True
    bucket.append(now)
    return False


@app.route("/")
def index():
    if not is_setup_complete():
        return render_template("setup.html")
    return render_template("index.html")


@app.get("/setup")
def setup_page():
    if is_setup_complete():
        return redirect("/")
    return render_template("setup.html")


@app.get("/signup")
def signup_page():
    """Always shows the workspace creation wizard, even if other workspaces already exist.
    Used to onboard additional businesses (multi-tenant) beyond the very first one."""
    return render_template("setup.html")


@app.post("/api/setup")
def api_setup():
    """Creates a new workspace and owner account. Works for the first-ever setup
    AND for onboarding additional businesses afterwards (multi-tenant)."""
    payload = request.get_json(force=True)
    workspace_name = (payload.get("workspace_name") or "").strip()
    owner_name     = (payload.get("owner_name") or "").strip()
    owner_email    = (payload.get("owner_email") or "").strip().lower()
    password       = (payload.get("password") or "").strip()
    region         = (payload.get("region") or "US").strip().upper()
    currency       = (payload.get("currency") or "USD").strip().upper()
    language       = (payload.get("language") or "en").strip().lower()

    if not workspace_name or not owner_name or not owner_email or not password:
        return json_error("All fields are required")
    if not is_valid_email(owner_email):
        return json_error("El email no tiene un formato valido")
    if len(password) < 6:
        return json_error("Password must be at least 6 characters")
    if rate_limited(f"setup:{request.remote_addr}", max_attempts=5, window_seconds=300):
        return json_error("Demasiados intentos de registro. Esperá unos minutos e intentá de nuevo.", 429)

    with closing(get_db()) as _conn_check:
        existing_user = _conn_check.execute(
            "SELECT id FROM users WHERE email = ?", (owner_email,)
        ).fetchone()
    if existing_user:
        return json_error("Ya existe una cuenta con ese email. Usa el login o un email diferente.", 400)

    # Derive slug from workspace name
    import re as _re
    slug = _re.sub(r"[^a-z0-9]+", "-", workspace_name.lower()).strip("-") or "workspace"

    now = datetime.utcnow().isoformat()
    REGION_DEFAULTS = {
        "AR": {"currency": "ARS", "language": "es"},
        "BR": {"currency": "BRL", "language": "pt"},
        "US": {"currency": "USD", "language": "en"},
    }
    defaults = REGION_DEFAULTS.get(region, {"currency": currency, "language": language})
    final_currency = currency or defaults["currency"]
    final_language = language or defaults["language"]

    with closing(get_db()) as conn:
        # Create workspace
        cur = conn.execute(
            "INSERT INTO workspaces (slug, name, region, currency, language, tone, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (slug, workspace_name, region, final_currency, final_language, "warm_professional", now),
        )
        workspace_id = cur.lastrowid

        # Create owner account
        conn.execute(
            "INSERT INTO users (workspace_id, email, password_hash, name, role, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (workspace_id, owner_email, generate_password_hash(password), owner_name, "owner", now),
        )

        # Default workspace profile
        conn.execute(
            "INSERT INTO workspace_profiles (workspace_id, response_style, personality_notes, forbidden_tone, updated_at) VALUES (?, ?, ?, ?, ?)",
            (workspace_id, "Professional", "Sound clear, helpful and commercially aware.", "Avoid sounding robotic or overly formal.", now),
        )

        # Default templates
        conn.execute(
            "INSERT INTO templates (workspace_id, name, category, body, created_at) VALUES (?, ?, ?, ?, ?)",
            (workspace_id, "Quote follow-up", "Sales", "Hi {{name}}, following up on the quote I sent. Happy to confirm it right now if you're ready.", now),
        )
        conn.execute(
            "INSERT INTO templates (workspace_id, name, category, body, created_at) VALUES (?, ?, ?, ?, ?)",
            (workspace_id, "New lead welcome", "Inbound", "Hi {{name}}, thanks for reaching out. I can help with pricing, availability and next steps.", now),
        )

        # Mark setup complete
        database_url = os.environ.get("DATABASE_URL", "")
        is_postgres = database_url and database_url.startswith("postgres")
        if is_postgres:
            conn.execute(
                "INSERT INTO setup (id, completed, workspace_name, owner_name, owner_email, region, currency, language, completed_at) VALUES (1, 1, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT (id) DO UPDATE SET completed=1, completed_at=EXCLUDED.completed_at",
                (workspace_name, owner_name, owner_email, region, final_currency, final_language, now),
            )
        else:
            conn.execute(
                "INSERT OR REPLACE INTO setup (id, completed, workspace_name, owner_name, owner_email, region, currency, language, completed_at) VALUES (1, 1, ?, ?, ?, ?, ?, ?, ?)",
                (workspace_name, owner_name, owner_email, region, final_currency, final_language, now),
            )
        conn.commit()

    return jsonify({"ok": True, "redirect": "/"})



@app.post("/api/login")
def api_login():
    if rate_limited(f"login:{request.remote_addr}", max_attempts=15, window_seconds=300):
        return json_error("Demasiados intentos de login desde esta conexión. Esperá unos minutos.", 429)
    payload = request.get_json(force=True)
    email = payload.get("email", "").strip().lower()
    password = payload.get("password", "")
    now = datetime.utcnow().isoformat()
    ip = request.remote_addr or ""
    with closing(get_db()) as conn:
        row = conn.execute(
            """SELECT users.*, workspaces.id AS wid, workspaces.name AS workspace_name,
                      workspaces.slug AS workspace_slug, workspaces.currency, workspaces.language
               FROM users JOIN workspaces ON users.workspace_id = workspaces.id
               WHERE users.email = ?""",
            (email,),
        ).fetchone()
        if not row:
            return json_error("Invalid credentials", 401)
        row = dict(row)
        workspace_id = row["wid"]

        # Inactive account
        if not row.get("active", 1):
            return json_error("Account is inactive. Contact the owner.", 403)

        # Brute-force lock
        locked = row.get("locked_until")
        if locked and locked > now:
            return json_error("Account locked for 15 minutes due to failed attempts.", 429)

        # Demo expiry
        if row.get("role") == "demo" and row.get("demo_expires_at"):
            if row["demo_expires_at"] < now:
                return json_error("Demo access expired. Contact the workspace owner.", 403)

        # Wrong password
        if not check_password_hash(row["password_hash"], password):
            fails = (row.get("failed_login_count") or 0) + 1
            lock = None
            if fails >= 5:
                from datetime import timedelta
                lock = (datetime.utcnow() + timedelta(minutes=15)).isoformat()
            conn.execute("UPDATE users SET failed_login_count=?, locked_until=? WHERE id=?",
                         (fails, lock, row["id"]))
            conn.commit()
            log_access(row["id"], workspace_id, "login_failed", detail=f"attempt={fails} ip={ip}")
            return json_error("Invalid credentials", 401)

        # Success
        conn.execute("UPDATE users SET failed_login_count=0, locked_until=NULL, last_login_at=?, last_login_ip=? WHERE id=?",
                     (now, ip, row["id"]))
        conn.commit()

    session["user_id"] = row["id"]
    session["workspace_id"] = workspace_id
    log_access(row["id"], workspace_id, "login", detail=f"ip={ip} role={row['role']}")

    perms = ROLE_PERMISSIONS.get(row["role"], ROLE_PERMISSIONS["viewer"]).copy()
    if row.get("can_see_costs"): perms["see_costs"] = True
    if row.get("can_see_finances"): perms["see_finances"] = True
    if row.get("can_export"): perms["export"] = True

    return jsonify({"ok": True, "user": {
        "id": row["id"], "email": row["email"], "name": row["name"],
        "role": row["role"], "workspace_id": workspace_id,
        "workspace_name": row["workspace_name"], "workspace_slug": row["workspace_slug"],
        "currency": row["currency"], "language": row["language"],
        "permissions": perms, "watermark": bool(row.get("watermark_demos", 1) and row["role"] in ("seller","demo")),
        "demo_expires_at": row.get("demo_expires_at"),
    }})


@app.post("/api/logout")
def api_logout():
    session.clear()
    return jsonify({"ok": True})


@app.get("/api/me")
def api_me():
    user = current_user()
    if not user:
        return json_error("Not authenticated", 401)
    perms = ROLE_PERMISSIONS.get(user.get("role","viewer"), ROLE_PERMISSIONS["viewer"]).copy()
    user["permissions"] = perms
    user["watermark"] = user.get("role") in ("seller", "demo")
    return jsonify({"ok": True, "user": user})


@app.get("/api/dashboard")
def api_dashboard():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    return jsonify({"ok": True, "dashboard": workspace_dashboard(user["workspace_id"]), "user": user})


@app.post("/api/messages/reply")
def api_reply():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    conversation_id = int(payload.get("conversation_id"))
    text = payload.get("text", "").strip()
    if not text:
        return json_error("Reply text is required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute("INSERT INTO messages (conversation_id, role, text, created_at) VALUES (?, ?, ?, ?)", (conversation_id, "assistant", text, now))
        conn.execute("UPDATE conversations SET updated_at = ? WHERE id = ? AND workspace_id = ?", (now, conversation_id, user["workspace_id"]))
        conn.execute(
            "INSERT INTO traces (workspace_id, flow, customer, status, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (user["workspace_id"], "send_reply", f"conversation:{conversation_id}", "ok", "Reply sent from UI", now),
        )
        conn.commit()
    return jsonify({"ok": True})


@app.post("/api/webchat/inbound")
def api_webchat_inbound():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    conversation_id = payload.get("conversation_id")
    customer_name = payload.get("customer_name", "Webchat Lead").strip() or "Webchat Lead"
    text = payload.get("text", "").strip()
    if not text:
        return json_error("Message text is required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        if conversation_id:
            row = conn.execute(
                "SELECT id, customer_name FROM conversations WHERE id = ? AND workspace_id = ?",
                (conversation_id, user["workspace_id"]),
            ).fetchone()
            if not row:
                return json_error("Conversation not found", 404)
            conv_id = row["id"]
            customer_name = row["customer_name"]
        else:
            cur = conn.execute(
                "INSERT INTO conversations (workspace_id, customer_name, channel, status, country, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (user["workspace_id"], customer_name, "webchat", "New inquiry", user["workspace_slug"][:2].upper(), now, now),
            )
            conv_id = cur.lastrowid

        conn.execute("INSERT INTO messages (conversation_id, role, text, created_at) VALUES (?, ?, ?, ?)", (conv_id, "customer", text, now))
        ai_payload = generate_ai_reply(user["workspace_id"], text, user["language"], user["currency"])
        reply = ai_payload["reply"]
        conn.execute("INSERT INTO messages (conversation_id, role, text, created_at) VALUES (?, ?, ?, ?)", (conv_id, "assistant", reply, now))
        conn.execute("UPDATE conversations SET status = ?, updated_at = ? WHERE id = ?", ("Hot lead", now, conv_id))
        conn.execute(
            "INSERT INTO traces (workspace_id, flow, customer, status, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (user["workspace_id"], "webchat_inbound", customer_name, "ok", f"Inbound message processed and reply generated via {ai_payload.get('provider')}", now),
        )
        conn.commit()
    return jsonify({"ok": True, "conversation_id": conv_id, "reply": reply})


@app.get("/api/products")
def api_products():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        products = [dict(r) for r in conn.execute("SELECT * FROM products WHERE workspace_id = ? ORDER BY id DESC", (user["workspace_id"],)).fetchall()]
    return jsonify({"ok": True, "products": products})


@app.post("/api/products")
def api_products_create():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    required = ["sku", "name", "category", "cost", "price", "stock"]
    if any(payload.get(key) in [None, ""] for key in required):
        return json_error("Missing product fields")
    if not is_positive_number(payload.get("cost")):
        return json_error("El costo no puede ser negativo")
    if not is_positive_number(payload.get("price")):
        return json_error("El precio no puede ser negativo")
    try:
        if int(payload.get("stock")) < 0:
            return json_error("El stock no puede ser negativo")
    except (TypeError, ValueError):
        return json_error("Stock invalido")
    now = datetime.utcnow().isoformat()
    stock_min_val = int(payload.get("stock_min", 0) or 0)
    if stock_min_val < 0:
        return json_error("El stock minimo no puede ser negativo")
    with closing(get_db()) as conn:
        conn.execute(
            """
            INSERT INTO products (workspace_id, sku, name, category, cost, price, stock, stock_min, unit, barcode, competitor_price, demand_score, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user["workspace_id"],
                payload["sku"],
                payload["name"],
                payload["category"],
                float(payload["cost"]),
                float(payload["price"]),
                int(payload["stock"]),
                stock_min_val,
                sanitize_text(payload.get("unit", "unit"), 20) or "unit",
                sanitize_text(payload.get("barcode", ""), 60),
                float(payload.get("competitor_price", 0)),
                int(payload.get("demand_score", 60)),
                now,
            ),
        )
        conn.commit()
    return jsonify({"ok": True})



@app.put("/api/products/<int:product_id>")
def api_products_update(product_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        row = conn.execute("SELECT id FROM products WHERE id = ? AND workspace_id = ?", (product_id, user["workspace_id"])).fetchone()
        if not row:
            return json_error("Product not found", 404)
        fields = []
        values = []
        for field, cast in [("sku", str), ("name", str), ("category", str), ("cost", float), ("price", float),
                             ("stock", int), ("stock_min", int), ("unit", str), ("barcode", str),
                             ("notes", str), ("competitor_price", float), ("demand_score", int)]:
            if field in payload and payload[field] != "" and payload[field] is not None:
                if field in ("cost", "price") and not is_positive_number(payload[field]):
                    return json_error(f"{field} no puede ser negativo")
                if field in ("stock", "stock_min") and int(payload[field]) < 0:
                    return json_error(f"{field} no puede ser negativo")
                fields.append(f"{field} = ?")
                values.append(cast(payload[field]))
        if "active" in payload:
            fields.append("active = ?")
            values.append(1 if payload["active"] else 0)
        if not fields:
            return json_error("No fields to update")
        values += [now, product_id, user["workspace_id"]]
        conn.execute(f"UPDATE products SET {', '.join(fields)}, updated_at = ? WHERE id = ? AND workspace_id = ?", values)
        conn.commit()
        updated = dict(conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone())
    return jsonify({"ok": True, "product": updated})


@app.delete("/api/products/<int:product_id>")
def api_products_delete(product_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        row = conn.execute("SELECT id FROM products WHERE id = ? AND workspace_id = ?", (product_id, user["workspace_id"])).fetchone()
        if not row:
            return json_error("Product not found", 404)
        conn.execute("DELETE FROM products WHERE id = ? AND workspace_id = ?", (product_id, user["workspace_id"]))
        conn.commit()
    return jsonify({"ok": True})


@app.post("/api/products/<int:product_id>/stock")
def api_products_stock(product_id: int):
    """Adjust stock: delta can be positive (restock) or negative (manual deduct)."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    delta = int(payload.get("delta", 0))
    note = payload.get("note", "")
    if delta == 0:
        return json_error("Delta cannot be zero")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        row = conn.execute("SELECT * FROM products WHERE id = ? AND workspace_id = ?", (product_id, user["workspace_id"])).fetchone()
        if not row:
            return json_error("Product not found", 404)
        new_stock = max(0, row["stock"] + delta)
        conn.execute("UPDATE products SET stock = ?, updated_at = ? WHERE id = ?", (new_stock, now, product_id))
        conn.execute(
            "INSERT INTO stock_movements (workspace_id, product_id, delta, stock_after, movement_type, note, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (user["workspace_id"], product_id, delta, new_stock, "manual", note, now),
        )
        conn.commit()
        updated = dict(conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone())
    return jsonify({"ok": True, "product": updated, "new_stock": new_stock})


@app.post("/api/products/bulk")
def api_products_bulk():
    """Bulk import products from a list. Upserts by SKU."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    items = payload.get("products", [])
    if not items or not isinstance(items, list):
        return json_error("products array required")
    if len(items) > 2000:
        return json_error("Maximo 2000 productos por carga. Dividi el archivo en partes mas chicas.")
    now = datetime.utcnow().isoformat()
    created = updated = 0
    error_details = []
    with closing(get_db()) as conn:
        for idx, item in enumerate(items):
            row_label = f"Fila {idx+1}"
            try:
                sku = str(item.get("sku", "")).strip()
                name = str(item.get("name", "")).strip()
                if not sku:
                    error_details.append(f"{row_label}: falta SKU"); continue
                if not name:
                    error_details.append(f"{row_label} (SKU {sku}): falta nombre"); continue
                if not is_positive_number(item.get("cost", 0)):
                    error_details.append(f"{row_label} (SKU {sku}): costo negativo o invalido"); continue
                if not is_positive_number(item.get("price", 0)):
                    error_details.append(f"{row_label} (SKU {sku}): precio negativo o invalido"); continue
                try:
                    stock_val = int(item.get("stock", 0))
                    if stock_val < 0:
                        error_details.append(f"{row_label} (SKU {sku}): stock negativo"); continue
                except (TypeError, ValueError):
                    error_details.append(f"{row_label} (SKU {sku}): stock invalido"); continue

                stock_min_val = int(item.get("stock_min", 0) or 0)
                if stock_min_val < 0:
                    error_details.append(f"{row_label} (SKU {sku}): stock_min negativo"); continue

                existing = conn.execute(
                    "SELECT id FROM products WHERE workspace_id = ? AND sku = ?",
                    (user["workspace_id"], sku),
                ).fetchone()
                if existing:
                    conn.execute(
                        """UPDATE products SET name=?, category=?, cost=?, price=?, stock=?, stock_min=?, competitor_price=?, demand_score=?, updated_at=?
                           WHERE id=?""",
                        (name, sanitize_text(item.get("category","General"), 100), float(item.get("cost",0)),
                         float(item.get("price",0)), stock_val, stock_min_val,
                         float(item.get("competitor_price",0)), int(item.get("demand_score",60)), now,
                         existing["id"]),
                    )
                    updated += 1
                else:
                    conn.execute(
                        """INSERT INTO products (workspace_id,sku,name,category,cost,price,stock,stock_min,competitor_price,demand_score,created_at)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                        (user["workspace_id"], sku, name, sanitize_text(item.get("category","General"), 100),
                         float(item.get("cost",0)), float(item.get("price",0)),
                         stock_val, stock_min_val, float(item.get("competitor_price",0)),
                         int(item.get("demand_score",60)), now),
                    )
                    created += 1
            except Exception as row_err:
                error_details.append(f"{row_label}: {str(row_err)[:100]}")
        conn.commit()
    return jsonify({
        "ok": True, "created": created, "updated": updated, "errors": len(error_details),
        "error_details": error_details[:50],  # cap the detail list so a huge bad file doesn't blow up the response
    })



@app.get("/api/products/<int:product_id>/movements")
def api_product_movements(product_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        row = conn.execute("SELECT id FROM products WHERE id = ? AND workspace_id = ?", (product_id, user["workspace_id"])).fetchone()
        if not row:
            return json_error("Product not found", 404)
        movements = [dict(r) for r in conn.execute(
            "SELECT * FROM stock_movements WHERE product_id = ? ORDER BY id DESC LIMIT 50",
            (product_id,),
        ).fetchall()]
    return jsonify({"ok": True, "movements": movements})


@app.post("/api/quotes")
def api_quotes():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    customer = payload.get("customer", "Unknown")
    items = payload.get("items", [])
    if not items:
        return json_error("At least one item is required")
    total = round(sum(float(i["qty"]) * float(i["unit_price"]) for i in items), 2)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "INSERT INTO ledger_entries (workspace_id, entry_type, concept, category, amount, currency, state, due_date, source, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (user["workspace_id"], "Income", f"Quote for {customer}", "Sales", total, user["currency"], "Pending", now[:10], "auto", now),
        )
        conn.execute(
            "INSERT INTO traces (workspace_id, flow, customer, status, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (user["workspace_id"], "create_quote", customer, "ok", f"Quote created with total {total}", now),
        )
        conn.commit()
    return jsonify({"ok": True, "total": total})


@app.get("/api/tasks")
def api_tasks():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        tasks = [dict(r) for r in conn.execute("SELECT * FROM tasks WHERE workspace_id = ? ORDER BY impact DESC, id DESC", (user["workspace_id"],)).fetchall()]
    return jsonify({"ok": True, "tasks": tasks})


@app.post("/api/tasks")
def api_tasks_create():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    title = payload.get("title", "").strip()
    if not title:
        return json_error("Title is required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "INSERT INTO tasks (workspace_id, title, area, owner, status, priority, impact, due_label, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                user["workspace_id"],
                title,
                payload.get("area", "Admin"),
                payload.get("owner", user["name"]),
                payload.get("status", "Backlog"),
                payload.get("priority", "Medium"),
                int(payload.get("impact", 5)),
                payload.get("due_label", "This week"),
                now,
            ),
        )
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/ledger")
def api_ledger():
    try:
        user = require_auth()
        require_permission(user, "see_finances")
    except PermissionError as e:
        return json_error(str(e), 403)
    with closing(get_db()) as conn:
        entries = [dict(r) for r in conn.execute("SELECT * FROM ledger_entries WHERE workspace_id = ? ORDER BY id DESC", (user["workspace_id"],)).fetchall()]
    return jsonify({"ok": True, "entries": entries})


@app.post("/api/ledger")
def api_ledger_create():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    concept = sanitize_text(payload.get("concept", ""), max_len=300)
    if not concept:
        return json_error("Concept is required")
    entry_type = payload.get("entry_type", "Expense")
    if entry_type not in ("Income", "Expense"):
        return json_error("entry_type debe ser 'Income' o 'Expense'")
    if not is_positive_number(payload.get("amount", 0), allow_zero=False):
        return json_error("El monto debe ser mayor a cero")
    state = payload.get("state", "Pending")
    if state not in ("Pending", "Paid", "Overdue", "Cancelled"):
        return json_error("Estado invalido")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "INSERT INTO ledger_entries (workspace_id, entry_type, concept, category, amount, currency, state, due_date, source, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                user["workspace_id"],
                entry_type,
                concept,
                sanitize_text(payload.get("category", "Operations"), max_len=100),
                float(payload.get("amount", 0)),
                payload.get("currency", user["currency"]),
                state,
                payload.get("due_date") or now[:10],
                "manual",
                now,
            ),
        )
        conn.commit()
    return jsonify({"ok": True})


@app.put("/api/ledger/<int:entry_id>")
def api_ledger_update(entry_id):
    """Edit a manually-entered ledger entry. Auto-generated entries (from sales or vendor
    invoices) can't be edited here — that would corrupt the automatic accounting trail."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    with closing(get_db()) as conn:
        row = conn.execute(
            "SELECT source FROM ledger_entries WHERE id=? AND workspace_id=?",
            (entry_id, user["workspace_id"])
        ).fetchone()
        if not row:
            return json_error("Entry not found", 404)
        if row["source"] != "manual":
            return json_error("Este asiento fue generado automaticamente y no se puede editar directamente", 403)
        fields, values = [], []
        if "concept" in payload:
            fields.append("concept = ?"); values.append(sanitize_text(payload["concept"], 300))
        if "category" in payload:
            fields.append("category = ?"); values.append(sanitize_text(payload["category"], 100))
        if "amount" in payload:
            if not is_positive_number(payload["amount"], allow_zero=False):
                return json_error("El monto debe ser mayor a cero")
            fields.append("amount = ?"); values.append(float(payload["amount"]))
        if "state" in payload and payload["state"] in ("Pending", "Paid", "Overdue", "Cancelled"):
            fields.append("state = ?"); values.append(payload["state"])
        if "due_date" in payload:
            fields.append("due_date = ?"); values.append(payload["due_date"])
        if not fields:
            return json_error("Nada para actualizar")
        values += [entry_id, user["workspace_id"]]
        conn.execute(f"UPDATE ledger_entries SET {', '.join(fields)} WHERE id=? AND workspace_id=?", values)
        conn.commit()
    return jsonify({"ok": True})


@app.delete("/api/ledger/<int:entry_id>")
def api_ledger_delete(entry_id):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        row = conn.execute(
            "SELECT source FROM ledger_entries WHERE id=? AND workspace_id=?",
            (entry_id, user["workspace_id"])
        ).fetchone()
        if not row:
            return json_error("Entry not found", 404)
        if row["source"] != "manual":
            return json_error("Este asiento fue generado automaticamente y no se puede borrar directamente", 403)
        conn.execute("DELETE FROM ledger_entries WHERE id=? AND workspace_id=?", (entry_id, user["workspace_id"]))
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/knowledge")
def api_knowledge():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        articles = [dict(r) for r in conn.execute("SELECT * FROM knowledge_articles WHERE workspace_id = ? ORDER BY id DESC", (user["workspace_id"],)).fetchall()]
    return jsonify({"ok": True, "articles": articles})


@app.post("/api/knowledge")
def api_knowledge_create():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    title = payload.get("title", "").strip()
    content = payload.get("content", "").strip()
    if not title or not content:
        return json_error("Title and content are required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "INSERT INTO knowledge_articles (workspace_id, title, content, created_at) VALUES (?, ?, ?, ?)",
            (user["workspace_id"], title, content, now),
        )
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/sources")
def api_sources():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        sources = fetch_workspace_sources(conn, user["workspace_id"])
    return jsonify({"ok": True, "sources": sources})


@app.post("/api/sources")
def api_sources_create():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    title = payload.get("title", "").strip()
    content = payload.get("content", "").strip()
    if not title or not content:
        return json_error("Title and content are required")
    domain = payload.get("domain", "general").strip() or "general"
    source_type = payload.get("source_type", "manual").strip() or "manual"
    excerpt = content[:220]
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "INSERT INTO source_documents (workspace_id, title, domain, source_type, content, excerpt, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (user["workspace_id"], title, domain, source_type, content, excerpt, now),
        )
        conn.commit()
    return jsonify({"ok": True})


@app.post("/api/sources/upload")
def api_sources_upload():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    uploaded = request.files.get("file")
    domain = request.form.get("domain", "general").strip() or "general"
    if not uploaded or not uploaded.filename:
        return json_error("File is required")
    filename = Path(uploaded.filename).name
    raw = uploaded.read()
    try:
        if filename.lower().endswith('.json'):
            parsed = json.loads(raw.decode('utf-8', errors='ignore'))
            if isinstance(parsed, dict):
                content = json.dumps(parsed, ensure_ascii=False, indent=2)
            else:
                content = json.dumps(parsed[:20], ensure_ascii=False, indent=2)
        else:
            content = raw.decode('utf-8', errors='ignore')
    except Exception:
        content = ''
    if not content.strip():
        return json_error("This local uploader currently supports text-like files such as txt, md, csv and json")
    content = content.strip()
    excerpt = content[:220]
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "INSERT INTO source_documents (workspace_id, title, domain, source_type, content, excerpt, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (user["workspace_id"], filename, domain, "upload", content, excerpt, now),
        )
        conn.commit()
    return jsonify({"ok": True, "title": filename, "excerpt": excerpt})


@app.get("/api/templates")
def api_templates():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        templates = [dict(r) for r in conn.execute("SELECT * FROM templates WHERE workspace_id = ? ORDER BY id DESC", (user["workspace_id"],)).fetchall()]
    return jsonify({"ok": True, "templates": templates})


@app.post("/api/templates")
def api_templates_create():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    name = payload.get("name", "").strip()
    body = payload.get("body", "").strip()
    if not name or not body:
        return json_error("Name and body are required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "INSERT INTO templates (workspace_id, name, category, body, created_at) VALUES (?, ?, ?, ?, ?)",
            (user["workspace_id"], name, payload.get("category", "General"), body, now),
        )
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/profile")
def api_profile():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        profile = get_workspace_profile(conn, user["workspace_id"])
    return jsonify({"ok": True, "profile": profile})


@app.post("/api/profile")
def api_profile_update():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    style = payload.get("response_style", "Warm Professional").strip() or "Warm Professional"
    personality = payload.get("personality_notes", "").strip() or "Sound natural and clear."
    forbidden = payload.get("forbidden_tone", "").strip() or "Avoid robotic phrasing."
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "INSERT INTO workspace_profiles (workspace_id, response_style, personality_notes, forbidden_tone, updated_at) VALUES (?, ?, ?, ?, ?) ON CONFLICT(workspace_id) DO UPDATE SET response_style=excluded.response_style, personality_notes=excluded.personality_notes, forbidden_tone=excluded.forbidden_tone, updated_at=excluded.updated_at",
            (user["workspace_id"], style, personality, forbidden, now),
        )
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/pricing/<int:product_id>")
def api_pricing(product_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    target_margin = float(request.args.get("target_margin", 0.42))
    with closing(get_db()) as conn:
        row = conn.execute("SELECT * FROM products WHERE id = ? AND workspace_id = ?", (product_id, user["workspace_id"])).fetchone()
        if not row:
            return json_error("Product not found", 404)
        rec = pricing_recommendation(dict(row), target_margin)
    return jsonify({"ok": True, "recommendation": rec})



# ── Orchestrator: smart inbound (replaces basic webchat for agent pipeline) ───
@app.post("/api/agent/inbound")
def api_agent_inbound():
    """
    Full autonomous pipeline: Orchestrator → Sales Agent | Accounting Agent.
    Drop-in upgrade to /api/webchat/inbound.
    """
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    conversation_id = payload.get("conversation_id")
    customer_name = payload.get("customer_name", "Lead").strip() or "Lead"
    text = payload.get("text", "").strip()
    if not text:
        return json_error("Message text is required")
    now = datetime.utcnow().isoformat()

    with closing(get_db()) as conn:
        if conversation_id:
            row = conn.execute(
                "SELECT id, customer_name FROM conversations WHERE id = ? AND workspace_id = ?",
                (conversation_id, user["workspace_id"]),
            ).fetchone()
            if not row:
                return json_error("Conversation not found", 404)
            conv_id = row["id"]
            customer_name = row["customer_name"]
        else:
            cur = conn.execute(
                "INSERT INTO conversations (workspace_id, customer_name, channel, status, country, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (user["workspace_id"], customer_name, "webchat", "Active", user.get("workspace_slug","")[:2].upper(), now, now),
            )
            conv_id = cur.lastrowid
        conn.execute("INSERT INTO messages (conversation_id, role, text, created_at) VALUES (?, ?, ?, ?)", (conv_id, "customer", text, now))
        conn.commit()

    result = orchestrator_process_message(
        user["workspace_id"], conv_id, customer_name, text,
        user["language"], user["currency"],
    )
    reply = result["reply"]

    with closing(get_db()) as conn:
        conn.execute("INSERT INTO messages (conversation_id, role, text, created_at) VALUES (?, ?, ?, ?)", (conv_id, "assistant", reply, now))
        conn.execute("UPDATE conversations SET status = 'Active', updated_at = ? WHERE id = ?", (now, conv_id))
        conn.commit()

    return jsonify({
        "ok": True,
        "conversation_id": conv_id,
        "reply": reply,
        "agent": result.get("agent"),
        "action": result.get("action"),
        "data": result.get("data", {}),
    })


# ── Deals ─────────────────────────────────────────────────────────────────────
@app.get("/api/deals")
def api_deals():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        deals = [dict(r) for r in conn.execute(
            "SELECT * FROM deals WHERE workspace_id = ? ORDER BY id DESC LIMIT 50",
            (user["workspace_id"],),
        ).fetchall()]
    return jsonify({"ok": True, "deals": deals})


@app.post("/api/deals/negotiate")
def api_deals_negotiate():
    """Manually trigger the Sales Agent to build a negotiated offer."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    customer_name = payload.get("customer_name", "").strip()
    message = payload.get("message", "").strip()
    items = payload.get("items", [])
    if not customer_name or not items:
        return json_error("customer_name and items are required")
    result = sales_agent_negotiate(
        user["workspace_id"],
        payload.get("conversation_id"),
        customer_name, message, items,
        user["language"], user["currency"],
    )
    return jsonify({"ok": True, **result})


@app.post("/api/deals/<int:deal_id>/accept")
def api_deals_accept(deal_id: int):
    """Customer accepted — Accounting Agent closes the deal and issues invoice instantly."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True) or {}
    customer_tax_id = payload.get("customer_tax_id", "")
    try:
        result = accounting_agent_close_deal(user["workspace_id"], deal_id, customer_tax_id)
    except ValueError as e:
        return json_error(str(e), 400)
    return jsonify({"ok": True, **result})


@app.post("/api/deals/<int:deal_id>/reject")
def api_deals_reject(deal_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        deal = conn.execute(
            "SELECT status, conversation_id FROM deals WHERE id = ? AND workspace_id = ?",
            (deal_id, user["workspace_id"])
        ).fetchone()
        if not deal:
            return json_error("Deal not found", 404)
        if deal["status"] in ("closed", "rejected"):
            return json_error(f"Este deal ya esta en estado '{deal['status']}' y no se puede rechazar de nuevo", 409)
        conn.execute(
            "UPDATE deals SET status = 'rejected', closed_at = ? WHERE id = ? AND workspace_id = ?",
            (now, deal_id, user["workspace_id"]),
        )
        # Keep the visual Pipeline in sync — mirrors the auto-close behavior on accept
        if deal["conversation_id"]:
            conn.execute(
                "UPDATE pipeline_cards SET stage='Perdido', updated_at=? WHERE workspace_id=? AND conversation_id=?",
                (now, user["workspace_id"], deal["conversation_id"])
            )
        conn.commit()
    return jsonify({"ok": True, "status": "rejected"})


# ── Invoices ──────────────────────────────────────────────────────────────────
@app.get("/api/invoices")
def api_invoices():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        invoices = [dict(r) for r in conn.execute(
            "SELECT * FROM invoices WHERE workspace_id = ? ORDER BY id DESC LIMIT 100",
            (user["workspace_id"],),
        ).fetchall()]
    return jsonify({"ok": True, "invoices": invoices})


@app.get("/api/invoices/<int:invoice_id>")
def api_invoice_detail(invoice_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        row = conn.execute(
            "SELECT * FROM invoices WHERE id = ? AND workspace_id = ?",
            (invoice_id, user["workspace_id"]),
        ).fetchone()
        if not row:
            return json_error("Invoice not found", 404)
    return jsonify({"ok": True, "invoice": dict(row)})


# ── Agent audit trail ─────────────────────────────────────────────────────────
@app.get("/api/agents/events")
def api_agent_events():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    limit = min(int(request.args.get("limit", 50)), 200)
    with closing(get_db()) as conn:
        events = [dict(r) for r in conn.execute(
            "SELECT * FROM agent_events WHERE workspace_id = ? ORDER BY id DESC LIMIT ?",
            (user["workspace_id"], limit),
        ).fetchall()]
    return jsonify({"ok": True, "events": events})


@app.post("/api/agents/audit")
def api_agents_audit():
    """Trigger the Auditor Agent to produce a full compliance report."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    report = auditor_agent_review(user["workspace_id"])
    return jsonify({"ok": True, "report": report})


@app.post("/api/agents/events/<int:event_id>/review")
def api_agent_event_review(event_id: int):
    """Human marks an agent event as reviewed."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "UPDATE agent_events SET reviewed_at = ? WHERE id = ? AND workspace_id = ?",
            (now, event_id, user["workspace_id"]),
        )
        conn.commit()
    return jsonify({"ok": True, "reviewed_at": now})


# ── Tax profile ───────────────────────────────────────────────────────────────
@app.get("/api/tax-profile")
def api_tax_profile_get():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        profile = get_tax_profile(conn, user["workspace_id"])
    return jsonify({"ok": True, "tax_profile": profile})


@app.post("/api/tax-profile")
def api_tax_profile_update():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            """INSERT INTO tax_profiles
               (workspace_id, country_code, tax_authority, tax_id,
                default_tax_pct, fiscal_regime, invoice_prefix, next_invoice_seq, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(workspace_id) DO UPDATE SET
                 country_code = excluded.country_code,
                 tax_authority = excluded.tax_authority,
                 tax_id = excluded.tax_id,
                 default_tax_pct = excluded.default_tax_pct,
                 fiscal_regime = excluded.fiscal_regime,
                 invoice_prefix = excluded.invoice_prefix,
                 updated_at = excluded.updated_at""",
            (user["workspace_id"],
             payload.get("country_code", "US"),
             payload.get("tax_authority", ""),
             payload.get("tax_id", ""),
             float(payload.get("default_tax_pct", 0)),
             payload.get("fiscal_regime", ""),
             payload.get("invoice_prefix", "INV"),
             1,
             now),
        )
        conn.commit()
    return jsonify({"ok": True})


# ── Negotiation margins (per workspace or per product) ────────────────────────
@app.get("/api/margins")
def api_margins_get():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM negotiation_margins WHERE workspace_id = ? ORDER BY id",
            (user["workspace_id"],),
        ).fetchall()]
    if not rows:
        rows = [get_negotiation_margins(conn, user["workspace_id"])]
    return jsonify({"ok": True, "margins": rows})


@app.post("/api/margins")
def api_margins_update():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            """INSERT INTO negotiation_margins
               (workspace_id, product_id, min_margin_pct, max_discount_pct, auto_approve_below_pct, updated_at)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (user["workspace_id"],
             payload.get("product_id"),
             float(payload.get("min_margin_pct", 0.25)),
             float(payload.get("max_discount_pct", 0.15)),
             float(payload.get("auto_approve_below_pct", 0.05)),
             now),
        )
        conn.commit()
    return jsonify({"ok": True})


# ── Commission / value-based billing summary ──────────────────────────────────
@app.get("/api/commissions")
def api_commissions():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        events = [dict(r) for r in conn.execute(
            "SELECT * FROM commission_events WHERE workspace_id = ? ORDER BY id DESC LIMIT 100",
            (user["workspace_id"],),
        ).fetchall()]
        summary = conn.execute(
            """SELECT currency,
                      SUM(gross_value) as total_gross,
                      SUM(commission_amount) as total_commission,
                      SUM(CASE WHEN billed = 0 THEN commission_amount ELSE 0 END) as unbilled
               FROM commission_events WHERE workspace_id = ? GROUP BY currency""",
            (user["workspace_id"],),
        ).fetchall()
    return jsonify({
        "ok": True,
        "events": events,
        "summary": [dict(r) for r in summary],
    })



# ══════════════════════════════════════════════════════════════════════════════
# REPORTS ENGINE — Excel exports, P&L, cashflow, inventory report, sales funnel
# ══════════════════════════════════════════════════════════════════════════════

def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        return None


def _xl_header_style(ws, row, cols, fill_color="1B3A6B"):
    """Apply header style to a row in openpyxl worksheet."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="CCCCCC")
    border = Border(bottom=border_side)
    for i, col_name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col_name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _xl_data_row(ws, row, values, number_format=None):
    from openpyxl.styles import Alignment
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.alignment = Alignment(vertical="center")
        if number_format and isinstance(val, (int, float)):
            cell.number_format = number_format


@app.get("/api/reports/summary")
def api_reports_summary():
    """Returns a rich financial + sales summary for the dashboard charts."""
    try:
        user = require_auth()
        require_permission(user, "see_finances")
    except PermissionError as e:
        return json_error(str(e), 403)
    wid = user["workspace_id"]
    cur = user["currency"]
    with closing(get_db()) as conn:
        # Ledger breakdown
        ledger = [dict(r) for r in conn.execute(
            "SELECT entry_type, category, amount, state, due_date, created_at FROM ledger_entries WHERE workspace_id = ? ORDER BY id DESC",
            (wid,),
        ).fetchall()]
        # Deals breakdown
        deals = [dict(r) for r in conn.execute(
            "SELECT status, negotiated_total, discount_pct, currency, created_at FROM deals WHERE workspace_id = ? ORDER BY id DESC",
            (wid,),
        ).fetchall()]
        # Products
        products = [dict(r) for r in conn.execute(
            "SELECT name, category, cost, price, stock FROM products WHERE workspace_id = ? ORDER BY stock DESC",
            (wid,),
        ).fetchall()]
        # Invoices
        invoices = [dict(r) for r in conn.execute(
            "SELECT total, tax_amount, status, created_at FROM invoices WHERE workspace_id = ? ORDER BY id DESC",
            (wid,),
        ).fetchall()]

    # P&L calculation
    income = sum(e["amount"] for e in ledger if e["entry_type"] == "Income")
    expenses = sum(e["amount"] for e in ledger if e["entry_type"] == "Expense")
    net = income - expenses
    paid_income = sum(e["amount"] for e in ledger if e["entry_type"] == "Income" and e["state"] == "Paid")
    pending_income = income - paid_income

    # Income by category
    income_by_cat = {}
    expense_by_cat = {}
    for e in ledger:
        cat = e.get("category") or "Sin categoría"
        if e["entry_type"] == "Income":
            income_by_cat[cat] = income_by_cat.get(cat, 0) + e["amount"]
        else:
            expense_by_cat[cat] = expense_by_cat.get(cat, 0) + e["amount"]

    # Monthly trend (last 6 months)
    from collections import defaultdict
    monthly = defaultdict(lambda: {"income": 0, "expenses": 0})
    for e in ledger:
        month = (e.get("created_at") or e.get("due_date") or "")[:7]
        if month:
            if e["entry_type"] == "Income":
                monthly[month]["income"] += e["amount"]
            else:
                monthly[month]["expenses"] += e["amount"]
    monthly_sorted = sorted(monthly.items())[-6:]

    # Deal funnel
    deal_funnel = {}
    for d in deals:
        s = d["status"]
        deal_funnel[s] = deal_funnel.get(s, 0) + 1
    total_deal_value = sum(d["negotiated_total"] for d in deals if d["status"] == "closed")

    # Inventory value
    inventory_value = sum(p["cost"] * p["stock"] for p in products)
    low_stock = [p for p in products if p["stock"] <= 5]

    # Tax collected
    total_tax = sum(i["tax_amount"] for i in invoices if i["status"] in ("issued","paid"))

    return jsonify({
        "ok": True,
        "currency": cur,
        "pl": {
            "income": income, "expenses": expenses, "net": net,
            "paid_income": paid_income, "pending_income": pending_income,
            "margin_pct": round((net / income * 100) if income > 0 else 0, 1),
        },
        "income_by_category": income_by_cat,
        "expense_by_category": expense_by_cat,
        "monthly_trend": [{"month": m, **v} for m, v in monthly_sorted],
        "deals": {
            "funnel": deal_funnel,
            "total_closed_value": total_deal_value,
            "avg_discount": round(sum(d["discount_pct"] for d in deals) / len(deals), 1) if deals else 0,
        },
        "inventory": {
            "total_value": inventory_value,
            "low_stock_count": len(low_stock),
            "low_stock_items": low_stock[:5],
            "total_products": len(products),
        },
        "invoices": {
            "total_issued": len(invoices),
            "total_tax_collected": total_tax,
        },
        "ledger_entries": len(ledger),
    })


@app.get("/api/reports/ledger-explained")
def api_ledger_explained():
    """Returns the full ledger with plain-language explanations for each entry."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        entries = [dict(r) for r in conn.execute(
            "SELECT * FROM ledger_entries WHERE workspace_id = ? ORDER BY id DESC",
            (user["workspace_id"],),
        ).fetchall()]

    explained = []
    for e in entries:
        t = e["entry_type"]
        amount = e["amount"]
        cur = e.get("currency", user["currency"])
        state = e.get("state", "")
        explanation = ""
        if t == "Income" and state == "Paid":
            explanation = f"Ingreso cobrado de {format_money(amount, cur)}. Ya está en caja."
        elif t == "Income" and state == "Pending":
            explanation = f"Ingreso pendiente de cobro de {format_money(amount, cur)}. Todavía no entró el dinero."
        elif t == "Expense" and state == "Paid":
            explanation = f"Gasto pagado de {format_money(amount, cur)}. Salió de caja."
        elif t == "Expense" and state == "Pending":
            explanation = f"Gasto pendiente de pago de {format_money(amount, cur)}. Deuda por pagar."
        else:
            explanation = f"Movimiento de {format_money(amount, cur)} — {t} ({state})."
        explained.append({**e, "explanation": explanation})

    # Running balance
    balance = 0
    for e in reversed(explained):
        if e["entry_type"] == "Income":
            balance += e["amount"]
        else:
            balance -= e["amount"]
        e["running_balance"] = round(balance, 2)

    return jsonify({"ok": True, "entries": list(reversed(explained)), "final_balance": round(balance, 2)})


@app.get("/api/export/excel")
def api_export_excel():
    """Export full business data to Excel: P&L, Ledger, Inventory, Deals, Invoices."""
    try:
        user = require_auth()
        require_permission(user, "export")
    except PermissionError as e:
        return json_error(str(e), 403)
    openpyxl = _require_openpyxl()
    if not openpyxl:
        return json_error("openpyxl not installed. Run: pip install openpyxl", 500)

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    import io

    wid = user["workspace_id"]
    cur = user["currency"]
    ws_name = user["workspace_name"]
    now_str = datetime.utcnow().strftime("%Y-%m-%d")

    with closing(get_db()) as conn:
        ledger   = [dict(r) for r in conn.execute("SELECT * FROM ledger_entries WHERE workspace_id = ? ORDER BY id DESC", (wid,)).fetchall()]
        products = [dict(r) for r in conn.execute("SELECT * FROM products WHERE workspace_id = ? ORDER BY name", (wid,)).fetchall()]
        deals    = [dict(r) for r in conn.execute("SELECT * FROM deals WHERE workspace_id = ? ORDER BY id DESC", (wid,)).fetchall()]
        invoices = [dict(r) for r in conn.execute("SELECT * FROM invoices WHERE workspace_id = ? ORDER BY id DESC", (wid,)).fetchall()]

    wb = Workbook()

    # ── Sheet 1: Summary P&L ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen P&L"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 18
    _xl_header_style(ws1, 1, ["Concepto", "Monto"], fill_color="1B3A6B")
    income = sum(e["amount"] for e in ledger if e["entry_type"] == "Income")
    expenses = sum(e["amount"] for e in ledger if e["entry_type"] == "Expense")
    net = income - expenses
    rows = [
        ("Ingresos totales", income), ("Gastos totales", expenses),
        ("Resultado neto", net), ("", ""),
        ("Ingresos cobrados", sum(e["amount"] for e in ledger if e["entry_type"]=="Income" and e["state"]=="Paid")),
        ("Ingresos pendientes", sum(e["amount"] for e in ledger if e["entry_type"]=="Income" and e["state"]=="Pending")),
        ("Gastos pagados", sum(e["amount"] for e in ledger if e["entry_type"]=="Expense" and e["state"]=="Paid")),
        ("Gastos pendientes", sum(e["amount"] for e in ledger if e["entry_type"]=="Expense" and e["state"]=="Pending")),
        ("", ""), ("Valor inventario", sum(p["cost"]*p["stock"] for p in products)),
        ("Deals cerrados", sum(d["negotiated_total"] for d in deals if d["status"]=="closed")),
    ]
    for i, (label, val) in enumerate(rows, 2):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=val if isinstance(val, (int, float)) else val)
        if label in ("Resultado neto",):
            fill = PatternFill("solid", fgColor="D1FAE5" if net >= 0 else "FEE2E2")
            ws1.cell(row=i, column=1).fill = fill
            ws1.cell(row=i, column=2).fill = fill
            ws1.cell(row=i, column=2).font = Font(bold=True)

    # ── Sheet 2: Libro contable ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Libro Contable")
    for col, w in zip("ABCDEFGH", [20,12,30,18,14,12,12,12]):
        ws2.column_dimensions[get_column_letter(col.encode()[0]-64)].width = w
    cols2 = ["Fecha", "Tipo", "Concepto", "Categoría", "Monto", "Moneda", "Estado", "Vencimiento"]
    _xl_header_style(ws2, 1, cols2)
    income_fill = PatternFill("solid", fgColor="ECFDF5")
    expense_fill = PatternFill("solid", fgColor="FFF7ED")
    for i, e in enumerate(ledger, 2):
        vals = [e.get("created_at","")[:10], e.get("entry_type",""), e.get("concept",""),
                e.get("category",""), e.get("amount",0), e.get("currency", cur),
                e.get("state",""), e.get("due_date","")]
        _xl_data_row(ws2, i, vals)
        fill = income_fill if e.get("entry_type")=="Income" else expense_fill
        for col in range(1, 9):
            ws2.cell(row=i, column=col).fill = fill

    # ── Sheet 3: Inventario ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Inventario")
    for idx, w in enumerate([8,30,10,18,14,14,10,8,14,10], 1):
        ws3.column_dimensions[get_column_letter(idx)].width = w
    cols3 = ["ID","Nombre","SKU","Categoría","Costo","Precio","Stock","Mín","Valor Total","Margen %"]
    _xl_header_style(ws3, 1, cols3, fill_color="1B3A6B")
    for i, p in enumerate(products, 2):
        margin = round((p["price"]-p["cost"])/p["price"]*100, 1) if p["price"] > 0 else 0
        total_val = round(p["cost"]*p["stock"], 2)
        vals = [p["id"],p["name"],p["sku"],p.get("category",""),p["cost"],p["price"],p["stock"],p.get("stock_min",0),total_val,margin]
        _xl_data_row(ws3, i, vals)
        if p["stock"] == 0:
            ws3.cell(row=i, column=7).fill = PatternFill("solid", fgColor="FEE2E2")
        elif p["stock"] <= p.get("stock_min",0):
            ws3.cell(row=i, column=7).fill = PatternFill("solid", fgColor="FEF3C7")

    # ── Sheet 4: Deals ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Deals")
    for idx, w in enumerate([8,22,16,14,14,10,14,18], 1):
        ws4.column_dimensions[get_column_letter(idx)].width = w
    cols4 = ["ID","Cliente","Estado","Total Neg.","Original","Desc %","Moneda","Fecha"]
    _xl_header_style(ws4, 1, cols4, fill_color="4F46E5")
    for i, d in enumerate(deals, 2):
        vals = [d["id"],d["customer_name"],d["status"],d["negotiated_total"],d["original_total"],d["discount_pct"],d["currency"],d.get("created_at","")[:10]]
        _xl_data_row(ws4, i, vals)
        if d["status"] == "closed":
            ws4.cell(row=i, column=3).fill = PatternFill("solid", fgColor="ECFDF5")

    # ── Sheet 5: Facturas ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Facturas")
    for idx, w in enumerate([8,16,22,10,10,10,10,10,16], 1):
        ws5.column_dimensions[get_column_letter(idx)].width = w
    cols5 = ["ID","Número","Cliente","Subtotal","IVA %","IVA $","Total","Estado","Fecha"]
    _xl_header_style(ws5, 1, cols5, fill_color="059669")
    for i, inv in enumerate(invoices, 2):
        vals = [inv["id"],inv["invoice_number"],inv["customer_name"],inv["subtotal"],inv["tax_pct"],inv["tax_amount"],inv["total"],inv["status"],inv.get("issued_at","")[:10]]
        _xl_data_row(ws5, i, vals)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Banzai_{ws_name.replace(' ','_')}_{now_str}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.get("/api/export/ledger-csv")
def api_export_ledger_csv():
    """Export ledger as CSV."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    import csv, io
    with closing(get_db()) as conn:
        entries = [dict(r) for r in conn.execute(
            "SELECT * FROM ledger_entries WHERE workspace_id = ? ORDER BY id DESC",
            (user["workspace_id"],),
        ).fetchall()]
    buf = io.StringIO()
    if entries:
        writer = csv.DictWriter(buf, fieldnames=entries[0].keys())
        writer.writeheader()
        writer.writerows(entries)
    buf.seek(0)
    from flask import Response
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment;filename=ledger.csv"})



# ══════════════════════════════════════════════════════════════════════════════
# INDUSTRY PLAYBOOK API — CRUD for rubros, used by Banzai Admin
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/industries")
def api_industries_list():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    # Return static playbooks merged with DB overrides
    playbooks = get_playbooks(user["workspace_id"])
    result = []
    for slug, data in playbooks.items():
        result.append({"slug": slug, "name": data.get("name", slug.replace("_"," ").title()),
                       "tactics": data.get("tactics",[]), "objections": data.get("objections",{}),
                       "upsell": data.get("upsell",""), "kpis": data.get("kpis",[]),
                       "keywords": data.get("keywords",[]), "is_custom": slug not in INDUSTRY_PLAYBOOKS})
    return jsonify({"ok": True, "industries": result, "total": len(result)})


@app.post("/api/industries")
def api_industries_create():
    """Create or update an industry playbook. workspace_id=0 for global."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    slug = (payload.get("slug") or "").strip().lower().replace(" ", "_")
    name = (payload.get("name") or "").strip()
    if not slug or not name:
        return json_error("slug and name are required")
    now = datetime.utcnow().isoformat()
    data = {
        "name": name,
        "tactics":   payload.get("tactics", []),
        "objections": payload.get("objections", {}),
        "upsell":    payload.get("upsell", ""),
        "kpis":      payload.get("kpis", []),
        "keywords":  payload.get("keywords", []),
    }
    ws_id = 0  # global by default
    with closing(get_db()) as conn:
        conn.execute(
            """INSERT INTO industry_playbooks (workspace_id, slug, name, data_json, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?)
               ON CONFLICT(workspace_id, slug) DO UPDATE SET
                 name = excluded.name, data_json = excluded.data_json, updated_at = excluded.updated_at""",
            (ws_id, slug, name, json.dumps(data, ensure_ascii=False), now, now),
        )
        conn.commit()
    return jsonify({"ok": True, "slug": slug, "name": name})


@app.put("/api/industries/<slug>")
def api_industries_update(slug: str):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        row = conn.execute("SELECT id, data_json FROM industry_playbooks WHERE slug = ?", (slug,)).fetchone()
        if row:
            existing = json.loads(row["data_json"])
        else:
            existing = INDUSTRY_PLAYBOOKS.get(slug, {})
        updated_data = {
            "name":       payload.get("name", existing.get("name", slug)),
            "tactics":    payload.get("tactics", existing.get("tactics", [])),
            "objections": payload.get("objections", existing.get("objections", {})),
            "upsell":     payload.get("upsell", existing.get("upsell", "")),
            "kpis":       payload.get("kpis", existing.get("kpis", [])),
            "keywords":   payload.get("keywords", existing.get("keywords", [])),
        }
        conn.execute(
            """INSERT INTO industry_playbooks (workspace_id, slug, name, data_json, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?)
               ON CONFLICT(workspace_id, slug) DO UPDATE SET
                 name = excluded.name, data_json = excluded.data_json, updated_at = excluded.updated_at""",
            (0, slug, updated_data["name"], json.dumps(updated_data, ensure_ascii=False), now, now),
        )
        conn.commit()
    return jsonify({"ok": True, "slug": slug})


@app.delete("/api/industries/<slug>")
def api_industries_delete(slug: str):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if slug == "default":
        return json_error("Cannot delete the default industry", 400)
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM industry_playbooks WHERE slug = ?", (slug,))
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/admin/token")
def api_admin_token():
    """Generate a short-lived token for the Admin panel to authenticate."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    import hmac, hashlib, time
    secret = app.secret_key.encode() if isinstance(app.secret_key, str) else app.secret_key
    ts = str(int(time.time()))
    sig = hmac.new(secret, f"{user['workspace_id']}:{ts}".encode(), hashlib.sha256).hexdigest()[:16]
    token = f"{user['workspace_id']}:{ts}:{sig}"
    return jsonify({"ok": True, "token": token, "workspace_id": user["workspace_id"],
                    "workspace_name": user["workspace_name"], "banzai_url": "http://127.0.0.1:5000"})



# ══════════════════════════════════════════════════════════════════════════════
# BILLING — Plans · Payments · Invoices for Banzai itself (the SaaS billing)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/billing/plans")
def api_billing_plans():
    try: require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        plans = [dict(r) for r in conn.execute(
            "SELECT * FROM billing_plans WHERE active = 1 ORDER BY price ASC"
        ).fetchall()]
    return jsonify({"ok": True, "plans": plans})


@app.post("/api/billing/plans")
def api_billing_plans_create():
    try: require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    if not p.get("name") or not p.get("price"):
        return json_error("name and price are required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        cur = conn.execute(
            "INSERT INTO billing_plans (workspace_id,name,description,price,currency,interval,stripe_price_id,mp_preference_id,active,created_at) VALUES (?,?,?,?,?,?,?,?,1,?)",
            (0, p["name"], p.get("description",""), float(p["price"]), p.get("currency","USD"),
             p.get("interval","month"), p.get("stripe_price_id",""), p.get("mp_preference_id",""), now)
        )
        conn.commit()
        plan = dict(conn.execute("SELECT * FROM billing_plans WHERE id=?", (cur.lastrowid,)).fetchone())
    return jsonify({"ok": True, "plan": plan})


@app.put("/api/billing/plans/<int:plan_id>")
def api_billing_plans_update(plan_id):
    try: require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        fields, vals = [], []
        for f in ["name","description","price","currency","interval","stripe_price_id","mp_preference_id","active"]:
            if f in p:
                fields.append(f"{f}=?"); vals.append(float(p[f]) if f=="price" else p[f])
        if not fields: return json_error("No fields to update")
        vals.append(plan_id)
        conn.execute(f"UPDATE billing_plans SET {','.join(fields)} WHERE id=?", vals)
        conn.commit()
    return jsonify({"ok": True})


@app.delete("/api/billing/plans/<int:plan_id>")
def api_billing_plans_delete(plan_id):
    try: require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        conn.execute("UPDATE billing_plans SET active=0 WHERE id=?", (plan_id,))
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/billing/payments")
def api_billing_payments():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        payments = [dict(r) for r in conn.execute(
            "SELECT * FROM billing_payments WHERE workspace_id=? ORDER BY id DESC", (user["workspace_id"],)
        ).fetchall()]
    return jsonify({"ok": True, "payments": payments})


@app.post("/api/billing/payments")
def api_billing_payments_create():
    """Register a manual payment (transfer, cash, etc)."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    required = ["customer_name","customer_email","amount"]
    if any(not p.get(f) for f in required): return json_error("customer_name, customer_email, amount required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        # Generate Banzai invoice number for this SaaS payment
        tax = get_tax_profile(conn, user["workspace_id"])
        inv_num = _next_invoice_number(conn, user["workspace_id"])
        cur = conn.execute(
            """INSERT INTO billing_payments
               (workspace_id,plan_id,customer_name,customer_email,amount,currency,
                method,status,reference,notes,banzai_invoice_number,paid_at,created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (user["workspace_id"], p.get("plan_id"), p["customer_name"], p["customer_email"],
             float(p["amount"]), p.get("currency", user["currency"]),
             p.get("method","manual"), p.get("status","paid"),
             p.get("reference",""), p.get("notes",""), inv_num,
             now if p.get("status","paid")=="paid" else None, now)
        )
        payment_id = cur.lastrowid
        # Auto-post to ledger
        conn.execute(
            "INSERT INTO ledger_entries (workspace_id,entry_type,concept,category,amount,currency,state,due_date,source,created_at) VALUES (?,?,?,?,?,?,?,?,'auto',?)",
            (user["workspace_id"], "Income",
             f"Pago Banzai — {p['customer_name']} — {inv_num}",
             "Suscripción Banzai", float(p["amount"]), p.get("currency",user["currency"]),
             "Paid", now[:10], now)
        )
        conn.commit()
    return jsonify({"ok": True, "payment_id": payment_id, "invoice_number": inv_num})


@app.post("/api/billing/mercadopago/preference")
def api_mp_preference():
    """Create a MercadoPago preference (payment link)."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    mp_token = os.environ.get("MP_ACCESS_TOKEN","")
    if not mp_token: return json_error("MP_ACCESS_TOKEN not configured in .env", 400)
    p = request.get_json(force=True)
    try:
        import requests as _req
        payload = {
            "items": [{"title": p.get("title","Banzai Suscripción"), "quantity": 1,
                       "unit_price": float(p.get("amount",0)), "currency_id": p.get("currency","ARS")}],
            "payer": {"email": p.get("email","")},
            "back_urls": {"success": f"{APP_URL}/billing/success",
                          "failure": f"{APP_URL}/billing/cancel",
                          "pending": f"{APP_URL}/billing/pending"},
            "auto_return": "approved",
            "external_reference": p.get("reference",""),
        }
        r = _req.post("https://api.mercadopago.com/checkout/preferences",
                      json=payload, headers={"Authorization": f"Bearer {mp_token}"}, timeout=15)
        data = r.json()
        if r.status_code != 201:
            return json_error(f"MercadoPago error: {data.get('message','Unknown error')}", 400)
        return jsonify({"ok": True, "init_point": data["init_point"], "id": data["id"],
                        "sandbox_init_point": data.get("sandbox_init_point","")})
    except Exception as e:
        return json_error(f"MercadoPago request failed: {e}", 500)


@app.post("/api/webhooks/mercadopago")
def api_mp_webhook():
    """Receive MercadoPago payment notifications."""
    payload = request.get_json(silent=True) or {}
    now = datetime.utcnow().isoformat()
    topic = payload.get("type","") or request.args.get("topic","")
    resource_id = payload.get("data",{}).get("id","") or request.args.get("id","")
    with closing(get_db()) as conn:
        row = conn.execute("SELECT id FROM workspaces ORDER BY id LIMIT 1").fetchone()
        wid = row["id"] if row else 1
        conn.execute("INSERT INTO traces (workspace_id,flow,customer,status,detail,created_at) VALUES (?,?,?,?,?,?)",
                     (wid, "mp_webhook", topic, "received", f"id={resource_id} | {str(payload)[:200]}", now))
        conn.commit()
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# RELEASE NAMES — Name your versions
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/releases")
def api_releases():
    try: require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        releases = [dict(r) for r in conn.execute(
            "SELECT * FROM release_names ORDER BY id DESC"
        ).fetchall()]
    current = {"version": APP_VERSION, "release_name": get_release_name(APP_VERSION)}
    return jsonify({"ok": True, "releases": releases, "current": current})


@app.post("/api/releases")
def api_releases_create():
    """Name a version. If version matches APP_VERSION it becomes the current name."""
    try: require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    version = (p.get("version") or APP_VERSION).strip()
    name = (p.get("release_name") or "").strip()
    if not name: return json_error("release_name is required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "INSERT INTO release_names (version,release_name,description,created_at) VALUES (?,?,?,?) ON CONFLICT(version) DO UPDATE SET release_name=excluded.release_name, description=excluded.description",
            (version, name, p.get("description",""), now)
        )
        conn.commit()
    return jsonify({"ok": True, "version": version, "release_name": name})


# ══════════════════════════════════════════════════════════════════════════════
# FEEDBACK SURVEYS — Auto-send, collect, approve updates
# ══════════════════════════════════════════════════════════════════════════════

def _generate_token() -> str:
    import secrets
    return secrets.token_urlsafe(20)


@app.get("/api/surveys")
def api_surveys_list():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        surveys = [dict(r) for r in conn.execute(
            "SELECT * FROM feedback_surveys WHERE workspace_id=? ORDER BY id DESC",
            (user["workspace_id"],)
        ).fetchall()]
    pending_updates = sum(1 for s in surveys if s["update_status"]=="requested")
    return jsonify({"ok": True, "surveys": surveys, "pending_updates": pending_updates})


@app.post("/api/surveys/send")
def api_surveys_send():
    """Send a feedback survey to a customer (or batch)."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    recipients = p.get("recipients", [])
    if not recipients:
        # Auto-pull from recent conversations
        with closing(get_db()) as conn:
            rows = conn.execute(
                """SELECT DISTINCT customer_name,
                   COALESCE(
                     (SELECT text FROM messages WHERE conversation_id=c.id AND role='customer' LIMIT 1),
                     customer_name||'@banzai.local'
                   ) as email
                   FROM conversations c WHERE workspace_id=? ORDER BY id DESC LIMIT 20""",
                (user["workspace_id"],)
            ).fetchall()
            recipients = [{"name": r["customer_name"], "email": r["email"]} for r in rows]
    if not recipients: return json_error("No recipients found")
    now = datetime.utcnow().isoformat()
    sent = []
    with closing(get_db()) as conn:
        for rec in recipients[:50]:
            token = _generate_token()
            conn.execute(
                "INSERT INTO feedback_surveys (workspace_id,token,customer_name,customer_email,sent_at,status) VALUES (?,?,?,?,?,'sent')",
                (user["workspace_id"], token, rec.get("name","Cliente"), rec.get("email",""), now)
            )
            sent.append({"name": rec.get("name"), "token": token,
                         "survey_url": f"{APP_URL}/feedback/{token}"})
        conn.commit()
    return jsonify({"ok": True, "sent": len(sent), "surveys": sent})


@app.get("/feedback/<token>")
def survey_page(token):
    """Public survey page — no auth required."""
    with closing(get_db()) as conn:
        row = conn.execute("SELECT * FROM feedback_surveys WHERE token=?", (token,)).fetchone()
        if not row: return "Encuesta no encontrada.", 404
        if row["responded_at"]: return render_template("survey_thanks.html")
        return render_template("survey.html", token=token, name=row["customer_name"])


@app.post("/api/surveys/respond/<token>")
def api_surveys_respond(token):
    """Customer submits their survey response."""
    p = request.get_json(force=True)
    nps = p.get("nps_score")
    response_text = (p.get("response_text") or "").strip()
    suggested_update = (p.get("suggested_update") or "").strip()
    if nps is None: return json_error("nps_score is required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        row = conn.execute("SELECT * FROM feedback_surveys WHERE token=?", (token,)).fetchone()
        if not row: return json_error("Survey not found", 404)
        if row["responded_at"]: return json_error("Already responded", 400)
        update_status = "requested" if suggested_update else "none"
        conn.execute(
            """UPDATE feedback_surveys SET
               nps_score=?, response_text=?, suggested_update=?,
               responded_at=?, status='responded', update_status=?
               WHERE token=?""",
            (int(nps), response_text, suggested_update, now, update_status, token)
        )
        conn.commit()

        # Fire automations based on NPS: low scores need attention, high scores are referral opportunities
        try:
            nps_val = int(nps)
            trigger_type = "nps_low" if nps_val <= 6 else "nps_high" if nps_val >= 9 else None
            if trigger_type:
                trigger_automations(row["workspace_id"], trigger_type, {
                    "customer_name": row["customer_name"],
                    "customer_email": row["customer_email"],
                    "nps_score": nps_val,
                })
        except Exception:
            pass

    return jsonify({"ok": True, "message": "¡Gracias por tu respuesta!"})


@app.get("/api/surveys/<int:survey_id>/review")
def api_survey_review(survey_id):
    """Get a survey response ready for owner review."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        row = conn.execute(
            "SELECT * FROM feedback_surveys WHERE id=? AND workspace_id=?",
            (survey_id, user["workspace_id"])
        ).fetchone()
        if not row: return json_error("Not found", 404)
    return jsonify({"ok": True, "survey": dict(row)})


@app.post("/api/surveys/<int:survey_id>/activate-update")
def api_survey_activate_update(survey_id):
    """
    Owner reviews the suggested update and MANUALLY activates it.
    This saves the update to the knowledge base so the AI uses it.
    The owner must explicitly press this — it is never automatic.
    """
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    confirmed_update = (p.get("confirmed_update") or "").strip()
    if not confirmed_update: return json_error("confirmed_update text is required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        row = conn.execute(
            "SELECT * FROM feedback_surveys WHERE id=? AND workspace_id=?",
            (survey_id, user["workspace_id"])
        ).fetchone()
        if not row: return json_error("Survey not found", 404)
        conn.execute(
            "UPDATE feedback_surveys SET update_status='activated', update_activated_at=? WHERE id=?",
            (now, survey_id)
        )
        # Save confirmed update as a knowledge article so the AI uses it
        conn.execute(
            "INSERT INTO knowledge_articles (workspace_id,title,content,created_at) VALUES (?,?,?,?)",
            (user["workspace_id"],
             f"Actualización activada — {now[:10]}",
             f"Actualizacion solicitada por cliente y aprobada. Detalle: {confirmed_update} (Origen: feedback de {row['customer_name']} - NPS: {row['nps_score']})",
             now)
        )
        conn.commit()
    return jsonify({"ok": True, "message": "Actualización activada y agregada a la base de conocimiento del agente."})


@app.post("/api/surveys/<int:survey_id>/reject-update")
def api_survey_reject_update(survey_id):
    """Owner rejects a suggested update."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "UPDATE feedback_surveys SET update_status='rejected', update_activated_at=? WHERE id=? AND workspace_id=?",
            (now, survey_id, user["workspace_id"])
        )
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/surveys/pending-updates")
def api_surveys_pending_updates():
    """Returns surveys with suggested updates waiting for owner review."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM feedback_surveys WHERE workspace_id=? AND update_status='requested' ORDER BY id DESC",
            (user["workspace_id"],)
        ).fetchall()]
    return jsonify({"ok": True, "pending": rows, "count": len(rows)})



# ══════════════════════════════════════════════════════════════════════════════
# USER MANAGEMENT — Create sellers, manage roles, audit access
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/users")
def api_users_list():
    try:
        user = require_auth()
        require_permission(user, "manage_users")
    except PermissionError as e:
        return json_error(str(e), 403)
    with closing(get_db()) as conn:
        users = [dict(r) for r in conn.execute(
            """SELECT id, email, name, role, active, can_see_costs, can_see_finances,
                      can_export, watermark_demos, demo_expires_at,
                      last_login_at, last_login_ip, failed_login_count, created_at
               FROM users WHERE workspace_id = ? ORDER BY id DESC""",
            (user["workspace_id"],)
        ).fetchall()]
        # Remove password hashes - never expose them
        for u in users:
            u.pop("password_hash", None)
    return jsonify({"ok": True, "users": users})


@app.post("/api/users")
def api_users_create():
    """Owner creates a seller/viewer/demo account."""
    try:
        user = require_auth()
        require_permission(user, "manage_users")
    except PermissionError as e:
        return json_error(str(e), 403)

    p = request.get_json(force=True)
    name  = (p.get("name") or "").strip()
    email = (p.get("email") or "").strip().lower()
    role  = (p.get("role") or "seller").strip()
    password = (p.get("password") or "").strip()

    if not name or not email or not password:
        return json_error("name, email and password are required")
    if role not in ("seller", "viewer", "demo"):
        return json_error("role must be seller, viewer, or demo")
    if len(password) < 6:
        return json_error("Password must be at least 6 characters")

    now = datetime.utcnow().isoformat()
    demo_expires = None
    if role == "demo":
        days = int(p.get("demo_days", 30))
        from datetime import timedelta
        demo_expires = (datetime.utcnow() + timedelta(days=days)).isoformat()

    with closing(get_db()) as conn:
        existing = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
        if existing:
            return json_error("Email already registered", 409)
        conn.execute(
            """INSERT INTO users (workspace_id, email, password_hash, name, role, active,
               can_see_costs, can_see_finances, can_export, can_manage_products,
               watermark_demos, demo_expires_at, invited_by, created_at)
               VALUES (?,?,?,?,?,1,?,?,?,?,?,?,?,?)""",
            (user["workspace_id"], email, generate_password_hash(password), name, role,
             int(p.get("can_see_costs", 0)), int(p.get("can_see_finances", 0)),
             int(p.get("can_export", 0)), int(p.get("can_manage_products", 0)),
             int(p.get("watermark_demos", 1) if role in ("seller","demo") else 0),
             demo_expires, user["id"], now)
        )
        conn.commit()
        new_user = dict(conn.execute("SELECT id, email, name, role, active, demo_expires_at, created_at FROM users WHERE email=?", (email,)).fetchone())

    log_access(user["id"], user["workspace_id"], "create_user",
               resource=f"user:{email}", detail=f"role={role}")
    return jsonify({"ok": True, "user": new_user})


@app.put("/api/users/<int:target_id>")
def api_users_update(target_id: int):
    """Update a user's role or permissions."""
    try:
        user = require_auth()
        require_permission(user, "manage_users")
    except PermissionError as e:
        return json_error(str(e), 403)

    p = request.get_json(force=True)
    now = datetime.utcnow().isoformat()

    with closing(get_db()) as conn:
        target = conn.execute(
            "SELECT * FROM users WHERE id=? AND workspace_id=?",
            (target_id, user["workspace_id"])
        ).fetchone()
        if not target:
            return json_error("User not found", 404)
        if dict(target)["role"] == "owner" and user["id"] != target_id:
            return json_error("Cannot modify another owner", 403)

        fields, vals = [], []
        for f in ["name", "role", "active", "can_see_costs", "can_see_finances",
                  "can_export", "can_manage_products", "watermark_demos"]:
            if f in p:
                fields.append(f"{f}=?")
                vals.append(p[f])
        # Demo expiry update
        if "demo_days" in p:
            from datetime import timedelta
            expires = (datetime.utcnow() + timedelta(days=int(p["demo_days"]))).isoformat()
            fields.append("demo_expires_at=?")
            vals.append(expires)
        if not fields:
            return json_error("No fields to update")
        vals.append(target_id)
        conn.execute(f"UPDATE users SET {','.join(fields)} WHERE id=?", vals)
        conn.commit()
        updated = dict(conn.execute("SELECT id,email,name,role,active,can_see_costs,can_see_finances,can_export,watermark_demos,demo_expires_at FROM users WHERE id=?", (target_id,)).fetchone())

    log_access(user["id"], user["workspace_id"], "update_user",
               resource=f"user:{target_id}", detail=str(list(p.keys())))
    return jsonify({"ok": True, "user": updated})


@app.delete("/api/users/<int:target_id>")
def api_users_delete(target_id: int):
    """Deactivate (not delete) a user. Owners cannot be deactivated by others."""
    try:
        user = require_auth()
        require_permission(user, "manage_users")
    except PermissionError as e:
        return json_error(str(e), 403)

    if target_id == user["id"]:
        return json_error("Cannot deactivate your own account", 400)

    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        target = conn.execute(
            "SELECT role FROM users WHERE id=? AND workspace_id=?",
            (target_id, user["workspace_id"])
        ).fetchone()
        if not target:
            return json_error("User not found", 404)
        if dict(target)["role"] == "owner":
            return json_error("Cannot deactivate an owner account", 403)
        conn.execute("UPDATE users SET active=0 WHERE id=?", (target_id,))
        conn.commit()

    log_access(user["id"], user["workspace_id"], "deactivate_user", resource=f"user:{target_id}")
    return jsonify({"ok": True})


@app.post("/api/users/<int:target_id>/reset-password")
def api_users_reset_password(target_id: int):
    """Owner resets a user's password."""
    try:
        user = require_auth()
        require_permission(user, "manage_users")
    except PermissionError as e:
        return json_error(str(e), 403)

    p = request.get_json(force=True)
    new_password = (p.get("new_password") or "").strip()
    if len(new_password) < 6:
        return json_error("Password must be at least 6 characters")

    with closing(get_db()) as conn:
        target = conn.execute(
            "SELECT id FROM users WHERE id=? AND workspace_id=?",
            (target_id, user["workspace_id"])
        ).fetchone()
        if not target:
            return json_error("User not found", 404)
        conn.execute(
            "UPDATE users SET password_hash=?, failed_login_count=0, locked_until=NULL WHERE id=?",
            (generate_password_hash(new_password), target_id)
        )
        conn.commit()

    log_access(user["id"], user["workspace_id"], "reset_password", resource=f"user:{target_id}")
    return jsonify({"ok": True})


@app.post("/api/users/invite")
def api_users_invite():
    """Generate an invite link for a seller. They set their own password on first login."""
    try:
        user = require_auth()
        require_permission(user, "manage_users")
    except PermissionError as e:
        return json_error(str(e), 403)

    p = request.get_json(force=True)
    name = (p.get("name") or "").strip()
    email = (p.get("email") or "").strip().lower()
    role = (p.get("role") or "seller").strip()
    if not name or not email:
        return json_error("name and email required")

    import secrets
    token = secrets.token_urlsafe(32)
    now = datetime.utcnow().isoformat()
    from datetime import timedelta
    expires = (datetime.utcnow() + timedelta(days=7)).isoformat()
    demo_expires = None
    if role == "demo":
        demo_expires = (datetime.utcnow() + timedelta(days=int(p.get("demo_days",30)))).isoformat()

    with closing(get_db()) as conn:
        existing = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
        if existing:
            return json_error("Email already registered", 409)
        conn.execute(
            """INSERT INTO users (workspace_id, email, password_hash, name, role, active,
               watermark_demos, demo_expires_at, invited_by, invite_token, invite_used, created_at)
               VALUES (?,?,?,?,?,0,?,?,?,?,0,?)""",
            (user["workspace_id"], email, generate_password_hash(token), name, role,
             int(role in ("seller","demo")), demo_expires, user["id"], token, now)
        )
        conn.commit()

    invite_url = f"{APP_URL}/invite/{token}"
    log_access(user["id"], user["workspace_id"], "invite_user",
               resource=f"user:{email}", detail=f"role={role}")
    return jsonify({"ok": True, "invite_url": invite_url, "expires": expires,
                    "name": name, "email": email, "role": role})


@app.get("/invite/<token>")
def invite_page(token):
    """Invite acceptance page — user sets their own password."""
    with closing(get_db()) as conn:
        row = conn.execute("SELECT * FROM users WHERE invite_token=?", (token,)).fetchone()
        if not row or dict(row)["invite_used"]:
            return render_template("invite_invalid.html")
    return render_template("invite_accept.html", token=token,
                           name=dict(row)["name"])


@app.post("/api/invite/accept")
def api_invite_accept():
    """User accepts invite and sets their password."""
    p = request.get_json(force=True)
    token = (p.get("token") or "").strip()
    password = (p.get("password") or "").strip()
    if len(password) < 6:
        return json_error("Password must be at least 6 characters")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        row = conn.execute("SELECT * FROM users WHERE invite_token=?", (token,)).fetchone()
        if not row:
            return json_error("Invalid invite link", 404)
        row = dict(row)
        if row["invite_used"]:
            return json_error("Invite already used", 400)
        conn.execute(
            "UPDATE users SET password_hash=?, active=1, invite_used=1, invite_token=NULL WHERE id=?",
            (generate_password_hash(password), row["id"])
        )
        conn.commit()
    return jsonify({"ok": True, "message": "Account activated. You can now log in."})


@app.get("/api/access-log")
def api_access_log():
    """Owner sees full access audit trail."""
    try:
        user = require_auth()
        require_permission(user, "manage_users")
    except PermissionError as e:
        return json_error(str(e), 403)
    limit = min(int(request.args.get("limit", 100)), 500)
    with closing(get_db()) as conn:
        logs = [dict(r) for r in conn.execute(
            """SELECT al.*, u.name as user_name, u.email as user_email, u.role as user_role
               FROM access_log al
               JOIN users u ON al.user_id = u.id
               WHERE al.workspace_id = ?
               ORDER BY al.id DESC LIMIT ?""",
            (user["workspace_id"], limit)
        ).fetchall()]
    return jsonify({"ok": True, "logs": logs, "total": len(logs)})


@app.get("/api/users/<int:target_id>/activity")
def api_user_activity(target_id: int):
    """Get activity summary for a specific seller."""
    try:
        user = require_auth()
        require_permission(user, "manage_users")
    except PermissionError as e:
        return json_error(str(e), 403)

    with closing(get_db()) as conn:
        target = conn.execute(
            "SELECT id, name, email, role, last_login_at, last_login_ip FROM users WHERE id=? AND workspace_id=?",
            (target_id, user["workspace_id"])
        ).fetchone()
        if not target:
            return json_error("User not found", 404)
        logs = [dict(r) for r in conn.execute(
            "SELECT action, resource, result, detail, created_at FROM access_log WHERE user_id=? ORDER BY id DESC LIMIT 50",
            (target_id,)
        ).fetchall()]
        # Count deals they've worked on via conversations
        convs = conn.execute(
            "SELECT COUNT(*) as n FROM conversations WHERE workspace_id=?",
            (user["workspace_id"],)
        ).fetchone()["n"]

    return jsonify({
        "ok": True,
        "user": dict(target),
        "activity_log": logs,
        "conversations_in_workspace": convs,
    })



# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 1: VISUAL SALES PIPELINE (Kanban)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/pipeline")
def api_pipeline_get():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    wid = user["workspace_id"]
    with closing(get_db()) as conn:
        pipeline = conn.execute("SELECT * FROM pipelines WHERE workspace_id=? ORDER BY id LIMIT 1", (wid,)).fetchone()
        if not pipeline:
            now = datetime.utcnow().isoformat()
            cur = conn.execute(
                "INSERT INTO pipelines (workspace_id,name,stages_json,created_at) VALUES (?,?,?,?)",
                (wid,"Pipeline principal",'["Nuevo","Contactado","Demo","Propuesta","Negociación","Cerrado","Perdido"]',now)
            )
            conn.commit()
            pipeline = dict(conn.execute("SELECT * FROM pipelines WHERE id=?", (cur.lastrowid,)).fetchone())
        else:
            pipeline = dict(pipeline)
        import json as _json
        stages = _json.loads(pipeline["stages_json"])
        cards = [dict(r) for r in conn.execute(
            "SELECT * FROM pipeline_cards WHERE workspace_id=? AND pipeline_id=? ORDER BY created_at DESC",
            (wid, pipeline["id"])
        ).fetchall()]
        # Organize by stage
        board = {s: [c for c in cards if c["stage"]==s] for s in stages}
        stats = {
            "total_cards": len(cards),
            "total_value": sum(c["deal_value"] for c in cards if c["stage"] not in ("Perdido",)),
            "closed_value": sum(c["deal_value"] for c in cards if c["stage"]=="Cerrado"),
            "win_rate": round(len([c for c in cards if c["stage"]=="Cerrado"]) / max(len(cards),1) * 100, 1),
        }
    return jsonify({"ok": True, "pipeline": pipeline, "stages": stages, "board": board, "stats": stats})


@app.post("/api/pipeline/cards")
def api_pipeline_card_create():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    if not p.get("customer_name"): return json_error("customer_name required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        pipeline = conn.execute("SELECT id FROM pipelines WHERE workspace_id=? ORDER BY id LIMIT 1", (user["workspace_id"],)).fetchone()
        if not pipeline:
            return json_error("No pipeline found")
        cur = conn.execute(
            """INSERT INTO pipeline_cards
               (workspace_id,pipeline_id,conversation_id,customer_name,deal_value,currency,
                stage,assigned_to,notes,expected_close,probability,tags,created_at,updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (user["workspace_id"], pipeline["id"], p.get("conversation_id"),
             p["customer_name"], float(p.get("deal_value",0)), p.get("currency",user["currency"]),
             p.get("stage","Nuevo"), p.get("assigned_to"), p.get("notes",""),
             p.get("expected_close",""), int(p.get("probability",50)),
             p.get("tags",""), now, now)
        )
        conn.commit()
        card = dict(conn.execute("SELECT * FROM pipeline_cards WHERE id=?", (cur.lastrowid,)).fetchone())
    return jsonify({"ok": True, "card": card})


@app.put("/api/pipeline/cards/<int:card_id>")
def api_pipeline_card_update(card_id):
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        row = conn.execute("SELECT id FROM pipeline_cards WHERE id=? AND workspace_id=?", (card_id, user["workspace_id"])).fetchone()
        if not row: return json_error("Card not found", 404)
        fields, vals = [], []
        for f in ["customer_name","deal_value","stage","notes","expected_close","probability","tags","assigned_to","currency"]:
            if f in p:
                fields.append(f"{f}=?"); vals.append(float(p[f]) if f=="deal_value" else int(p[f]) if f=="probability" else p[f])
        if not fields: return json_error("No fields to update")
        vals += [now, card_id]
        conn.execute(f"UPDATE pipeline_cards SET {','.join(fields)}, updated_at=? WHERE id=?", vals)
        conn.commit()
        card = dict(conn.execute("SELECT * FROM pipeline_cards WHERE id=?", (card_id,)).fetchone())
    return jsonify({"ok": True, "card": card})


@app.delete("/api/pipeline/cards/<int:card_id>")
def api_pipeline_card_delete(card_id):
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM pipeline_cards WHERE id=? AND workspace_id=?", (card_id, user["workspace_id"]))
        conn.commit()
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 2: AUTOMATION ENGINE — IF THIS, THEN THAT
# ══════════════════════════════════════════════════════════════════════════════

AUTOMATION_TRIGGERS = {
    "deal_closed":     "Cuando se cierra un deal",
    "deal_over_value": "Cuando un deal supera un monto",
    "nps_low":         "Cuando NPS < 7",
    "nps_high":        "Cuando NPS >= 9",
    "new_conversation": "Cuando entra una conversación nueva",
    "stock_low":       "Cuando el stock baja del mínimo",
    "invoice_issued":  "Cuando se emite una factura",
    "no_reply_24h":    "Cuando una conversación sin respuesta por 24h",
}

AUTOMATION_ACTIONS = {
    "send_whatsapp":   "Enviar mensaje por WhatsApp",
    "create_task":     "Crear una tarea",
    "move_pipeline":   "Mover card en el pipeline",
    "send_survey":     "Enviar encuesta de satisfacción",
    "notify_owner":    "Notificar al dueño",
    "add_to_contacts": "Agregar a contactos",
}


@app.get("/api/automations")
def api_automations_list():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        autos = [dict(r) for r in conn.execute(
            "SELECT * FROM automations WHERE workspace_id=? ORDER BY id DESC", (user["workspace_id"],)
        ).fetchall()]
    return jsonify({"ok": True, "automations": autos,
                    "available_triggers": AUTOMATION_TRIGGERS,
                    "available_actions": AUTOMATION_ACTIONS})


@app.post("/api/automations")
def api_automations_create():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    if not p.get("name") or not p.get("trigger_type") or not p.get("action_type"):
        return json_error("name, trigger_type, action_type required")
    now = datetime.utcnow().isoformat()
    import json as _j
    with closing(get_db()) as conn:
        cur = conn.execute(
            "INSERT INTO automations (workspace_id,name,trigger_type,trigger_config,action_type,action_config,active,created_at) VALUES (?,?,?,?,?,?,1,?)",
            (user["workspace_id"], p["name"], p["trigger_type"],
             _j.dumps(p.get("trigger_config",{})), p["action_type"],
             _j.dumps(p.get("action_config",{})), now)
        )
        conn.commit()
        auto = dict(conn.execute("SELECT * FROM automations WHERE id=?", (cur.lastrowid,)).fetchone())
    return jsonify({"ok": True, "automation": auto})


@app.put("/api/automations/<int:auto_id>")
def api_automations_update(auto_id):
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    import json as _j
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        row = conn.execute("SELECT id FROM automations WHERE id=? AND workspace_id=?", (auto_id, user["workspace_id"])).fetchone()
        if not row: return json_error("Automation not found", 404)
        fields, vals = [], []
        if "name" in p: fields.append("name=?"); vals.append(p["name"])
        if "active" in p: fields.append("active=?"); vals.append(int(p["active"]))
        if "action_config" in p: fields.append("action_config=?"); vals.append(_j.dumps(p["action_config"]))
        if "trigger_config" in p: fields.append("trigger_config=?"); vals.append(_j.dumps(p["trigger_config"]))
        if fields:
            vals.append(auto_id)
            conn.execute(f"UPDATE automations SET {','.join(fields)} WHERE id=?", vals)
            conn.commit()
    return jsonify({"ok": True})


@app.delete("/api/automations/<int:auto_id>")
def api_automations_delete(auto_id):
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM automations WHERE id=? AND workspace_id=?", (auto_id, user["workspace_id"]))
        conn.commit()
    return jsonify({"ok": True})


def trigger_automations(workspace_id: int, trigger_type: str, data: dict = None):
    """Fire automations matching trigger_type for a workspace. Non-blocking."""
    import json as _j
    try:
        now = datetime.utcnow().isoformat()
        with closing(get_db()) as conn:
            autos = [dict(r) for r in conn.execute(
                "SELECT * FROM automations WHERE workspace_id=? AND trigger_type=? AND active=1",
                (workspace_id, trigger_type)
            ).fetchall()]
            for auto in autos:
                try:
                    action_cfg = _j.loads(auto.get("action_config","{}"))
                    result = "ok"
                    detail = f"trigger={trigger_type}"
                    # Execute action
                    if auto["action_type"] == "create_task":
                        conn.execute(
                            "INSERT INTO tasks (workspace_id,title,area,owner,status,priority,impact,due_label,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
                            (workspace_id, action_cfg.get("title", f"Auto: {trigger_type}"),
                             action_cfg.get("area","Auto"), "Sistema", "Today", "High", 8, "Today", now)
                        )
                    elif auto["action_type"] == "notify_owner":
                        conn.execute(
                            "INSERT INTO traces (workspace_id,flow,customer,status,detail,created_at) VALUES (?,?,?,?,?,?)",
                            (workspace_id, "automation_notify", trigger_type, "fired",
                             f"Auto '{auto['name']}' fired: {str(data)[:200]}", now)
                        )
                    elif auto["action_type"] == "send_whatsapp":
                        phone = (data or {}).get("customer_phone")
                        msg_template = action_cfg.get("message", "")
                        customer_name = (data or {}).get("customer_name", "")
                        msg_text = msg_template.replace("{{name}}", customer_name).replace("{{nombre}}", customer_name)
                        if phone and msg_text:
                            try:
                                send_whatsapp_text(phone, msg_text, workspace_id=workspace_id)
                                detail = f"WhatsApp enviado a {phone}"
                            except Exception as wa_err:
                                result = "error"
                                detail = f"Error enviando WhatsApp: {wa_err}"
                        else:
                            result = "skipped"
                            detail = "Sin telefono de cliente o mensaje vacio"
                    elif auto["action_type"] == "move_pipeline":
                        deal_id = (data or {}).get("deal_id")
                        new_status = action_cfg.get("status", "negotiating")
                        if deal_id:
                            conn.execute("UPDATE deals SET status=? WHERE id=? AND workspace_id=?", (new_status, deal_id, workspace_id))
                            detail = f"Deal {deal_id} movido a {new_status}"
                        else:
                            result = "skipped"
                            detail = "Sin deal_id en el evento"
                    elif auto["action_type"] == "send_survey":
                        customer_name = (data or {}).get("customer_name", "")
                        customer_email = (data or {}).get("customer_email", "")
                        if customer_name:
                            import secrets as _secrets_mod
                            survey_token = _secrets_mod.token_urlsafe(24)
                            conn.execute(
                                "INSERT INTO feedback_surveys (workspace_id,customer_name,customer_email,token,status,sent_at) VALUES (?,?,?,?,?,?)",
                                (workspace_id, customer_name, customer_email or "", survey_token, "sent", now)
                            )
                            detail = f"Encuesta enviada a {customer_name}"
                        else:
                            result = "skipped"
                            detail = "Sin nombre de cliente en el evento"
                    elif auto["action_type"] == "add_to_contacts":
                        customer_name = (data or {}).get("customer_name", "")
                        customer_phone = (data or {}).get("customer_phone", "")
                        if customer_name:
                            existing = conn.execute(
                                "SELECT id FROM contacts WHERE workspace_id=? AND (name=? OR phone=?)",
                                (workspace_id, customer_name, customer_phone)
                            ).fetchone()
                            if not existing:
                                conn.execute(
                                    "INSERT INTO contacts (workspace_id,name,phone,created_at) VALUES (?,?,?,?)",
                                    (workspace_id, customer_name, customer_phone, now)
                                )
                                detail = f"Contacto agregado: {customer_name}"
                            else:
                                result = "skipped"
                                detail = "Contacto ya existia"
                        else:
                            result = "skipped"
                            detail = "Sin nombre de cliente en el evento"
                    conn.execute(
                        "UPDATE automations SET run_count=run_count+1, last_run_at=? WHERE id=?",
                        (now, auto["id"])
                    )
                    conn.execute(
                        "INSERT INTO automation_runs (automation_id,workspace_id,trigger_data,result,detail,created_at) VALUES (?,?,?,?,?,?)",
                        (auto["id"], workspace_id, str(data)[:300], result, detail, now)
                    )
                except Exception as e:
                    conn.execute(
                        "INSERT INTO automation_runs (automation_id,workspace_id,trigger_data,result,detail,created_at) VALUES (?,?,?,?,?,?)",
                        (auto["id"], workspace_id, str(data)[:300], "error", str(e)[:200], now)
                    )
            conn.commit()
    except Exception:
        pass


# Wire automations into existing events
# (Called after deal close and survey response — injected at runtime)


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 3: GOALS & KPI TRACKER
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/goals")
def api_goals_list():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    wid = user["workspace_id"]
    now = datetime.utcnow().isoformat()[:10]
    with closing(get_db()) as conn:
        goals = [dict(r) for r in conn.execute(
            "SELECT * FROM goals WHERE workspace_id=? ORDER BY id DESC", (wid,)
        ).fetchall()]
        # Auto-calculate current_value from live data
        ledger_income = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as v FROM ledger_entries WHERE workspace_id=? AND entry_type='Income' AND state='Paid'",
            (wid,)
        ).fetchone()["v"]
        closed_deals = conn.execute(
            "SELECT COUNT(*) as n, COALESCE(SUM(negotiated_total),0) as v FROM deals WHERE workspace_id=? AND status='closed'",
            (wid,)
        ).fetchone()
        surveys_responded = conn.execute(
            "SELECT COUNT(*) as n FROM feedback_surveys WHERE workspace_id=? AND status='responded'",
            (wid,)
        ).fetchone()["n"]
        avg_nps = conn.execute(
            "SELECT COALESCE(AVG(nps_score),0) as v FROM feedback_surveys WHERE workspace_id=? AND nps_score IS NOT NULL",
            (wid,)
        ).fetchone()["v"]
        # Update current_value for each goal based on metric
        live_values = {
            "revenue": ledger_income,
            "deals_closed": closed_deals["n"],
            "deals_value": closed_deals["v"],
            "nps": round(avg_nps, 1),
            "surveys": surveys_responded,
        }
        for g in goals:
            live = live_values.get(g["metric"])
            if live is not None:
                g["current_value"] = live
                pct = round(live / max(g["target_value"],1) * 100, 1)
                g["progress_pct"] = min(pct, 100)
                g["on_track"] = pct >= 50
            else:
                g["progress_pct"] = round(g["current_value"] / max(g["target_value"],1) * 100, 1)
                g["on_track"] = g["progress_pct"] >= 50
    return jsonify({"ok": True, "goals": goals,
                    "available_metrics": {
                        "revenue": "Ingresos totales cobrados",
                        "deals_closed": "Deals cerrados (cantidad)",
                        "deals_value": "Valor total de deals cerrados",
                        "nps": "NPS promedio",
                        "surveys": "Encuestas respondidas",
                    }})


@app.post("/api/goals")
def api_goals_create():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    if not p.get("name") or not p.get("metric") or not p.get("target_value"):
        return json_error("name, metric, target_value required")
    now = datetime.utcnow().isoformat()
    start = p.get("start_date", now[:10])
    end = p.get("end_date", now[:7] + "-30")
    with closing(get_db()) as conn:
        cur = conn.execute(
            "INSERT INTO goals (workspace_id,name,metric,target_value,period,start_date,end_date,status,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
            (user["workspace_id"], p["name"], p["metric"], float(p["target_value"]),
             p.get("period","month"), start, end, "active", now)
        )
        conn.commit()
        goal = dict(conn.execute("SELECT * FROM goals WHERE id=?", (cur.lastrowid,)).fetchone())
    return jsonify({"ok": True, "goal": goal})


@app.put("/api/goals/<int:goal_id>")
def api_goals_update(goal_id):
    """Adjust a goal's target, period, or dates after creation."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    fields = {}
    if "name" in payload:
        fields["name"] = sanitize_text(payload["name"], max_len=200)
    if "target_value" in payload:
        if not is_positive_number(payload["target_value"], allow_zero=False):
            return json_error("El objetivo debe ser mayor a cero")
        fields["target_value"] = float(payload["target_value"])
    if "end_date" in payload:
        fields["end_date"] = payload["end_date"]
    if "status" in payload and payload["status"] in ("active", "completed", "cancelled"):
        fields["status"] = payload["status"]
    if not fields:
        return json_error("Nada para actualizar")
    with closing(get_db()) as conn:
        existing = conn.execute("SELECT id FROM goals WHERE id=? AND workspace_id=?", (goal_id, user["workspace_id"])).fetchone()
        if not existing:
            return json_error("Goal not found", 404)
        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [goal_id, user["workspace_id"]]
        conn.execute(f"UPDATE goals SET {set_clause} WHERE id = ? AND workspace_id = ?", values)
        conn.commit()
    return jsonify({"ok": True})


@app.delete("/api/goals/<int:goal_id>")
def api_goals_delete(goal_id):
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM goals WHERE id=? AND workspace_id=?", (goal_id, user["workspace_id"]))
        conn.commit()
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 4: SMART CONTACTS (360° customer profile)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/contacts")
def api_contacts_list():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        contacts = [dict(r) for r in conn.execute(
            "SELECT * FROM contacts WHERE workspace_id=? ORDER BY total_revenue DESC, id DESC",
            (user["workspace_id"],)
        ).fetchall()]
    return jsonify({"ok": True, "contacts": contacts})


@app.post("/api/contacts")
def api_contacts_create():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    if not p.get("name"): return json_error("name required")
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        cur = conn.execute(
            "INSERT INTO contacts (workspace_id,name,email,phone,company,role,tags,notes,source,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (user["workspace_id"], p["name"], p.get("email",""), p.get("phone",""),
             p.get("company",""), p.get("role",""), p.get("tags",""), p.get("notes",""),
             p.get("source","manual"), now)
        )
        conn.commit()
        contact = dict(conn.execute("SELECT * FROM contacts WHERE id=?", (cur.lastrowid,)).fetchone())
    return jsonify({"ok": True, "contact": contact})


@app.get("/api/contacts/<int:contact_id>")
def api_contact_detail(contact_id):
    """360 profile: contact + all their deals + invoices + conversations + NPS."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        contact = conn.execute(
            "SELECT * FROM contacts WHERE id=? AND workspace_id=?",
            (contact_id, user["workspace_id"])
        ).fetchone()
        if not contact: return json_error("Contact not found", 404)
        contact = dict(contact)
        # All deals by this contact
        deals = [dict(r) for r in conn.execute(
            "SELECT * FROM deals WHERE workspace_id=? AND customer_name=? ORDER BY id DESC",
            (user["workspace_id"], contact["name"])
        ).fetchall()]
        # All invoices
        invoices = [dict(r) for r in conn.execute(
            "SELECT * FROM invoices WHERE workspace_id=? AND customer_name=? ORDER BY id DESC",
            (user["workspace_id"], contact["name"])
        ).fetchall()]
        # All surveys/NPS
        surveys = [dict(r) for r in conn.execute(
            "SELECT nps_score, response_text, responded_at FROM feedback_surveys WHERE workspace_id=? AND customer_name=? ORDER BY id DESC",
            (user["workspace_id"], contact["name"])
        ).fetchall()]
        # Conversations
        conversations = [dict(r) for r in conn.execute(
            "SELECT id, channel, status, created_at FROM conversations WHERE workspace_id=? AND customer_name=? ORDER BY id DESC LIMIT 10",
            (user["workspace_id"], contact["name"])
        ).fetchall()]
        # Build 360 summary
        total_revenue = sum(inv["total"] for inv in invoices if inv["status"] in ("issued","paid"))
        avg_nps = round(sum(s["nps_score"] for s in surveys if s["nps_score"] is not None) / max(len([s for s in surveys if s["nps_score"] is not None]),1), 1) if surveys else None
        contact["total_revenue"] = total_revenue
        contact["avg_nps"] = avg_nps
        contact["deal_count"] = len(deals)
        contact["invoice_count"] = len(invoices)
    return jsonify({
        "ok": True, "contact": contact,
        "deals": deals, "invoices": invoices,
        "surveys": surveys, "conversations": conversations,
        "summary": {
            "total_revenue": total_revenue,
            "deals_closed": len([d for d in deals if d["status"]=="closed"]),
            "avg_nps": avg_nps,
            "last_contact": conversations[0]["created_at"][:10] if conversations else None,
        }
    })


@app.post("/api/contacts/sync")
def api_contacts_sync():
    """Auto-sync contacts from conversations and deals."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    now = datetime.utcnow().isoformat()
    wid = user["workspace_id"]
    created = updated = 0
    with closing(get_db()) as conn:
        # Pull unique names from conversations
        names = [r["customer_name"] for r in conn.execute(
            "SELECT DISTINCT customer_name FROM conversations WHERE workspace_id=? AND channel != 'internal'",
            (wid,)
        ).fetchall()]
        for name in names:
            existing = conn.execute("SELECT id FROM contacts WHERE workspace_id=? AND name=?", (wid, name)).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO contacts (workspace_id,name,source,created_at) VALUES (?,?,'auto_sync',?)",
                    (wid, name, now)
                )
                created += 1
            else:
                # Update conversation count
                n = conn.execute(
                    "SELECT COUNT(*) as n FROM conversations WHERE workspace_id=? AND customer_name=?",
                    (wid, name)
                ).fetchone()["n"]
                rev = conn.execute(
                    "SELECT COALESCE(SUM(total),0) as v FROM invoices WHERE workspace_id=? AND customer_name=?",
                    (wid, name)
                ).fetchone()["v"]
                conn.execute(
                    "UPDATE contacts SET conversation_count=?, total_revenue=?, last_contact_at=? WHERE workspace_id=? AND name=?",
                    (n, rev, now, wid, name)
                )
                updated += 1
        conn.commit()
    return jsonify({"ok": True, "created": created, "updated": updated})


@app.put("/api/contacts/<int:contact_id>")
def api_contact_update(contact_id):
    """Manually edit a contact's details (name, email, phone, company, role, notes, tags)."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    if "email" in payload and payload["email"] and not is_valid_email(payload["email"]):
        return json_error("El email no tiene un formato valido")
    fields = {}
    for k in ("name", "email", "phone", "company", "role", "notes", "tags"):
        if k in payload:
            fields[k] = sanitize_text(payload[k], max_len=1000) if isinstance(payload[k], str) else payload[k]
    if not fields:
        return json_error("Nada para actualizar")
    with closing(get_db()) as conn:
        existing = conn.execute("SELECT id FROM contacts WHERE id=? AND workspace_id=?", (contact_id, user["workspace_id"])).fetchone()
        if not existing:
            return json_error("Contact not found", 404)
        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [contact_id, user["workspace_id"]]
        conn.execute(f"UPDATE contacts SET {set_clause} WHERE id = ? AND workspace_id = ?", values)
        conn.commit()
    return jsonify({"ok": True})


@app.delete("/api/contacts/<int:contact_id>")
def api_contact_delete(contact_id):
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM contacts WHERE id=? AND workspace_id=?", (contact_id, user["workspace_id"]))
        conn.commit()
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 5: AI BUSINESS ADVISOR — weekly insights
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/advisor/insights")
def api_advisor_insights():
    """Generate AI-powered business insights with period-over-period comparison,
    accounts payable awareness, and structured priority-ranked output. Cached per workspace
    for 6 hours to keep it fast, cheap, and consistent instead of re-querying the model on every visit."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    wid = user["workspace_id"]
    cur_currency = user["currency"]
    force_refresh = request.args.get("refresh") == "1"

    # ── Check cache first (6 hour freshness window) ──
    if not force_refresh:
        with closing(get_db()) as conn:
            cached = conn.execute(
                "SELECT * FROM advisor_insights_cache WHERE workspace_id = ?", (wid,)
            ).fetchone()
        if cached:
            try:
                cached_age = (datetime.utcnow() - datetime.fromisoformat(cached["generated_at"])).total_seconds()
                if cached_age < 6 * 3600:
                    cached_insights = json.loads(cached["insights_json"])
                    return jsonify({
                        "ok": True,
                        "insights": cached_insights,
                        "rule_insights": cached_insights,  # backward compat with older frontend
                        "insights_text": None,
                        "context": json.loads(cached["context_json"]),
                        "source": cached["source"],
                        "generated_at": cached["generated_at"],
                        "from_cache": True,
                    })
            except Exception:
                pass

    now_dt = datetime.utcnow()
    week_ago = (now_dt - timedelta(days=7)).isoformat()
    two_weeks_ago = (now_dt - timedelta(days=14)).isoformat()

    with closing(get_db()) as conn:
        income = conn.execute("SELECT COALESCE(SUM(amount),0) as v FROM ledger_entries WHERE workspace_id=? AND entry_type='Income'", (wid,)).fetchone()["v"]
        expenses = conn.execute("SELECT COALESCE(SUM(amount),0) as v FROM ledger_entries WHERE workspace_id=? AND entry_type='Expense'", (wid,)).fetchone()["v"]

        # Period-over-period comparison: this week vs previous week
        income_this_week = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as v FROM ledger_entries WHERE workspace_id=? AND entry_type='Income' AND created_at >= ?",
            (wid, week_ago)
        ).fetchone()["v"]
        income_prior_week = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as v FROM ledger_entries WHERE workspace_id=? AND entry_type='Income' AND created_at >= ? AND created_at < ?",
            (wid, two_weeks_ago, week_ago)
        ).fetchone()["v"]
        deals_this_week = conn.execute(
            "SELECT COUNT(*) as n FROM deals WHERE workspace_id=? AND status='closed' AND closed_at >= ?",
            (wid, week_ago)
        ).fetchone()["n"]
        deals_prior_week = conn.execute(
            "SELECT COUNT(*) as n FROM deals WHERE workspace_id=? AND status='closed' AND closed_at >= ? AND closed_at < ?",
            (wid, two_weeks_ago, week_ago)
        ).fetchone()["n"]

        deals_closed = conn.execute("SELECT COUNT(*) as n, COALESCE(SUM(negotiated_total),0) as v FROM deals WHERE workspace_id=? AND status='closed'", (wid,)).fetchone()
        deals_open = conn.execute("SELECT COUNT(*) as n FROM deals WHERE workspace_id=? AND status IN ('negotiating','offer_sent')", (wid,)).fetchone()["n"]
        avg_nps = conn.execute("SELECT COALESCE(AVG(nps_score),0) as v FROM feedback_surveys WHERE workspace_id=? AND nps_score IS NOT NULL", (wid,)).fetchone()["v"]
        low_stock = conn.execute("SELECT COUNT(*) as n FROM products WHERE workspace_id=? AND stock<=stock_min AND active=1", (wid,)).fetchone()["n"]
        top_products = [dict(r) for r in conn.execute(
            "SELECT name, price, stock, demand_score FROM products WHERE workspace_id=? ORDER BY demand_score DESC LIMIT 5", (wid,)
        ).fetchall()]
        pending_updates = conn.execute("SELECT COUNT(*) as n FROM feedback_surveys WHERE workspace_id=? AND update_status='requested'", (wid,)).fetchone()["n"]
        automations_fired = conn.execute("SELECT COALESCE(SUM(run_count),0) as v FROM automations WHERE workspace_id=?", (wid,)).fetchone()["v"]

        # Accounts payable awareness (ties in the vendor invoices system)
        payable_pending = conn.execute(
            "SELECT COUNT(*) as n, COALESCE(SUM(amount),0) as v FROM vendor_invoices WHERE workspace_id=? AND status='pending'",
            (wid,)
        ).fetchone()
        payable_overdue = conn.execute(
            "SELECT COUNT(*) as n, COALESCE(SUM(amount),0) as v FROM vendor_invoices WHERE workspace_id=? AND status='pending' AND due_date IS NOT NULL AND due_date < ?",
            (wid, now_dt.strftime("%Y-%m-%d"))
        ).fetchone()

    income_trend_pct = round(((income_this_week - income_prior_week) / income_prior_week * 100), 1) if income_prior_week > 0 else None
    deals_trend = deals_this_week - deals_prior_week

    context = {
        "income": income, "expenses": expenses, "net": income - expenses,
        "income_this_week": income_this_week, "income_prior_week": income_prior_week,
        "income_trend_pct": income_trend_pct,
        "deals_this_week": deals_this_week, "deals_prior_week": deals_prior_week, "deals_trend": deals_trend,
        "deals_closed": deals_closed["n"], "deals_value": deals_closed["v"],
        "deals_open": deals_open, "avg_nps": round(avg_nps, 1),
        "low_stock_count": low_stock, "top_products": top_products,
        "pending_client_updates": pending_updates,
        "automations_run": automations_fired,
        "payable_pending_count": payable_pending["n"], "payable_pending_total": payable_pending["v"],
        "payable_overdue_count": payable_overdue["n"], "payable_overdue_total": payable_overdue["v"],
        "currency": cur_currency,
    }

    # ── Structured AI insights (JSON, not raw text) ──
    openai_key = os.environ.get("OPENAI_API_KEY", "")
    insights = None
    source = "rule_based"

    if openai_key and OpenAI:
        try:
            client_ai = OpenAI(api_key=openai_key)
            resp = client_ai.chat.completions.create(
                model=os.environ.get("OPENAI_MODEL", "gpt-4o-mini"),
                messages=[{
                    "role": "system",
                    "content": (
                        "Sos el asesor de negocios IA de Banzai, experto en PyMEs. Analizás datos reales y devolvés "
                        "insights priorizados por urgencia real, no genéricos. Prestá atención especial a: pagos "
                        "vencidos a proveedores (urgencia máxima), tendencias semana contra semana (no solo totales), "
                        "y oportunidades concretas de ingresos. "
                        "Respondé SOLO con un array JSON valido de exactamente 5 objetos, sin texto adicional, sin markdown. "
                        'Cada objeto: {"emoji": string, "title": string corto, "body": string 1-2 oraciones especificas con numeros reales, '
                        '"action": string con la accion concreta a tomar, "priority": "high"|"medium"|"low"}. '
                        "Ordená del más urgente al menos urgente."
                    )
                }, {
                    "role": "user",
                    "content": "Datos del negocio: " + json.dumps(context, ensure_ascii=False)
                }],
                max_tokens=900, temperature=0.6
            )
            raw = resp.choices[0].message.content.strip()
            if raw.startswith("```"):
                raw = "\n".join(raw.split("\n")[1:-1])
            insights = json.loads(raw)
            if isinstance(insights, dict):
                insights = insights.get("insights", [])
            source = "openai"
        except Exception as e:
            insights = None
            source = "rule_based"

    # ── Rule-based fallback (also used if AI output was malformed) ──
    if not insights or not isinstance(insights, list):
        insights = []
        if context["payable_overdue_count"] > 0:
            insights.append({"emoji":"🚨","title":f"{context['payable_overdue_count']} factura(s) vencida(s) sin pagar",
                "body": f"Le debés {format_money(context['payable_overdue_total'], cur_currency)} a proveedores en facturas ya vencidas.",
                "action":"Andá a Facturas Proveedores y priorizá esos pagos antes de que te corten el crédito.", "priority":"high"})
        if income_trend_pct is not None and income_trend_pct < -15:
            insights.append({"emoji":"📉","title":"Ingresos en baja esta semana",
                "body": f"Facturaste {income_trend_pct}% menos que la semana pasada.",
                "action":"Revisá el pipeline — puede haber deals estancados que el agente no está cerrando.", "priority":"high"})
        elif income_trend_pct is not None and income_trend_pct > 15:
            insights.append({"emoji":"📈","title":"Ingresos en alza esta semana",
                "body": f"Facturaste {income_trend_pct}% más que la semana pasada.",
                "action":"Aprovechá el momentum — mandá un broadcast a clientes inactivos ahora.", "priority":"medium"})
        if income > 0:
            margin = round((income - expenses) / income * 100, 1)
            insights.append({"emoji":"📊","title":"Margen neto","body":f"Tu margen es {margin}%.",
                "action":"Revisá la categoría de gastos más alta en Finanzas." if margin < 30 else "Margen sólido, mantenelo así.",
                "priority": "medium" if margin < 20 else "low"})
        if deals_open > 0:
            insights.append({"emoji":"🎯","title":f"{deals_open} deals abiertos",
                "body": f"Tenés {deals_open} negociaciones activas sin cerrar.",
                "action":"Revisá Deals y hacé seguimiento manual a los que llevan más de 3 días sin movimiento.", "priority":"medium"})
        if low_stock > 0:
            insights.append({"emoji":"📦","title":f"{low_stock} producto(s) con stock bajo",
                "body":"Podés perder ventas que el agente ya está negociando por falta de stock.",
                "action":"Inventario → filtrar por stock bajo y reponer.", "priority":"high"})
        if pending_updates > 0:
            insights.append({"emoji":"💡","title":f"{pending_updates} sugerencia(s) de clientes sin revisar",
                "body":"Tus clientes propusieron mejoras que todavía no evaluaste.",
                "action":"Admin → Encuestas → revisar y activar las relevantes.", "priority":"low"})
        if avg_nps > 0:
            nps_label = "Promotores" if avg_nps >= 9 else "Pasivos" if avg_nps >= 7 else "Detractores"
            insights.append({"emoji":"⭐","title":f"NPS promedio: {round(avg_nps,1)}",
                "body": f"Tu base de clientes son mayormente {nps_label}.",
                "action":"Pediles referidos activamente." if avg_nps >= 8 else "Priorizá resolver los problemas mencionados en encuestas.",
                "priority": "medium" if avg_nps < 7 else "low"})
        if not insights:
            insights.append({"emoji":"✅","title":"Todo en orden","body":"No hay alertas urgentes en este momento.",
                "action":"Seguí cargando datos para insights más precisos.", "priority":"low"})
        priority_rank = {"high": 0, "medium": 1, "low": 2}
        insights.sort(key=lambda x: priority_rank.get(x.get("priority","low"), 2))
        insights = insights[:5]

    # ── Save to cache ──
    generated_at = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        is_pg = "postgres" in os.environ.get("DATABASE_URL", "")
        if is_pg:
            conn.execute(
                "INSERT INTO advisor_insights_cache (workspace_id, insights_json, context_json, source, generated_at) VALUES (%s,%s,%s,%s,%s) "
                "ON CONFLICT(workspace_id) DO UPDATE SET insights_json=%s, context_json=%s, source=%s, generated_at=%s",
                (wid, json.dumps(insights), json.dumps(context), source, generated_at,
                 json.dumps(insights), json.dumps(context), source, generated_at)
            )
        else:
            conn.execute(
                "INSERT OR REPLACE INTO advisor_insights_cache (workspace_id, insights_json, context_json, source, generated_at) VALUES (?,?,?,?,?)",
                (wid, json.dumps(insights), json.dumps(context), source, generated_at)
            )
        conn.commit()

    return jsonify({
        "ok": True,
        "insights": insights,
        "rule_insights": insights,  # backward compat with older frontend that reads rule_insights
        "insights_text": None,
        "context": context,
        "source": source,
        "generated_at": generated_at,
        "from_cache": False,
    })


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 6: MULTI-CHANNEL BROADCAST
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/api/broadcast")
def api_broadcast():
    """Send a message to multiple contacts via WhatsApp or as internal notification."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    p = request.get_json(force=True)
    message = (p.get("message") or "").strip()
    recipients = p.get("recipients", [])
    channel = p.get("channel", "internal")
    if not message: return json_error("message required")
    if not recipients: return json_error("recipients required")
    now = datetime.utcnow().isoformat()
    wid = user["workspace_id"]
    sent = failed = 0
    with closing(get_db()) as conn:
        for rec in recipients[:200]:
            name = rec.get("name","Contact")
            phone = rec.get("phone","")
            # Personalize message
            personalized = message.replace("{{name}}", name).replace("{{nombre}}", name)
            if channel == "whatsapp" and phone:
                try:
                    send_whatsapp_text(phone, personalized, workspace_id=user["workspace_id"])
                    sent += 1
                except Exception:
                    failed += 1
            else:
                # Internal notification / trace
                conn.execute(
                    "INSERT INTO traces (workspace_id,flow,customer,status,detail,created_at) VALUES (?,?,?,?,?,?)",
                    (wid, "broadcast", name, "sent", personalized[:300], now)
                )
                sent += 1
        conn.commit()
    return jsonify({"ok": True, "sent": sent, "failed": failed,
                    "total": len(recipients), "channel": channel})


@app.get("/api/broadcast/history")
def api_broadcast_history():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM traces WHERE workspace_id=? AND flow='broadcast' ORDER BY id DESC LIMIT 100",
            (user["workspace_id"],)
        ).fetchall()]
    return jsonify({"ok": True, "history": rows})



@app.post("/api/users/quick-create")
def api_users_quick_create():
    """Owner creates a user with username@banzai84.com format instantly."""
    try:
        user = require_auth()
        require_permission(user, "manage_users")
    except PermissionError as e:
        return json_error(str(e), 403)

    p = request.get_json(force=True)
    username = (p.get("username") or "").strip().lower().replace(" ", "")
    name     = (p.get("name") or "").strip()
    password = (p.get("password") or "").strip()
    role     = (p.get("role") or "seller").strip()

    if not username or not name or not password:
        return json_error("username, name and password are required")
    if len(password) < 6:
        return json_error("Password must be at least 6 characters")
    if role not in ("seller", "viewer", "demo", "owner"):
        return json_error("Invalid role")

    email = f"{username}@banzai84.com"
    now = datetime.utcnow().isoformat()

    demo_expires = None
    if role == "demo":
        from datetime import timedelta
        demo_expires = (datetime.utcnow() + timedelta(days=int(p.get("demo_days", 30)))).isoformat()

    with closing(get_db()) as conn:
        existing = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
        if existing:
            return json_error(f"{email} already exists", 409)
        conn.execute(
            """INSERT INTO users
               (workspace_id, email, password_hash, name, role, active,
                can_see_costs, can_see_finances, can_export, can_manage_products,
                watermark_demos, demo_expires_at, invited_by, created_at)
               VALUES (?,?,?,?,?,1,?,?,?,?,?,?,?,?)""",
            (user["workspace_id"], email, generate_password_hash(password),
             name, role,
             int(p.get("can_see_costs", 0)), int(p.get("can_see_finances", 0)),
             int(p.get("can_export", 0)), int(p.get("can_manage_products", 0)),
             int(role in ("seller","demo")),
             demo_expires, user["id"], now)
        )
        conn.commit()

    log_access(user["id"], user["workspace_id"], "quick_create_user",
               resource=email, detail=f"role={role}")
    return jsonify({"ok": True, "email": email, "name": name, "role": role, "password_set": True})


@app.post("/api/simulations/run")
def api_simulations_run():
    scenarios = [
        {"email": "owner@banzai.local", "password": "demo1234", "message": "Necesito precio de lavandina 5L y entrega hoy", "customer": "Mini Mercado Sur"},
        {"email": "owner@northbridge.local", "password": "demo1234", "message": "Can you prepare a bulk quote for paper products?", "customer": "Blue Harbor Shop"},
        {"email": "owner@auroraops.local", "password": "demo1234", "message": "Vocês conseguem agendar uma reunião hoje?", "customer": "Studio Bella"},
    ]
    results = []
    client = app.test_client()
    for sc in scenarios:
        login_resp = client.post("/api/login", json={"email": sc["email"], "password": sc["password"]})
        if login_resp.status_code != 200:
            results.append({"workspace": sc["email"], "ok": False, "error": "login failed"})
            continue
        send_resp = client.post("/api/webchat/inbound", json={"customer_name": sc["customer"], "text": sc["message"]})
        if send_resp.status_code != 200:
            results.append({"workspace": sc["email"], "ok": False, "error": "inbound failed"})
            continue
        data = send_resp.get_json()
        results.append({
            "workspace": sc["email"],
            "ok": True,
            "reply": data["reply"],
            "conversation_id": data["conversation_id"],
        })
        client.post("/api/logout")
    return jsonify({"ok": True, "results": results})


@app.get("/api/integrations/status")
def api_integrations_status():
    return jsonify({"ok": True, "integrations": integration_status()})


@app.get("/api/workspace/whatsapp-number")
def api_workspace_whatsapp_get():
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    with closing(get_db()) as conn:
        row = conn.execute("SELECT whatsapp_number FROM workspaces WHERE id=?", (user["workspace_id"],)).fetchone()
    return jsonify({"ok": True, "whatsapp_number": row["whatsapp_number"] if row else None})


@app.put("/api/workspace/whatsapp-number")
def api_workspace_whatsapp_set():
    """Assign this business's own WhatsApp number so inbound webhooks route to the right workspace.
    Required for multi-tenant: each business needs a unique number registered here."""
    try: user = require_auth()
    except PermissionError: return json_error("Not authenticated", 401)
    if not user_can(user, "manage_settings"):
        return json_error("Permission denied", 403)
    payload = request.get_json(force=True)
    number = sanitize_text(payload.get("whatsapp_number", ""), max_len=30)
    if number and not _re_mod.match(r"^\+?[0-9]{8,15}$", number.replace(" ", "")):
        return json_error("El numero debe tener formato internacional, ej: +5491122334455")
    with closing(get_db()) as conn:
        if number:
            conflict = conn.execute(
                "SELECT id FROM workspaces WHERE whatsapp_number=? AND id != ?",
                (number, user["workspace_id"])
            ).fetchone()
            if conflict:
                return json_error("Ese numero ya esta asignado a otro negocio en Banzai")
        conn.execute("UPDATE workspaces SET whatsapp_number=? WHERE id=?", (number or None, user["workspace_id"]))
        conn.commit()
    return jsonify({"ok": True, "whatsapp_number": number or None})


@app.post("/api/ai/reply")
def api_ai_reply():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    text = payload.get("text", "").strip()
    if not text:
        return json_error("Message text is required")
    result = generate_ai_reply(user["workspace_id"], text, user["language"], user["currency"])
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "INSERT INTO traces (workspace_id, flow, customer, status, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (user["workspace_id"], "ai_reply", user["email"], "ok", f"Reply generated via {result['provider']}", now),
        )
        conn.commit()
    return jsonify({"ok": True, **result})


@app.get("/api/webhooks/whatsapp")
def api_whatsapp_verify():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge", "")
    if mode == "subscribe" and token and token == os.environ.get("WHATSAPP_VERIFY_TOKEN"):
        return challenge, 200
    return "forbidden", 403


@app.post("/api/webhooks/whatsapp")
def api_whatsapp_inbound():
    payload = request.get_json(silent=True) or {}
    now = datetime.utcnow().isoformat()

    # Route to the correct business by the number Meta says the message was sent to,
    # same safety logic as the Twilio webhook: never silently default when multiple
    # businesses exist and none match, to avoid cross-tenant message leaks.
    display_phone = None
    try:
        display_phone = payload.get("entry", [{}])[0].get("changes", [{}])[0].get("value", {}).get("metadata", {}).get("display_phone_number")
    except Exception:
        pass

    with closing(get_db()) as conn:
        row = None
        if display_phone:
            row = conn.execute("SELECT id, language, currency FROM workspaces WHERE whatsapp_number = ?", (display_phone,)).fetchone()
        if not row:
            all_ws = conn.execute("SELECT id, language, currency FROM workspaces LIMIT 2").fetchall()
            if len(all_ws) == 1:
                row = all_ws[0]
            else:
                _logger.error(f"Meta WhatsApp webhook: no workspace matches number {display_phone} and multiple workspaces exist — message dropped")
                return jsonify({"ok": True})
        workspace_id = row["id"] if row else 1
        language = row["language"] if row else "en"
        currency = row["currency"] if row else "USD"

        # Parse real Meta webhook payload
        sender_phone = None
        incoming_text = None
        try:
            entry = payload.get("entry", [{}])[0]
            change = entry.get("changes", [{}])[0]
            value = change.get("value", {})
            messages = value.get("messages", [])
            if messages:
                msg = messages[0]
                sender_phone = msg.get("from", "")
                if msg.get("type") == "text":
                    incoming_text = msg["text"]["body"]
        except Exception:
            pass

        conn.execute(
            "INSERT INTO traces (workspace_id, flow, customer, status, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (workspace_id, "whatsapp_webhook", sender_phone or "external", "received", json.dumps(payload)[:400], now),
        )

        # Auto-reply if we got a real message
        if sender_phone and incoming_text:
            try:
                # Find or create conversation for this phone number
                conv_row = conn.execute(
                    "SELECT id FROM conversations WHERE workspace_id = ? AND channel = 'whatsapp' AND customer_name = ?",
                    (workspace_id, sender_phone),
                ).fetchone()
                if conv_row:
                    conv_id = conv_row["id"]
                    conn.execute("UPDATE conversations SET status = ?, updated_at = ? WHERE id = ?", ("Active", now, conv_id))
                else:
                    cur = conn.execute(
                        "INSERT INTO conversations (workspace_id, customer_name, channel, status, country, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (workspace_id, sender_phone, "whatsapp", "New inquiry", "WA", now, now),
                    )
                    conv_id = cur.lastrowid

                conn.execute("INSERT INTO messages (conversation_id, role, text, created_at) VALUES (?, ?, ?, ?)", (conv_id, "customer", incoming_text, now))
                orch_result = orchestrator_process_message(workspace_id, conv_id, sender_phone, incoming_text, language, currency)
                reply_text = orch_result["reply"]
                conn.execute("INSERT INTO messages (conversation_id, role, text, created_at) VALUES (?, ?, ?, ?)", (conv_id, "assistant", reply_text, now))
                conn.execute(
                    "INSERT INTO traces (workspace_id, flow, customer, status, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (workspace_id, "whatsapp_autoreply", sender_phone, "ok", f"Reply via {ai_payload.get('provider')}", now),
                )
                conn.commit()
                # Send the reply back via WhatsApp
                try:
                    send_whatsapp_text(sender_phone, reply_text, workspace_id=workspace_id)
                except Exception as send_exc:
                    with closing(get_db()) as conn2:
                        conn2.execute(
                            "INSERT INTO traces (workspace_id, flow, customer, status, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                            (workspace_id, "whatsapp_send_error", sender_phone, "error", str(send_exc)[:300], now),
                        )
                        conn2.commit()
            except Exception as exc:
                conn.execute(
                    "INSERT INTO traces (workspace_id, flow, customer, status, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (workspace_id, "whatsapp_process_error", sender_phone or "unknown", "error", str(exc)[:300], now),
                )
                conn.commit()
        else:
            conn.commit()
    return jsonify({"ok": True, "received": True})


@app.post("/api/whatsapp/send-test")
def api_whatsapp_send_test():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    payload = request.get_json(force=True)
    to = payload.get("to", "").strip()
    text = payload.get("text", "").strip()
    if not to or not text:
        return json_error("Both 'to' and 'text' are required")
    try:
        result = send_whatsapp_text(to, text, workspace_id=user["workspace_id"])
        return jsonify({"ok": True, "result": result})
    except Exception as exc:
        return json_error(f"WhatsApp send failed: {exc}", 400)


@app.post("/api/billing/create-checkout-session")
def api_billing_checkout_session():
    try:
        require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    status = integration_status()
    if not status["stripe"]["configured"]:
        return json_error("Stripe integration is not configured", 400)
    payload = request.get_json(force=True)
    email = payload.get("email", "").strip()
    price_id = payload.get("price_id", "").strip()
    if not email or not price_id:
        return json_error("Both 'email' and 'price_id' are required")
    try:
        session_obj = stripe.checkout.Session.create(
            mode="subscription",
            customer_email=email,
            line_items=[{"price": price_id, "quantity": 1}],
            success_url=f"{APP_URL}/billing/success",
            cancel_url=f"{APP_URL}/billing/cancel",
        )
        return jsonify({"ok": True, "url": session_obj.url, "id": session_obj.id})
    except Exception as exc:
        return json_error(f"Stripe checkout failed: {exc}", 400)


@app.post("/api/webhooks/stripe")
def api_stripe_webhook():
    if stripe is None:
        return json_error("Stripe library not installed", 400)
    secret = os.environ.get("STRIPE_WEBHOOK_SECRET")
    payload = request.data
    sig_header = request.headers.get("Stripe-Signature", "")
    try:
        if secret:
            event = stripe.Webhook.construct_event(payload, sig_header, secret)
        else:
            event = json.loads(payload.decode("utf-8") or "{}")
    except Exception as exc:
        return json_error(f"Invalid webhook: {exc}", 400)

    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        row = conn.execute("SELECT id FROM workspaces ORDER BY id LIMIT 1").fetchone()
        workspace_id = row["id"] if row else 1
        conn.execute(
            "INSERT INTO traces (workspace_id, flow, customer, status, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (workspace_id, "stripe_webhook", "billing", "received", json.dumps(event)[:400], now),
        )
        conn.commit()
    return jsonify({"ok": True, "event_type": event.get("type") if isinstance(event, dict) else str(type(event))})


@app.get("/api/backend/blueprint")
def api_backend_blueprint():
    return jsonify(
        {
            "ok": True,
            "layers": [
                {"layer": "Auth", "purpose": "Handles login, sessions and roles"},
                {"layer": "API", "purpose": "Exposes conversations, products, knowledge, templates, tasks, ledger and quote actions"},
                {"layer": "Database", "purpose": "Stores business data in SQLite"},
                {"layer": "AI service", "purpose": "Creates humanized replies, uses expert source snippets and pricing recommendations"},
                {"layer": "Source library", "purpose": "Stores uploaded and curated knowledge sources for stronger answers"},
                {"layer": "Channel integration", "purpose": "Processes a real webchat channel and is ready for WhatsApp webhook connection"},
                {"layer": "External AI", "purpose": "Can call OpenAI from the backend when OPENAI_API_KEY is configured"},
                {"layer": "Billing", "purpose": "Can create Stripe checkout sessions when Stripe credentials are configured"},
                {"layer": "Health + version", "purpose": "Reports app status and installed edition"},
                {"layer": "Backup", "purpose": "Lets the user download a safe snapshot of the local database"},
                {"layer": "Go-live pack", "purpose": "Bundles positioning, pricing, rollout and sales docs for commercialization"},
            ],
            "routes": [
                {"method": "POST", "route": "/api/login"},
                {"method": "GET", "route": "/api/dashboard"},
                {"method": "POST", "route": "/api/messages/reply"},
                {"method": "POST", "route": "/api/webchat/inbound"},
                {"method": "POST", "route": "/api/quotes"},
                {"method": "POST", "route": "/api/tasks"},
                {"method": "POST", "route": "/api/ledger"},
                {"method": "POST", "route": "/api/profile"},
                {"method": "POST", "route": "/api/sources"},
                {"method": "POST", "route": "/api/sources/upload"},
                {"method": "GET", "route": "/api/integrations/status"},
                {"method": "POST", "route": "/api/ai/reply"},
                {"method": "GET", "route": "/api/webhooks/whatsapp"},
                {"method": "POST", "route": "/api/webhooks/whatsapp"},
                {"method": "POST", "route": "/api/whatsapp/send-test"},
                {"method": "POST", "route": "/api/billing/create-checkout-session"},
                {"method": "POST", "route": "/api/webhooks/stripe"},
                {"method": "GET", "route": "/api/go-live-pack"},
                {"method": "GET", "route": "/api/backup"},
            ],
        }
    )


@app.get("/api/go-live-pack")
def api_go_live_pack():
    pack_dir = BASE_DIR / "go_live_pack"
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for file in pack_dir.glob("*.md"):
            zf.write(file, arcname=f"go_live_pack/{file.name}")
    buffer.seek(0)
    filename = f"banzai-go-live-pack-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.zip"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/zip")


@app.get("/api/backup")
def api_backup():
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(DB_PATH, arcname="banzai.db")
    buffer.seek(0)
    filename = f"banzai-backup-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.zip"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/zip")


@app.get("/api/health")
def api_health():
    return jsonify({"ok": True, "status": "healthy", "app": APP_NAME, "version": APP_VERSION})


@app.get("/api/version")
def api_version():
    return jsonify({
        "ok": True,
        "app": APP_NAME,
        "version": APP_VERSION,
        "release_name": get_release_name(APP_VERSION),
        "edition": APP_EDITION,
        "updatable": True,
        "installable": True,
        "integrations": integration_status(),
    })


@app.get("/feedback/thanks")
def survey_thanks():
    return render_template("survey_thanks.html")


@app.get("/manifest.webmanifest")
def manifest():
    return send_from_directory(app.static_folder, "manifest.webmanifest", mimetype="application/manifest+json")


@app.get("/service-worker.js")
def service_worker():
    return send_from_directory(app.static_folder, "service-worker.js", mimetype="application/javascript")


init_db()
auto_setup_if_needed()


# ═══════════════════════════════════════════════════════════════════════════════
# BANZAI ADMIN PANEL — embedded at /bz-admin/
# Access: /bz-admin/ with ADMIN_USERNAME / ADMIN_PASSWORD
# ═══════════════════════════════════════════════════════════════════════════════

BZ_ADMIN_HTML = '<!doctype html>\n<html lang="es">\n<head>\n  <meta charset="utf-8">\n  <meta name="viewport" content="width=device-width,initial-scale=1">\n  <title>Banzai Admin</title>\n  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">\n  <style>\n    *{box-sizing:border-box;margin:0;padding:0;}\n    :root{\n      --bg:#06080F;--surface:#0C0F1A;--surface2:#111520;--surface3:#161B2E;\n      --border:rgba(255,255,255,.07);--border2:rgba(255,255,255,.12);\n      --text:#F0F2FA;--text2:#9AA3C2;--text3:#4E5B7A;\n      --brand:#6366F1;--brand-dim:rgba(99,102,241,.12);--brand-dark:#4F46E5;\n      --green:#10B981;--green-dim:rgba(16,185,129,.12);\n      --amber:#F59E0B;--amber-dim:rgba(245,158,11,.12);\n      --red:#EF4444;--red-dim:rgba(239,68,68,.12);\n      --r:10px;--r-lg:16px;\n    }\n    body{font-family:\'Inter\',sans-serif;background:var(--bg);color:var(--text);display:flex;min-height:100vh;font-size:14px;-webkit-font-smoothing:antialiased;}\n    ::-webkit-scrollbar{width:3px;} ::-webkit-scrollbar-thumb{background:var(--surface3);}\n    .sidebar{width:220px;flex-shrink:0;background:var(--surface);border-right:1px solid var(--border);display:flex;flex-direction:column;overflow-y:auto;position:fixed;top:0;left:0;bottom:0;}\n    .s-logo{padding:18px 16px 14px;display:flex;align-items:center;gap:10px;border-bottom:1px solid var(--border);}\n    .s-mark{width:32px;height:32px;background:var(--brand);border-radius:9px;display:flex;align-items:center;justify-content:center;font-weight:900;font-size:11px;color:#fff;box-shadow:0 0 12px rgba(99,102,241,.3);}\n    .s-name{font-size:15px;font-weight:800;letter-spacing:-.02em;}\n    .s-tag{font-size:9px;font-weight:700;background:var(--brand-dim);color:var(--brand);padding:2px 6px;border-radius:4px;letter-spacing:.06em;text-transform:uppercase;margin-left:4px;}\n    .s-group{padding:14px 14px 4px;font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;color:var(--text3);}\n    .nav-btn{display:flex;align-items:center;gap:9px;padding:9px 14px;border-radius:8px;cursor:pointer;color:var(--text2);font-size:13px;font-weight:500;transition:all .1s;margin:1px 6px;border:none;background:transparent;width:calc(100% - 12px);text-align:left;}\n    .nav-btn:hover{background:var(--surface2);color:var(--text);}\n    .nav-btn.active{background:var(--brand-dim);color:var(--brand);border:1px solid rgba(99,102,241,.2);}\n    .nav-icon{font-size:14px;width:16px;text-align:center;flex-shrink:0;}\n    .s-bottom{margin-top:auto;padding:10px;border-top:1px solid var(--border);}\n    .s-user{display:flex;align-items:center;gap:8px;padding:8px;}\n    .s-av{width:28px;height:28px;border-radius:8px;background:var(--brand-dim);display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:var(--brand);}\n    .s-info{flex:1;} .s-uname{font-size:12px;font-weight:600;} .s-urole{font-size:10px;color:var(--text3);}\n    .main{flex:1;margin-left:220px;overflow-y:auto;min-height:100vh;}\n    .page{display:none;padding:28px 32px;}\n    .page.active{display:block;}\n    .page-header{display:flex;align-items:flex-start;justify-content:space-between;margin-bottom:24px;}\n    .page-title{font-size:22px;font-weight:700;letter-spacing:-.02em;}\n    .page-sub{font-size:12px;color:var(--text3);margin-top:3px;}\n    .header-actions{display:flex;gap:8px;}\n    .panel{background:var(--surface);border:1px solid var(--border);border-radius:var(--r-lg);overflow:hidden;margin-bottom:14px;}\n    .panel-head{display:flex;align-items:center;justify-content:space-between;padding:13px 18px;border-bottom:1px solid var(--border);}\n    .panel-head h3{font-size:12px;font-weight:600;}\n    .panel-body{padding:16px 18px;}\n    .g2{display:grid;grid-template-columns:1fr 1fr;gap:14px;}\n    .g3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;}\n    .kpi-row{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:20px;}\n    .kpi{background:var(--surface);border:1px solid var(--border);border-radius:var(--r-lg);padding:14px 16px;position:relative;overflow:hidden;}\n    .kpi::before{content:\'\';position:absolute;top:0;left:0;right:0;height:2px;background:var(--kc,var(--brand));}\n    .kpi.green{--kc:var(--green);} .kpi.red{--kc:var(--red);} .kpi.amber{--kc:var(--amber);}\n    .kpi-label{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:var(--text3);margin-bottom:8px;}\n    .kpi-val{font-size:20px;font-weight:700;}\n    .badge{display:inline-flex;align-items:center;gap:3px;padding:2px 7px;border-radius:99px;font-size:10px;font-weight:700;}\n    .badge::before{content:\'\';width:5px;height:5px;border-radius:50%;background:currentColor;flex-shrink:0;}\n    .badge.green{background:var(--green-dim);color:var(--green);}\n    .badge.blue{background:var(--brand-dim);color:var(--brand);}\n    .badge.amber{background:var(--amber-dim);color:var(--amber);}\n    .badge.red{background:var(--red-dim);color:var(--red);}\n    .badge.gray{background:var(--surface3);color:var(--text3);}\n    .empty{display:flex;flex-direction:column;align-items:center;justify-content:center;padding:36px;text-align:center;color:var(--text3);gap:8px;}\n    button.primary{background:var(--brand);color:#fff;border:none;border-radius:8px;padding:9px 18px;font:inherit;font-size:13px;font-weight:600;cursor:pointer;transition:all .15s;display:inline-flex;align-items:center;gap:6px;}\n    button.primary:hover{background:var(--brand-dark);}\n    button.primary:disabled{opacity:.5;cursor:not-allowed;}\n    button.secondary{background:var(--surface3);color:var(--text2);border:1px solid var(--border2);border-radius:8px;padding:8px 14px;font:inherit;font-size:12px;font-weight:500;cursor:pointer;transition:all .15s;}\n    button.secondary:hover{border-color:var(--brand);color:var(--brand);}\n    button.ghost{background:transparent;color:var(--text3);border:1px solid var(--border);border-radius:8px;padding:7px 12px;font:inherit;font-size:12px;cursor:pointer;}\n    button.ghost:hover{color:var(--text2);border-color:var(--border2);}\n    button.danger{background:var(--red-dim);color:var(--red);border:1px solid rgba(239,68,68,.2);border-radius:8px;padding:7px 12px;font:inherit;font-size:12px;cursor:pointer;}\n    label{display:block;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:var(--text3);margin-bottom:4px;margin-top:12px;}\n    input,textarea,select{width:100%;background:var(--surface3);border:1px solid var(--border2);border-radius:8px;color:var(--text);font:inherit;font-size:13px;padding:9px 12px;outline:none;transition:border-color .15s;}\n    input:focus,textarea:focus,select:focus{border-color:var(--brand);}\n    textarea{resize:vertical;min-height:72px;}\n    .row{display:flex;align-items:center;justify-content:space-between;padding:9px 0;border-bottom:1px solid var(--border);font-size:12px;}\n    .row:last-child{border-bottom:none;}\n    #toast{position:fixed;bottom:20px;right:20px;z-index:999;display:flex;flex-direction:column;gap:6px;}\n    .t-item{background:var(--surface);border:1px solid var(--border2);border-radius:10px;padding:10px 14px;font-size:12px;font-weight:500;box-shadow:0 8px 32px rgba(0,0,0,.6);display:flex;align-items:center;gap:8px;}\n    .t-item.ok{border-color:rgba(16,185,129,.3);color:var(--green);}\n    .t-item.err{border-color:rgba(239,68,68,.3);color:var(--red);}\n    .chat-wrap{display:flex;flex-direction:column;height:calc(100vh - 200px);}\n    .chat-msgs{flex:1;overflow-y:auto;padding:16px;display:flex;flex-direction:column;gap:12px;}\n    .bubble{max-width:80%;padding:12px 16px;border-radius:14px;font-size:13px;line-height:1.6;}\n    .bubble.user{background:var(--brand);color:#fff;align-self:flex-end;border-bottom-right-radius:4px;}\n    .bubble.assistant{background:var(--surface3);border:1px solid var(--border2);align-self:flex-start;border-bottom-left-radius:4px;white-space:pre-wrap;}\n    .bubble.assistant pre{background:var(--surface);border-radius:6px;padding:10px;font-size:11px;margin-top:8px;overflow-x:auto;}\n    .chat-input-row{padding:12px;border-top:1px solid var(--border);display:flex;gap:8px;}\n    .chat-input-row textarea{min-height:52px;flex:1;}\n    .quick-btns{display:flex;gap:6px;flex-wrap:wrap;padding:8px 12px 0;}\n    .ind-card{background:var(--surface2);border:1px solid var(--border);border-radius:var(--r-lg);padding:14px;margin-bottom:8px;}\n    .ind-card-head{display:flex;align-items:center;gap:8px;margin-bottom:6px;}\n    .ind-name{font-size:13px;font-weight:600;flex:1;}\n    .ind-slug{font-size:10px;color:var(--text3);background:var(--surface3);padding:2px 6px;border-radius:4px;font-family:monospace;}\n    .chip{display:inline-flex;align-items:center;background:var(--surface3);border:1px solid var(--border);border-radius:5px;padding:2px 7px;font-size:11px;color:var(--text2);margin:2px;}\n  </style>\n</head>\n<body>\n\n<nav class="sidebar">\n  <div class="s-logo">\n    <div class="s-mark">BZ</div>\n    <div class="s-name">Banzai<span class="s-tag">Admin</span></div>\n  </div>\n  <div class="s-group">Dashboard</div>\n  <button class="nav-btn active" onclick="go(\'overview\')"><span class="nav-icon">⬡</span> Overview</button>\n  <button class="nav-btn" onclick="go(\'chat\')"><span class="nav-icon">✦</span> IA Experta</button>\n  <div class="s-group">Gestión</div>\n  <button class="nav-btn" onclick="go(\'industries\')"><span class="nav-icon">◈</span> Rubros del Agente</button>\n  <button class="nav-btn" onclick="go(\'users\')"><span class="nav-icon">👥</span> Vendedores</button>\n  <button class="nav-btn" onclick="go(\'billing\')"><span class="nav-icon">💳</span> Facturación</button>\n  <button class="nav-btn" onclick="go(\'surveys\')"><span class="nav-icon">💬</span> Encuestas</button>\n  <button class="nav-btn" onclick="go(\'releases\')"><span class="nav-icon">✦</span> Versiones</button>\n  <button class="nav-btn" onclick="go(\'updates\')"><span class="nav-icon">⚡</span> Actualizaciones</button>\n  <button class="nav-btn" onclick="go(\'reports\')"><span class="nav-icon">◉</span> Reportes</button>\n  <button class="nav-btn" onclick="go(\'events\')"><span class="nav-icon">◎</span> Eventos</button>\n  <div class="s-bottom">\n    <div class="s-user">\n      <div class="s-av">SA</div>\n      <div class="s-info">\n        <div class="s-uname">Super Admin</div>\n        <div class="s-urole">superadmin@banzai84.com</div>\n      </div>\n      <button class="ghost" onclick="logout()" style="padding:4px 8px;font-size:11px;">↩</button>\n    </div>\n  </div>\n</nav>\n\n<main class="main">\n\n  <!-- Overview -->\n  <div class="page active" id="page-overview">\n    <div class="page-header">\n      <div><div class="page-title">Overview</div><div class="page-sub">Panel de control de Banzai84</div></div>\n      <div class="header-actions">\n        <button class="ghost" onclick="loadOverview()">↺ Actualizar</button>\n        <button class="primary" onclick="go(\'chat\')">✦ IA Experta</button>\n      </div>\n    </div>\n    <div id="overview-status" style="margin-bottom:16px;"></div>\n    <div class="kpi-row" id="kpi-row"><div class="kpi"><div class="kpi-label">Cargando...</div><div class="kpi-val">—</div></div></div>\n    <div class="g2">\n      <div class="panel">\n        <div class="panel-head"><h3>Últimos eventos de agentes</h3><button class="ghost" onclick="loadEvents()" style="font-size:11px;">↺</button></div>\n        <div class="panel-body" style="max-height:280px;overflow-y:auto;" id="events-list"><div class="empty"><div>Sin eventos</div></div></div>\n      </div>\n      <div class="panel">\n        <div class="panel-head"><h3>Rubros activos del Sales Agent</h3><span class="badge blue" id="ind-badge">—</span></div>\n        <div class="panel-body" style="max-height:280px;overflow-y:auto;" id="ind-overview"><div class="empty"><div>Cargando...</div></div></div>\n      </div>\n    </div>\n  </div>\n\n  <!-- IA Chat -->\n  <div class="page" id="page-chat">\n    <div class="page-header">\n      <div><div class="page-title">IA Experta</div><div class="page-sub">Experta en Banzai y todos los aspectos de negocios</div></div>\n    </div>\n    <div class="panel" style="height:calc(100vh - 180px);display:flex;flex-direction:column;">\n      <div class="quick-btns">\n        <button class="ghost" style="font-size:11px;" onclick="ask(\'¿Cuántos rubros tengo y cuáles son?\')">¿Qué rubros tengo?</button>\n        <button class="ghost" style="font-size:11px;" onclick="ask(\'Creá un rubro para veterinarias con tácticas y objeciones\')">+ Veterinaria</button>\n        <button class="ghost" style="font-size:11px;" onclick="ask(\'¿Cómo está el negocio? Analizá los KPIs\')">Analizar KPIs</button>\n        <button class="ghost" style="font-size:11px;" onclick="ask(\'Dame el mejor script para cerrar una distribuidora\')">Script distribuidora</button>\n        <button class="ghost" style="font-size:11px;" onclick="ask(\'¿Cómo llego a $10.000 USD de MRR con Banzai?\')">Estrategia $10k MRR</button>\n      </div>\n      <div class="chat-msgs" id="chat-msgs"></div>\n      <div class="chat-input-row">\n        <textarea id="chat-input" placeholder="Preguntá cualquier cosa..." onkeydown="if(event.key===\'Enter\'&&!event.shiftKey){event.preventDefault();sendMsg();}"></textarea>\n        <button class="primary" onclick="sendMsg()" style="align-self:flex-end;">Enviar ↑</button>\n      </div>\n    </div>\n  </div>\n\n  <!-- Industries -->\n  <div class="page" id="page-industries">\n    <div class="page-header">\n      <div><div class="page-title">Rubros del Agente</div><div class="page-sub">Editá tácticas, objeciones y keywords</div></div>\n      <div class="header-actions">\n        <button class="primary" onclick="openIndForm()">+ Nuevo rubro</button>\n      </div>\n    </div>\n    <div class="panel" id="ind-form" style="display:none;margin-bottom:16px;">\n      <div class="panel-head"><h3 id="ind-form-title">Nuevo rubro</h3><button class="ghost" onclick="closeIndForm()" style="font-size:11px;">Cancelar</button></div>\n      <div class="panel-body">\n        <div class="g2">\n          <div>\n            <label>Slug</label><input id="f-slug" placeholder="ej: veterinaria">\n            <label>Nombre</label><input id="f-name" placeholder="ej: Veterinaria">\n            <label>Upsell</label><input id="f-upsell" placeholder="ej: El pack anual ahorra 20%">\n          </div>\n          <div>\n            <label>Tácticas (una por línea)</label><textarea id="f-tactics" placeholder="precio por volumen"></textarea>\n            <label>KPIs (una por línea)</label><textarea id="f-kpis" style="min-height:56px;" placeholder="ticket promedio"></textarea>\n          </div>\n        </div>\n        <label>Keywords (separadas por coma)</label>\n        <input id="f-keywords" placeholder="veterinaria, mascota, perro">\n        <label>Objeciones (objeción: respuesta — una por línea)</label>\n        <textarea id="f-objections" placeholder="caro: ¿Cuánto cuesta no tenerlo?"></textarea>\n        <div style="display:flex;gap:8px;margin-top:14px;justify-content:flex-end;">\n          <button class="ghost" onclick="closeIndForm()">Cancelar</button>\n          <button class="primary" onclick="saveIndustry()">Guardar rubro</button>\n        </div>\n      </div>\n    </div>\n    <div id="ind-list"><div class="empty"><div>Cargando rubros...</div></div></div>\n  </div>\n\n  <!-- Users -->\n  <div class="page" id="page-users">\n    <div class="page-header">\n      <div><div class="page-title">Vendedores</div><div class="page-sub">Creá usuarios @banzai84.com · Controlá permisos</div></div>\n    </div>\n    <div class="g2">\n      <div class="panel">\n        <div class="panel-head"><h3>Usuarios activos</h3><button class="ghost" onclick="loadUsers()" style="font-size:11px;">↺</button></div>\n        <div class="panel-body" id="users-list"><div class="empty"><div>Cargando...</div></div></div>\n      </div>\n      <div class="panel">\n        <div class="panel-head"><h3>Crear usuario @banzai84.com</h3></div>\n        <div class="panel-body">\n          <label>Nombre completo</label><input id="qc-name" placeholder="Carlos García">\n          <label>Username</label><input id="qc-user" placeholder="carlos.garcia">\n          <label>Contraseña</label><input id="qc-pass" type="password">\n          <label>Rol</label>\n          <select id="qc-role">\n            <option value="seller">Vendedor</option>\n            <option value="viewer">Visor</option>\n            <option value="demo">Demo temporal</option>\n          </select>\n          <div id="qc-result" style="display:none;margin-top:10px;padding:10px;border-radius:8px;font-size:12px;"></div>\n          <div style="margin-top:12px;"><button class="primary" style="width:100%;justify-content:center;" onclick="quickCreate()">Crear usuario</button></div>\n        </div>\n      </div>\n    </div>\n  </div>\n\n  <!-- Billing -->\n  <div class="page" id="page-billing">\n    <div class="page-header">\n      <div><div class="page-title">Facturación</div><div class="page-sub">Planes · Pagos · Historial</div></div>\n      <div class="header-actions"><button class="primary" onclick="openPlanForm()">+ Nuevo plan</button></div>\n    </div>\n    <div class="g2">\n      <div class="panel">\n        <div class="panel-head"><h3>Planes activos</h3></div>\n        <div class="panel-body" id="plans-list"><div class="empty"><div>Sin planes</div></div></div>\n      </div>\n      <div class="panel" id="plan-form">\n        <div class="panel-head"><h3>Nuevo plan</h3></div>\n        <div class="panel-body">\n          <label>Nombre</label><input id="plan-name" placeholder="Ej: Banzai Growth">\n          <label>Descripción</label><input id="plan-desc">\n          <div class="g2" style="gap:10px;">\n            <div><label>Precio</label><input id="plan-price" type="number"></div>\n            <div><label>Moneda</label><select id="plan-cur"><option>USD</option><option>ARS</option></select></div>\n          </div>\n          <label>Período</label><select id="plan-int"><option value="month">Mensual</option><option value="year">Anual</option></select>\n          <div style="margin-top:12px;"><button class="primary" style="width:100%;justify-content:center;" onclick="savePlan()">Guardar plan</button></div>\n        </div>\n      </div>\n    </div>\n    <div class="panel" style="margin-top:14px;">\n      <div class="panel-head"><h3>Registrar pago manual</h3></div>\n      <div class="panel-body">\n        <div class="g2">\n          <div><label>Cliente</label><input id="pay-name"><label>Email</label><input id="pay-email" type="email"></div>\n          <div>\n            <div class="g2" style="gap:10px;"><div><label>Monto</label><input id="pay-amount" type="number"></div><div><label>Moneda</label><select id="pay-cur"><option>USD</option><option>ARS</option></select></div></div>\n            <label>Método</label><select id="pay-method"><option value="transfer">Transferencia</option><option value="cash">Efectivo</option><option value="mp">MercadoPago</option></select>\n            <label>Referencia</label><input id="pay-ref">\n          </div>\n        </div>\n        <div style="margin-top:12px;"><button class="primary" onclick="registerPayment()">Registrar pago</button></div>\n      </div>\n    </div>\n    <div class="panel" style="margin-top:14px;">\n      <div class="panel-head"><h3>Historial de pagos</h3><button class="ghost" onclick="loadPayments()" style="font-size:11px;">↺</button></div>\n      <div class="panel-body" id="payments-list"><div class="empty"><div>Sin pagos</div></div></div>\n    </div>\n  </div>\n\n  <!-- Surveys -->\n  <div class="page" id="page-surveys">\n    <div class="page-header">\n      <div><div class="page-title">Encuestas</div><div class="page-sub">Feedback · Actualizaciones pendientes</div></div>\n      <div class="header-actions">\n        <button class="secondary" onclick="loadSurveys()">↺ Actualizar</button>\n        <button class="primary" onclick="sendSurveys()">📨 Enviar encuestas</button>\n      </div>\n    </div>\n    <div id="pending-section" style="margin-bottom:14px;"></div>\n    <div class="panel">\n      <div class="panel-head"><h3>Todas las encuestas</h3><span class="badge gray" id="surveys-count">—</span></div>\n      <div class="panel-body" id="surveys-list"><div class="empty"><div>Sin encuestas</div></div></div>\n    </div>\n  </div>\n\n  <!-- Releases -->\n  <div class="page" id="page-releases">\n    <div class="page-header">\n      <div><div class="page-title">Versiones</div><div class="page-sub">Nombrá cada versión como quieras</div></div>\n    </div>\n    <div class="g2">\n      <div class="panel">\n        <div class="panel-head"><h3>Versión actual</h3></div>\n        <div class="panel-body" id="current-version"><div class="empty"><div>Cargando...</div></div></div>\n      </div>\n      <div class="panel">\n        <div class="panel-head"><h3>Nombrar versión</h3></div>\n        <div class="panel-body">\n          <label>Nombre</label><input id="rel-name" placeholder="Ej: Operación Cóndor">\n          <label>Descripción (opcional)</label><textarea id="rel-desc" style="min-height:56px;"></textarea>\n          <div style="margin-top:12px;"><button class="primary" style="width:100%;justify-content:center;" onclick="saveRelease()">Guardar nombre</button></div>\n        </div>\n      </div>\n    </div>\n    <div class="panel" style="margin-top:14px;">\n      <div class="panel-head"><h3>Historial</h3></div>\n      <div class="panel-body" id="releases-list"><div class="empty"><div>Sin versiones</div></div></div>\n    </div>\n  </div>\n\n  <!-- Updates -->\n  <div class="page" id="page-updates">\n    <div class="page-header">\n      <div><div class="page-title">Actualizaciones con IA</div><div class="page-sub">Modificá Banzai84 con IA sin tocar código</div></div>\n    </div>\n    <div class="panel" style="margin-bottom:14px;">\n      <div class="panel-head"><h3>⚙ Configuración GitHub y Railway</h3><button class="ghost" onclick="toggleConfig()" id="config-toggle" style="font-size:11px;">Editar</button></div>\n      <div class="panel-body" id="updates-config-body">\n        <div id="config-status" style="font-size:12px;color:var(--text2);">Cargando...</div>\n      </div>\n      <div class="panel-body" id="updates-config-form" style="display:none;">\n        <label>GitHub Token (ghp_...)</label>\n        <input id="cfg-github-token" type="password" placeholder="ghp_...">\n        <label>GitHub Repo</label>\n        <input id="cfg-github-repo" placeholder="twilliamsfritz-alt/banzai84" value="twilliamsfritz-alt/banzai84">\n        <label>Railway Deploy Webhook (opcional)</label>\n        <input id="cfg-railway-webhook" placeholder="https://backboard.railway.app/webhooks/...">\n        <label>OpenAI API Key</label>\n        <input id="cfg-openai-key" type="password" placeholder="sk-...">\n        <div style="margin-top:12px;display:flex;gap:8px;">\n          <button class="primary" onclick="saveConfig()">Guardar</button>\n          <button class="ghost" onclick="toggleConfig()">Cancelar</button>\n        </div>\n      </div>\n    </div>\n    <div class="g2">\n      <div>\n        <div class="panel" style="margin-bottom:14px;">\n          <div class="panel-head"><h3>🤖 Generador de cambios con IA</h3></div>\n          <div class="panel-body">\n            <label>Archivo a modificar</label>\n            <select id="upd-file">\n              <option value="app.py">app.py — Backend principal</option>\n              <option value="static/app.js">static/app.js — Frontend JavaScript</option>\n              <option value="templates/index.html">templates/index.html — HTML del cliente</option>\n            </select>\n            <label>¿Qué querés cambiar? (en español)</label>\n            <textarea id="upd-desc" style="min-height:100px;" placeholder="Ej: Agregá un campo de prioridad en los deals del pipeline"></textarea>\n            <label>Mensaje del commit</label>\n            <input id="upd-msg" placeholder="ej: Agregar prioridad en deals">\n            <div style="margin-top:12px;display:flex;gap:8px;">\n              <button class="primary" onclick="generateChange()" id="btn-generate">✦ Generar con IA</button>\n              <button class="ghost" onclick="loadFileContent()">Ver archivo actual</button>\n            </div>\n          </div>\n        </div>\n        <div class="panel">\n          <div class="panel-head"><h3>⚡ Cambios rápidos</h3></div>\n          <div class="panel-body" style="display:flex;flex-direction:column;gap:6px;">\n            <button class="secondary" style="text-align:left;font-size:12px;" onclick="quickChange(\'Agrega un campo de prioridad alta/media/baja en los deals del pipeline\', \'app.py\', \'Agregar prioridad en deals\')">+ Campo prioridad en deals</button>\n            <button class="secondary" style="text-align:left;font-size:12px;" onclick="quickChange(\'Agrega soporte para rubro de farmacia independiente con tacticas del sector farmaceutico\', \'app.py\', \'Rubro farmacia\')">+ Rubro farmacia independiente</button>\n            <button class="secondary" style="text-align:left;font-size:12px;" onclick="quickChange(\'Agrega una seccion de metricas por vendedor que muestre deals cerrados, revenue y tasa de cierre de cada usuario con rol seller\', \'app.py\', \'Metricas por vendedor\')">+ Métricas por vendedor</button>\n            <button class="secondary" style="text-align:left;font-size:12px;" onclick="quickChange(\'Modifica el Sales Agent para que detecte cuando el cliente menciona competidores y responda destacando ventajas de Banzai\', \'app.py\', \'Sales Agent detecta competidores\')">+ Sales Agent detecta competidores</button>\n            <button class="secondary" style="text-align:left;font-size:12px;" onclick="quickChange(\'Agrega un resumen diario automatico al owner con deals del dia, ingresos y tareas pendientes\', \'app.py\', \'Resumen diario automatico\')">+ Resumen diario automático</button>\n          </div>\n        </div>\n      </div>\n      <div>\n        <div class="panel" style="height:100%;">\n          <div class="panel-head"><h3>👁 Preview del código</h3><span class="badge gray" id="upd-file-badge">—</span></div>\n          <div style="padding:0 16px 8px;font-size:11px;color:var(--text3);" id="upd-preview-info"></div>\n          <div style="margin:0 16px;">\n            <textarea id="upd-preview" style="min-height:340px;font-family:monospace;font-size:11px;background:var(--surface3);border:1px solid var(--border2);border-radius:8px;padding:12px;width:100%;resize:vertical;" placeholder="El código generado aparece acá. Podés editarlo antes de deployar."></textarea>\n          </div>\n          <div style="padding:12px 16px 16px;display:flex;gap:8px;">\n            <button class="primary" onclick="deployChange()" id="btn-deploy" disabled>🚀 Deploy a Railway</button>\n            <button class="secondary" onclick="copyCode()">Copiar</button>\n            <button class="ghost" onclick="clearPreview()">✕</button>\n          </div>\n          <div id="deploy-result" style="margin:0 16px 16px;font-size:12px;"></div>\n        </div>\n      </div>\n    </div>\n  </div>\n\n  <!-- Reports -->\n  <div class="page" id="page-reports">\n    <div class="page-header">\n      <div><div class="page-title">Reportes en vivo</div><div class="page-sub">P&L · Deals · Inventario</div></div>\n      <div class="header-actions"><button class="secondary" onclick="loadReports()">↺ Actualizar</button></div>\n    </div>\n    <div class="kpi-row" id="reports-kpis"></div>\n    <div class="g2" id="reports-detail"></div>\n  </div>\n\n  <!-- Events -->\n  <div class="page" id="page-events">\n    <div class="page-header">\n      <div><div class="page-title">Eventos de agentes</div><div class="page-sub">Log de actividad en tiempo real</div></div>\n      <div class="header-actions">\n        <button class="secondary" onclick="loadEvents(true)">↺ Actualizar</button>\n        <button class="primary" onclick="runAudit()">🔍 Auditar</button>\n      </div>\n    </div>\n    <div class="panel">\n      <div class="panel-head"><h3>Últimos 20 eventos</h3></div>\n      <div class="panel-body" id="events-full"><div class="empty"><div>Sin eventos</div></div></div>\n    </div>\n  </div>\n\n</main>\n\n<div id="toast"></div>\n\n<script>\nfunction $(id){ return document.getElementById(id); }\nfunction toast(msg, type){\n  if(!type) type=\'ok\';\n  var el=document.createElement(\'div\');\n  el.className=\'t-item \'+type;\n  el.textContent=msg;\n  $(\'toast\').appendChild(el);\n  setTimeout(function(){el.remove();},3500);\n}\nfunction fmt(v,cur){\n  if(!cur) cur=\'USD\';\n  return new Intl.NumberFormat(\'es-AR\',{style:\'currency\',currency:cur,maximumFractionDigits:0}).format(v||0);\n}\n\nfunction go(id){\n  document.querySelectorAll(\'.page\').forEach(function(p){p.classList.remove(\'active\');});\n  document.querySelectorAll(\'.nav-btn\').forEach(function(b){b.classList.remove(\'active\');});\n  $(\'page-\'+id).classList.add(\'active\');\n  document.querySelectorAll(\'.nav-btn\').forEach(function(b){\n    if(b.getAttribute(\'onclick\')==="go(\'"+id+"\')") b.classList.add(\'active\');\n  });\n  var loaders={\n    overview:loadOverview,\n    industries:loadIndustries,\n    users:loadUsers,\n    billing:function(){loadPlans();loadPayments();},\n    surveys:loadSurveys,\n    releases:loadReleases,\n    updates:loadUpdatesStatus,\n    reports:loadReports,\n    events:function(){loadEvents(true);}\n  };\n  if(loaders[id]) loaders[id]();\n}\n\nasync function api(path,opts){\n  if(!opts) opts={};\n  if(!opts.headers) opts.headers={\'Content-Type\':\'application/json\'};\n  try {\n    var r=await fetch(path,opts);\n    var d=await r.json().catch(function(){return {};});\n    if(!r.ok) throw new Error(d.error||(\'HTTP \'+r.status));\n    return d;\n  } catch(e){\n    console.warn(\'API error:\',path,e.message);\n    throw e;\n  }\n}\n\nasync function loadOverview(){\n  try {\n    var d=await api(\'/bz-admin/api/proxy/dashboard\');\n    var kpis=d.dashboard&&d.dashboard.kpis?d.dashboard.kpis:{};\n    var cur=d.dashboard&&d.dashboard.user?d.dashboard.user.currency||\'USD\':\'USD\';\n    $(\'kpi-row\').innerHTML=[\n      {l:\'Conversaciones\',v:kpis.open_chats||0,c:\'\'},\n      {l:\'Hot leads\',v:kpis.hot_leads||0,c:\'amber\'},\n      {l:\'Ingresos\',v:fmt(kpis.income||0,cur),c:\'green\'},\n      {l:\'Gastos\',v:fmt(kpis.expenses||0,cur),c:\'red\'},\n      {l:\'Flujo neto\',v:fmt(kpis.net||0,cur),c:(kpis.net||0)>=0?\'green\':\'red\'},\n    ].map(function(k){return \'<div class="kpi \'+k.c+\'"><div class="kpi-label">\'+k.l+\'</div><div class="kpi-val">\'+k.v+\'</div></div>\';}).join(\'\');\n    $(\'overview-status\').innerHTML=\'<div style="font-size:12px;color:var(--green);margin-bottom:8px;">✓ Conectado a Banzai</div>\';\n  } catch(e){\n    $(\'overview-status\').innerHTML=\'<div style="background:rgba(245,158,11,.08);border:1px solid rgba(245,158,11,.2);border-radius:10px;padding:12px 16px;font-size:12px;color:var(--amber);margin-bottom:8px;">⚠ Sin conexión a Banzai — iniciá Banzai en el puerto 5000 y recargá</div>\';\n    $(\'kpi-row\').innerHTML=\'\';\n  }\n  loadEvents(false);\n  loadIndustriesOverview();\n}\n\nasync function loadEvents(full){\n  try {\n    var r=await api(\'/bz-admin/api/proxy/agents/events\');\n    var evs=r.events||[];\n    var COLS={sales_agent:\'var(--brand)\',accounting_agent:\'var(--green)\',auditor_agent:\'#A78BFA\',ai_engine:\'var(--amber)\'};\n    var html=evs.length?evs.map(function(e){\n      return \'<div style="display:flex;gap:10px;padding:9px 0;border-bottom:1px solid var(--border);align-items:flex-start;">\'\n        +\'<div style="width:6px;height:6px;border-radius:50%;background:\'+(COLS[e.agent]||\'var(--text3)\')+\';flex-shrink:0;margin-top:5px;"></div>\'\n        +\'<div style="flex:1;min-width:0;">\'\n        +\'<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:\'+(COLS[e.agent]||\'var(--text3)\')+\';">\'+(e.agent||\'—\')+\'</div>\'\n        +\'<div style="font-size:12px;">\'+(e.action||\'\')+\'</div>\'\n        +\'<div style="font-size:11px;color:var(--text3);">\'+(e.output_summary||\'\')+\'</div>\'\n        +\'</div>\'\n        +\'<div style="font-size:10px;color:var(--text3);">\'+(e.created_at||\'\').substring(11,16)+\'</div>\'\n        +\'</div>\';\n    }).join(\'\'):\'<div class="empty"><div>Sin eventos todavía</div></div>\';\n    if($(\'events-list\')) $(\'events-list\').innerHTML=html;\n    if(full&&$(\'events-full\')) $(\'events-full\').innerHTML=html;\n  } catch(e){\n    if($(\'events-list\')) $(\'events-list\').innerHTML=\'<div class="empty"><div>Sin conexión</div></div>\';\n  }\n}\n\nasync function runAudit(){\n  try {\n    var r=await api(\'/bz-admin/api/proxy/agents/audit\',{method:\'POST\'});\n    toast(r.report&&r.report.health===\'ok\'?\'Auditoría: todo bien ✓\':\'Auditoría: issues detectados ⚠\',r.report&&r.report.health===\'ok\'?\'ok\':\'err\');\n  } catch(e){toast(\'Error en auditoría\',\'err\');}\n}\n\nasync function loadIndustriesOverview(){\n  try {\n    var r=await api(\'/bz-admin/api/proxy/industries\');\n    var inds=r.industries||[];\n    $(\'ind-badge\').textContent=inds.length+\' rubros\';\n    $(\'ind-overview\').innerHTML=inds.slice(0,15).map(function(i){\n      return \'<div class="row"><span style="font-size:12px;font-weight:500;">\'+(i.name||i.slug)+\'</span>\'\n        +\'<span class="badge \'+(i.is_custom?\'green\':\'gray\')+\'">\'+(i.is_custom?\'Custom\':\'Base\')+\'</span></div>\';\n    }).join(\'\');\n  } catch(e){$(\'ind-overview\').innerHTML=\'<div class="empty"><div>Sin conexión</div></div>\';}\n}\n\nvar editingSlug=null;\nvar indsCache=[];\n\nasync function loadIndustries(){\n  try {\n    var r=await api(\'/bz-admin/api/proxy/industries\');\n    indsCache=r.industries||[];\n    $(\'ind-list\').innerHTML=indsCache.length?indsCache.map(function(i){\n      return \'<div class="ind-card">\'\n        +\'<div class="ind-card-head">\'\n        +\'<span class="ind-name">\'+(i.name||i.slug)+\'</span>\'\n        +\'<span class="ind-slug">\'+i.slug+\'</span>\'\n        +\'<span class="badge \'+(i.is_custom?\'green\':\'gray\')+\'">\'+(i.is_custom?\'Custom\':\'Base\')+\'</span>\'\n        +\'<button class="secondary" style="padding:3px 8px;font-size:11px;" onclick="editInd(\\\'\'+i.slug+\'\\\')">Editar</button>\'\n        +(i.slug!==\'default\'?\'<button class="danger" style="padding:3px 8px;font-size:11px;" onclick="deleteInd(\\\'\'+i.slug+\'\\\')">✕</button>\':\'\')\n        +\'</div>\'\n        +\'<div>\'+(i.tactics||[]).map(function(t){return \'<span class="chip">\'+t+\'</span>\';}).join(\'\')+\'</div>\'\n        +\'</div>\';\n    }).join(\'\'):\'<div class="empty"><div>Sin rubros</div></div>\';\n  } catch(e){$(\'ind-list\').innerHTML=\'<div class="empty"><div>Error al cargar rubros</div></div>\';}\n}\n\nfunction openIndForm(){editingSlug=null;$(\'ind-form-title\').textContent=\'Nuevo rubro\';$(\'ind-form\').style.display=\'block\';[\'f-slug\',\'f-name\',\'f-upsell\',\'f-tactics\',\'f-kpis\',\'f-keywords\',\'f-objections\'].forEach(function(id){$(id).value=\'\';});$(\'ind-form\').scrollIntoView({behavior:\'smooth\'});}\nfunction closeIndForm(){$(\'ind-form\').style.display=\'none\';editingSlug=null;}\n\nfunction editInd(slug){\n  var i=indsCache.find(function(x){return x.slug===slug;});\n  if(!i) return;\n  editingSlug=slug;\n  $(\'ind-form-title\').textContent=\'Editar: \'+i.name;\n  $(\'f-slug\').value=slug;$(\'f-name\').value=i.name||\'\';$(\'f-upsell\').value=i.upsell||\'\';\n  $(\'f-tactics\').value=(i.tactics||[]).join(\'\\n\');\n  $(\'f-kpis\').value=(i.kpis||[]).join(\'\\n\');\n  $(\'f-keywords\').value=(i.keywords||[]).join(\', \');\n  $(\'f-objections\').value=Object.entries(i.objections||{}).map(function(kv){return kv[0]+\': \'+kv[1];}).join(\'\\n\');\n  $(\'ind-form\').style.display=\'block\';$(\'ind-form\').scrollIntoView({behavior:\'smooth\'});\n}\n\nasync function saveIndustry(){\n  var slug=$(\'f-slug\').value.trim().toLowerCase().replace(/\\s+/g,\'_\');\n  var name=$(\'f-name\').value.trim();\n  if(!slug||!name){toast(\'Slug y nombre obligatorios\',\'err\');return;}\n  var obj={};\n  $(\'f-objections\').value.split(\'\\n\').forEach(function(l){var i=l.indexOf(\':\');if(i>0)obj[l.slice(0,i).trim()]=l.slice(i+1).trim();});\n  var payload={slug:slug,name:name,tactics:$(\'f-tactics\').value.split(\'\\n\').map(function(s){return s.trim();}).filter(Boolean),\n    kpis:$(\'f-kpis\').value.split(\'\\n\').map(function(s){return s.trim();}).filter(Boolean),\n    keywords:$(\'f-keywords\').value.split(\',\').map(function(s){return s.trim();}).filter(Boolean),\n    objections:obj,upsell:$(\'f-upsell\').value.trim()};\n  try {\n    if(editingSlug) await api(\'/bz-admin/api/proxy/industries/\'+editingSlug,{method:\'PUT\',body:JSON.stringify(payload)});\n    else await api(\'/bz-admin/api/proxy/industries\',{method:\'POST\',body:JSON.stringify(payload)});\n    toast(\'Rubro guardado ✓\',\'ok\');closeIndForm();await loadIndustries();\n  } catch(e){toast(e.message,\'err\');}\n}\n\nasync function deleteInd(slug){\n  if(!confirm(\'¿Eliminar "\'+slug+\'"?\')) return;\n  try{await api(\'/bz-admin/api/proxy/industries/\'+slug,{method:\'DELETE\'});toast(\'Eliminado\',\'ok\');await loadIndustries();}\n  catch(e){toast(e.message,\'err\');}\n}\n\nasync function loadUsers(){\n  try {\n    var r=await api(\'/bz-admin/api/proxy/users\');\n    var users=r.users||[];\n    $(\'users-list\').innerHTML=users.length?users.map(function(u){\n      return \'<div class="row"><div>\'\n        +\'<div style="font-weight:600;font-size:13px;">\'+u.name+\'</div>\'\n        +\'<div style="font-size:11px;color:var(--text3);">\'+u.email+\'</div>\'\n        +\'</div><div style="display:flex;align-items:center;gap:6px;">\'\n        +\'<span class="badge \'+(u.role===\'owner\'?\'green\':u.role===\'seller\'?\'blue\':\'gray\')+\'">\'+u.role+\'</span>\'\n        +(!u.active?\'<span class="badge red">Inactivo</span>\':\'\')\n        +(u.role!==\'owner\'?\'<button onclick="toggleUser(\'+u.id+\',\'+(u.active?0:1)+\')" style="padding:3px 8px;font-size:10px;background:var(--surface3);border:1px solid var(--border2);border-radius:5px;color:var(--text2);cursor:pointer;">\'+(u.active?\'Suspender\':\'Activar\')+\'</button>\':\'\')\n        +\'</div></div>\';\n    }).join(\'\'):\'<div class="empty"><div>Sin usuarios</div></div>\';\n  } catch(e){$(\'users-list\').innerHTML=\'<div class="empty"><div>Error al cargar</div></div>\';}\n}\n\nasync function quickCreate(){\n  var name=$(\'qc-name\').value.trim(),user=$(\'qc-user\').value.trim(),pass=$(\'qc-pass\').value,role=$(\'qc-role\').value;\n  if(!name||!user||!pass){toast(\'Completá todos los campos\',\'err\');return;}\n  try {\n    var r=await api(\'/bz-admin/api/proxy/users/quick-create\',{method:\'POST\',body:JSON.stringify({name:name,username:user,password:pass,role:role})});\n    var res=$(\'qc-result\');\n    res.style.display=\'block\';res.style.background=\'var(--green-dim)\';res.style.border=\'1px solid rgba(16,185,129,.2)\';\n    res.innerHTML=\'✓ <strong>\'+r.email+\'</strong> · Contraseña: <strong>\'+pass+\'</strong>\';\n    $(\'qc-name\').value=\'\';$(\'qc-user\').value=\'\';$(\'qc-pass\').value=\'\';\n    toast(\'Usuario \'+r.email+\' creado ✓\',\'ok\');await loadUsers();\n  } catch(e){toast(e.message,\'err\');}\n}\n\nasync function toggleUser(id,active){\n  try{await api(\'/bz-admin/api/proxy/users/\'+id,{method:\'PUT\',body:JSON.stringify({active:active})});toast(active?\'Activado\':\'Suspendido\',\'ok\');await loadUsers();}\n  catch(e){toast(e.message,\'err\');}\n}\n\nasync function loadPlans(){\n  try {\n    var r=await api(\'/bz-admin/api/proxy/billing/plans\');\n    $(\'plans-list\').innerHTML=(r.plans||[]).length?(r.plans||[]).map(function(p){\n      return \'<div class="row"><div><div style="font-weight:600;">\'+p.name+\'</div><div style="font-size:11px;color:var(--text3);">\'+(p.description||\'\')+\'</div></div>\'\n        +\'<div style="font-size:14px;font-weight:700;color:var(--brand);">$\'+p.price+\' \'+p.currency+\'/\'+(p.interval===\'month\'?\'mes\':\'año\')+\'</div></div>\';\n    }).join(\'\'):\'<div class="empty"><div>Sin planes</div></div>\';\n  } catch(e){}\n}\n\nasync function savePlan(){\n  var payload={name:$(\'plan-name\').value.trim(),description:$(\'plan-desc\').value.trim(),price:parseFloat($(\'plan-price\').value||0),currency:$(\'plan-cur\').value,interval:$(\'plan-int\').value};\n  if(!payload.name){toast(\'Nombre obligatorio\',\'err\');return;}\n  try{await api(\'/bz-admin/api/proxy/billing/plans\',{method:\'POST\',body:JSON.stringify(payload)});toast(\'Plan creado ✓\',\'ok\');await loadPlans();}\n  catch(e){toast(e.message,\'err\');}\n}\n\nfunction openPlanForm(){$(\'plan-form\').scrollIntoView({behavior:\'smooth\'});}\n\nasync function loadPayments(){\n  try {\n    var r=await api(\'/bz-admin/api/proxy/billing/payments\');\n    $(\'payments-list\').innerHTML=(r.payments||[]).length?(r.payments||[]).map(function(p){\n      return \'<div class="row"><div><div style="font-weight:500;">\'+p.customer_name+\'</div><div style="font-size:10px;color:var(--text3);">\'+(p.banzai_invoice_number||\'\')+\' · \'+p.method+\' · \'+(p.created_at||\'\').substring(0,10)+\'</div></div>\'\n        +\'<div style="display:flex;align-items:center;gap:6px;"><span style="font-weight:700;color:var(--green);">$\'+p.amount+\' \'+p.currency+\'</span>\'\n        +\'<span class="badge \'+(p.status===\'paid\'?\'green\':\'amber\')+\'">\'+p.status+\'</span></div></div>\';\n    }).join(\'\'):\'<div class="empty"><div>Sin pagos</div></div>\';\n  } catch(e){}\n}\n\nasync function registerPayment(){\n  var payload={customer_name:$(\'pay-name\').value.trim(),customer_email:$(\'pay-email\').value.trim(),amount:parseFloat($(\'pay-amount\').value||0),currency:$(\'pay-cur\').value,method:$(\'pay-method\').value,reference:$(\'pay-ref\').value.trim(),status:\'paid\'};\n  if(!payload.customer_name||!payload.amount){toast(\'Completá los campos\',\'err\');return;}\n  try{var r=await api(\'/bz-admin/api/proxy/billing/payments\',{method:\'POST\',body:JSON.stringify(payload)});toast(\'Pago registrado — \'+r.invoice_number+\' ✓\',\'ok\');await loadPayments();}\n  catch(e){toast(e.message,\'err\');}\n}\n\nasync function loadSurveys(){\n  try {\n    var r=await api(\'/bz-admin/api/proxy/surveys\');\n    var surveys=r.surveys||[],pending=r.pending_updates||0;\n    $(\'surveys-count\').textContent=surveys.length;\n    if(pending>0){\n      var pendingHtml=\'<div style="background:rgba(245,158,11,.06);border:1px solid rgba(245,158,11,.2);border-radius:12px;padding:14px 18px;margin-bottom:12px;">\'\n        +\'<div style="font-size:13px;font-weight:600;color:var(--amber);margin-bottom:10px;">⚡ \'+pending+\' actualizaciones pendientes de tu aprobación</div>\';\n      surveys.filter(function(s){return s.update_status===\'requested\';}).forEach(function(s){\n        pendingHtml+=\'<div style="background:var(--surface2);border-radius:10px;padding:14px;margin-bottom:8px;">\'\n          +\'<div style="font-size:13px;font-weight:600;margin-bottom:6px;">\'+s.customer_name+\' · NPS \'+s.nps_score+\'</div>\'\n          +\'<div style="font-size:12px;color:var(--text2);margin-bottom:8px;">"\'+(s.response_text||\'\')+\'"</div>\'\n          +\'<div style="background:var(--surface3);border-radius:8px;padding:10px;margin-bottom:10px;font-size:13px;">\'+s.suggested_update+\'</div>\'\n          +\'<textarea id="upd-\'+s.id+\'" style="min-height:56px;margin-bottom:8px;">\'+s.suggested_update+\'</textarea>\'\n          +\'<div style="display:flex;gap:8px;">\'\n          +\'<button class="primary" onclick="activateUpdate(\'+s.id+\')" style="flex:1;justify-content:center;">✓ Activar</button>\'\n          +\'<button class="danger" onclick="rejectUpdate(\'+s.id+\')">✕ Rechazar</button>\'\n          +\'</div></div>\';\n      });\n      pendingHtml+=\'</div>\';\n      $(\'pending-section\').innerHTML=pendingHtml;\n    } else {\n      $(\'pending-section\').innerHTML=\'\';\n    }\n    $(\'surveys-list\').innerHTML=surveys.length?surveys.map(function(s){\n      return \'<div class="row"><div><div style="font-weight:500;">\'+s.customer_name+\'</div><div style="font-size:11px;color:var(--text3);">\'+s.customer_email+\' · \'+(s.sent_at||\'\').substring(0,10)+\'</div></div>\'\n        +\'<div style="display:flex;align-items:center;gap:6px;">\'\n        +(s.nps_score!==null?\'<span class="badge \'+(s.nps_score>=9?\'green\':s.nps_score>=7?\'blue\':\'red\')+\'">NPS \'+s.nps_score+\'</span>\':\'<span class="badge gray">—</span>\')\n        +\'<span class="badge \'+(s.status===\'responded\'?\'green\':\'gray\')+\'">\'+(s.status===\'responded\'?\'Respondió\':\'Pendiente\')+\'</span>\'\n        +\'</div></div>\';\n    }).join(\'\'):\'<div class="empty"><div>Sin encuestas</div></div>\';\n  } catch(e){$(\'surveys-list\').innerHTML=\'<div class="empty"><div>Error al cargar</div></div>\';}\n}\n\nasync function sendSurveys(){try{var r=await api(\'/bz-admin/api/proxy/surveys/send\',{method:\'POST\',body:JSON.stringify({})});toast(r.sent+\' encuestas enviadas ✓\',\'ok\');await loadSurveys();}catch(e){toast(e.message,\'err\');}}\nasync function activateUpdate(id){var txt=$(\'upd-\'+id)?$(\'upd-\'+id).value.trim():\'\';if(!txt){toast(\'Escribí la versión aprobada\',\'err\');return;}if(!confirm(\'¿Activar?\'))return;try{await api(\'/bz-admin/api/proxy/surveys/\'+id+\'/activate\',{method:\'POST\',body:JSON.stringify({confirmed_update:txt})});toast(\'Activada ✓\',\'ok\');await loadSurveys();}catch(e){toast(e.message,\'err\');}}\nasync function rejectUpdate(id){try{await api(\'/bz-admin/api/proxy/surveys/\'+id+\'/reject\',{method:\'POST\',body:JSON.stringify({})});toast(\'Rechazada\',\'ok\');await loadSurveys();}catch(e){toast(e.message,\'err\');}}\n\nasync function loadReleases(){\n  try {\n    var r=await api(\'/bz-admin/api/proxy/releases\');\n    var cur=r.current||{};\n    $(\'current-version\').innerHTML=\'<div style="text-align:center;padding:16px;"><div style="font-size:32px;font-weight:800;color:var(--brand);">\'+(cur.release_name||cur.version||\'—\')+\'</div><div style="font-size:12px;color:var(--text3);margin-top:6px;">versión \'+cur.version+\'</div></div>\';\n    $(\'releases-list\').innerHTML=(r.releases||[]).length?(r.releases||[]).map(function(rel){\n      return \'<div class="row"><div><div style="font-size:14px;font-weight:700;color:var(--brand);">\'+rel.release_name+\'</div><div style="font-size:11px;color:var(--text3);">v\'+rel.version+\' · \'+(rel.created_at||\'\').substring(0,10)+\'</div></div></div>\';\n    }).join(\'\'):\'<div class="empty"><div>Sin versiones</div></div>\';\n  } catch(e){}\n}\n\nasync function saveRelease(){\n  var name=$(\'rel-name\').value.trim();\n  if(!name){toast(\'El nombre es obligatorio\',\'err\');return;}\n  try{await api(\'/bz-admin/api/proxy/releases\',{method:\'POST\',body:JSON.stringify({release_name:name,description:$(\'rel-desc\').value.trim()})});toast(\'Versión "\'+name+\'" guardada ✓\',\'ok\');$(\'rel-name\').value=\'\';$(\'rel-desc\').value=\'\';await loadReleases();}\n  catch(e){toast(e.message,\'err\');}\n}\n\nasync function loadReports(){\n  try {\n    var r=await api(\'/bz-admin/api/proxy/reports\');\n    var pl=r.pl||{},cur=r.currency||\'USD\';\n    $(\'reports-kpis\').innerHTML=[\n      {l:\'Ingresos\',v:fmt(pl.income||0,cur),c:\'green\'},\n      {l:\'Gastos\',v:fmt(pl.expenses||0,cur),c:\'red\'},\n      {l:\'Neto\',v:fmt(pl.net||0,cur),c:(pl.net||0)>=0?\'green\':\'red\'},\n      {l:\'Margen\',v:(pl.margin_pct||0)+\'%\',c:\'\'},\n      {l:\'Deals cerrados\',v:fmt((r.deals&&r.deals.total_closed_value)||0,cur),c:\'green\'},\n    ].map(function(k){return \'<div class="kpi \'+k.c+\'"><div class="kpi-label">\'+k.l+\'</div><div class="kpi-val">\'+k.v+\'</div></div>\';}).join(\'\');\n    $(\'reports-detail\').innerHTML=[\n      {title:\'Ingresos por categoría\',data:r.income_by_category||{}},\n      {title:\'Gastos por categoría\',data:r.expense_by_category||{}},\n    ].map(function(s){\n      return \'<div class="panel"><div class="panel-head"><h3>\'+s.title+\'</h3></div><div class="panel-body">\'\n        +(Object.keys(s.data).length?Object.entries(s.data).map(function(kv){return \'<div class="row"><span>\'+kv[0]+\'</span><strong>\'+fmt(kv[1],cur)+\'</strong></div>\';}).join(\'\'):\'<div class="empty"><div>Sin datos</div></div>\')\n        +\'</div></div>\';\n    }).join(\'\');\n  } catch(e){$(\'reports-kpis\').innerHTML=\'<div style="grid-column:1/-1;" class="empty"><div>Sin conexión a Banzai</div></div>\';}\n}\n\nvar chatHistory=[];\nfunction bubble(role,text){\n  var el=$(\'chat-msgs\');\n  var b=document.createElement(\'div\');\n  b.className=\'bubble \'+role;\n  if(role===\'assistant\'){\n    b.innerHTML=text.replace(/\\n/g,\'<br>\');\n  } else {\n    b.textContent=text;\n  }\n  el.appendChild(b);\n  el.scrollTop=el.scrollHeight;\n}\n\nfunction ask(msg){$(\'chat-input\').value=msg;sendMsg();}\n\nasync function sendMsg(){\n  var input=$(\'chat-input\').value.trim();\n  if(!input) return;\n  bubble(\'user\',input);\n  chatHistory.push({role:\'user\',content:input});\n  $(\'chat-input\').value=\'\';\n  var typing=document.createElement(\'div\');\n  typing.className=\'bubble assistant\';typing.id=\'typing\';typing.style.opacity=\'.5\';typing.textContent=\'Pensando...\';\n  $(\'chat-msgs\').appendChild(typing);\n  $(\'chat-msgs\').scrollTop=$(\'chat-msgs\').scrollHeight;\n  try {\n    var r=await api(\'/bz-admin/api/ai-chat\',{method:\'POST\',body:JSON.stringify({message:input,history:chatHistory.slice(-6)})});\n    $(\'typing\')&&$(\'typing\').remove();\n    var reply=r.reply||\'Sin respuesta\';\n    bubble(\'assistant\',reply);\n    chatHistory.push({role:\'assistant\',content:reply});\n  } catch(e){$(\'typing\')&&$(\'typing\').remove();bubble(\'assistant\',\'Error: \'+e.message);}\n}\n\n// ── Updates System ──────────────────────────────────────────────────────────\nvar updatesCurrentContent=\'\';\nvar updatesFilePath=\'app.py\';\n\nasync function loadUpdatesStatus(){\n  try {\n    var r=await api(\'/bz-admin/api/updates/status\');\n    var html=\'\';\n    if(r.github_configured){\n      html+=\'<div style="margin-bottom:4px;color:var(--green);">✓ GitHub conectado — \'+r.github_repo+\'</div>\';\n    } else {\n      html+=\'<div style="margin-bottom:4px;color:var(--amber);">⚠ GitHub no configurado — hacé clic en Editar</div>\';\n    }\n    if(r.railway_configured){\n      html+=\'<div style="color:var(--green);">✓ Railway webhook activo</div>\';\n    } else {\n      html+=\'<div style="color:var(--text3);">— Railway webhook opcional (el deploy funciona igual via GitHub)</div>\';\n    }\n    $(\'config-status\').innerHTML=html;\n  } catch(e){\n    $(\'config-status\').innerHTML=\'<span style="color:var(--red);">Error al cargar estado</span>\';\n  }\n}\n\nfunction toggleConfig(){\n  var form=$(\'updates-config-form\');\n  var body=$(\'updates-config-body\');\n  var isHidden=form.style.display===\'none\';\n  form.style.display=isHidden?\'block\':\'none\';\n  body.style.display=isHidden?\'none\':\'block\';\n  $(\'config-toggle\').textContent=isHidden?\'Cancelar\':\'Editar\';\n}\n\nasync function saveConfig(){\n  var payload={\n    github_token:$(\'cfg-github-token\').value.trim(),\n    github_repo:$(\'cfg-github-repo\').value.trim(),\n    railway_webhook:$(\'cfg-railway-webhook\').value.trim(),\n    openai_key:$(\'cfg-openai-key\').value.trim(),\n  };\n  try{\n    await api(\'/bz-admin/api/updates/save-config\',{method:\'POST\',body:JSON.stringify(payload)});\n    toast(\'Configuración guardada ✓\',\'ok\');\n    toggleConfig();\n    await loadUpdatesStatus();\n  } catch(e){toast(e.message,\'err\');}\n}\n\nasync function loadFileContent(){\n  var fp=$(\'upd-file\').value;\n  $(\'upd-file-badge\').textContent=fp;\n  $(\'upd-preview\').value=\'Cargando \'+fp+\'...\';\n  try {\n    var r=await api(\'/bz-admin/api/updates/file?path=\'+encodeURIComponent(fp));\n    updatesCurrentContent=r.content;\n    updatesFilePath=fp;\n    $(\'upd-preview\').value=r.content;\n    var lines=r.content.split(\'\\n\').length;\n    $(\'upd-preview-info\').textContent=fp+\' — \'+lines+\' líneas\';\n    toast(\'Archivo cargado ✓\',\'ok\');\n  } catch(e){$(\'upd-preview\').value=\'Error: \'+e.message;toast(e.message,\'err\');}\n}\n\nasync function generateChange(){\n  var desc=$(\'upd-desc\').value.trim();\n  var fp=$(\'upd-file\').value;\n  if(!desc){toast(\'Describí qué querés cambiar\',\'err\');return;}\n  $(\'btn-generate\').disabled=true;\n  $(\'btn-generate\').textContent=\'✦ Generando...\';\n  $(\'upd-preview\').value=\'La IA está analizando el código y generando los cambios...\\n\\nEsto puede tomar 20-40 segundos...\';\n  $(\'upd-file-badge\').textContent=fp;\n  updatesFilePath=fp;\n  if(!updatesCurrentContent||updatesFilePath!==fp){\n    try{var fr=await api(\'/bz-admin/api/updates/file?path=\'+encodeURIComponent(fp));updatesCurrentContent=fr.content;}\n    catch(e){updatesCurrentContent=\'\';}\n  }\n  try {\n    var r=await api(\'/bz-admin/api/updates/generate\',{method:\'POST\',body:JSON.stringify({\n      description:desc,file_path:fp,current_content:updatesCurrentContent\n    })});\n    $(\'upd-preview\').value=r.code;\n    updatesCurrentContent=r.code;\n    $(\'upd-preview-info\').textContent=fp+\' generado — \'+r.code.split(\'\\n\').length+\' líneas. Revisá antes de deployar.\';\n    $(\'btn-deploy\').disabled=false;\n    toast(\'Código generado ✓ — revisá antes de deployar\',\'ok\');\n  } catch(e){\n    $(\'upd-preview\').value=\'Error: \'+e.message;\n    toast(e.message,\'err\');\n  }\n  $(\'btn-generate\').disabled=false;\n  $(\'btn-generate\').textContent=\'✦ Generar con IA\';\n}\n\nfunction quickChange(desc,file,msg){\n  $(\'upd-desc\').value=desc;\n  $(\'upd-file\').value=file;\n  $(\'upd-msg\').value=msg;\n  updatesFilePath=file;\n  updatesCurrentContent=\'\';\n  generateChange();\n}\n\nasync function deployChange(){\n  var code=$(\'upd-preview\').value.trim();\n  var fp=$(\'upd-file\').value;\n  var msg=$(\'upd-msg\').value.trim()||\'Update via Banzai Admin\';\n  if(!code){toast(\'No hay código para deployar\',\'err\');return;}\n  if(!confirm(\'¿Deployar "\'+msg+\'" a Railway?\')) return;\n  $(\'btn-deploy\').disabled=true;\n  $(\'btn-deploy\').textContent=\'🚀 Deployando...\';\n  $(\'deploy-result\').innerHTML=\'\';\n  try {\n    var r=await api(\'/bz-admin/api/updates/deploy\',{method:\'POST\',body:JSON.stringify({code:code,file_path:fp,message:msg})});\n    $(\'deploy-result\').innerHTML=\'<div style="background:var(--green-dim);border:1px solid rgba(16,185,129,.2);border-radius:8px;padding:12px;margin-top:8px;">\'\n      +\'<div style="font-weight:700;color:var(--green);margin-bottom:6px;">✓ \'+r.message+\'</div>\'\n      +(r.railway_triggered?\'<div style="font-size:11px;color:var(--green);">✓ Railway redeploy disparado automáticamente</div>\':\'<div style="font-size:11px;color:var(--text3);">Railway va a detectar el cambio en ~30 segundos</div>\')\n      +(r.commit_url?\'<div style="margin-top:6px;"><a href="\'+r.commit_url+\'" target="_blank" style="font-size:11px;color:var(--brand);">Ver commit en GitHub →</a></div>\':\'\')\n      +\'</div>\';\n    toast(\'Deployado exitosamente ✓\',\'ok\');\n    $(\'upd-desc\').value=\'\';$(\'upd-preview\').value=\'\';$(\'btn-deploy\').disabled=true;\n  } catch(e){\n    $(\'deploy-result\').innerHTML=\'<div style="background:var(--red-dim);border:1px solid rgba(239,68,68,.2);border-radius:8px;padding:12px;color:var(--red);">✗ \'+e.message+\'</div>\';\n    toast(e.message,\'err\');\n  }\n  $(\'btn-deploy\').textContent=\'🚀 Deploy a Railway\';\n  $(\'btn-deploy\').disabled=false;\n}\n\nfunction copyCode(){\n  var code=$(\'upd-preview\').value;\n  if(!code){toast(\'No hay código para copiar\',\'err\');return;}\n  navigator.clipboard.writeText(code).then(function(){toast(\'Código copiado ✓\',\'ok\');}).catch(function(){toast(\'Error al copiar\',\'err\');});\n}\n\nfunction clearPreview(){\n  $(\'upd-preview\').value=\'\';$(\'upd-preview-info\').textContent=\'\';$(\'upd-file-badge\').textContent=\'—\';\n  $(\'btn-deploy\').disabled=true;$(\'deploy-result\').innerHTML=\'\';updatesCurrentContent=\'\';\n}\n\nasync function logout(){await api(\'/bz-admin/api/admin-logout\',{method:\'POST\'});window.location.href=\'/bz-admin/\';}\n\nloadOverview();\nbubble(\'assistant\',\'¡Hola! Soy la IA experta de Banzai Admin.\\n\\nPuedo ayudarte con:\\n• Crear y editar rubros del Sales Agent\\n• Analizar el negocio en tiempo real\\n• Scripts de ventas por industria\\n• Estrategia de crecimiento\\n\\n¿Qué necesitás?\');\n</script>\n</body>\n</html>\n'

BZ_ADMIN_LOGIN_HTML = '<!doctype html>\n<html lang="es">\n<head>\n  <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">\n  <title>Banzai Admin — Conectar</title>\n  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">\n  <style>\n    *{box-sizing:border-box;margin:0;padding:0;}\n    :root{--bg:#06080F;--surface:#0C0F1A;--surface2:#111520;--surface3:#161B2E;--border:rgba(255,255,255,.07);--border2:rgba(255,255,255,.12);--text:#F0F2FA;--text2:#9AA3C2;--text3:#4E5B7A;--brand:#6366F1;--brand-dim:rgba(99,102,241,.12);--brand-glow:rgba(99,102,241,.25);--brand-dark:#4F46E5;--green:#10B981;--red:#EF4444;--r:10px;--r-lg:16px;--r-xl:22px;}\n    body{font-family:\'Inter\',system-ui,sans-serif;background:var(--bg);color:var(--text);min-height:100vh;display:flex;align-items:center;justify-content:center;padding:40px 24px;-webkit-font-smoothing:antialiased;}\n    body::before{content:\'\';position:fixed;top:-20%;left:-10%;width:60vw;height:60vw;border-radius:50%;background:radial-gradient(circle,rgba(99,102,241,.07) 0%,transparent 65%);pointer-events:none;}\n    .dot-grid{position:fixed;inset:0;background-image:radial-gradient(rgba(255,255,255,.025) 1px,transparent 1px);background-size:28px 28px;pointer-events:none;}\n    .shell{max-width:420px;width:100%;position:relative;z-index:1;}\n    .wordmark{display:flex;align-items:center;gap:12px;justify-content:center;margin-bottom:36px;}\n    .wm-mark{width:44px;height:44px;background:var(--brand);border-radius:13px;display:flex;align-items:center;justify-content:center;font-weight:900;font-size:16px;color:#fff;box-shadow:0 0 24px var(--brand-glow);}\n    .wm-name{font-size:24px;font-weight:800;letter-spacing:-.03em;}\n    .admin-tag{font-size:11px;font-weight:700;background:rgba(99,102,241,.15);color:var(--brand);padding:2px 8px;border-radius:4px;margin-left:8px;letter-spacing:.06em;text-transform:uppercase;}\n    .card{background:var(--surface);border:1px solid var(--border2);border-radius:var(--r-xl);padding:36px;box-shadow:0 32px 80px rgba(0,0,0,.5);}\n    h2{font-size:20px;font-weight:700;margin-bottom:4px;}\n    .sub{font-size:13px;color:var(--text3);margin-bottom:24px;}\n    .url-badge{background:var(--surface3);border:1px solid var(--border);border-radius:var(--r);padding:10px 14px;font-size:12px;color:var(--text2);margin-bottom:20px;display:flex;align-items:center;gap:8px;}\n    .url-badge .dot{width:6px;height:6px;border-radius:50%;background:var(--green);box-shadow:0 0 6px var(--green);flex-shrink:0;}\n    label{display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.07em;color:var(--text3);margin-bottom:5px;margin-top:16px;}\n    input{width:100%;background:var(--surface3);border:1px solid var(--border2);border-radius:var(--r);color:var(--text);font:inherit;font-size:14px;padding:11px 14px;outline:none;transition:border-color .15s,box-shadow .15s;}\n    input:focus{border-color:var(--brand);box-shadow:0 0 0 3px var(--brand-dim);}\n    .btn{width:100%;background:var(--brand);color:#fff;border:none;border-radius:var(--r);padding:13px;font:inherit;font-size:14px;font-weight:600;cursor:pointer;margin-top:20px;transition:all .15s;display:flex;align-items:center;justify-content:center;gap:8px;}\n    .btn:hover{background:var(--brand-dark);box-shadow:0 0 20px var(--brand-glow);}\n    .err{color:var(--red);font-size:12px;margin-top:10px;min-height:16px;}\n    .footer{text-align:center;font-size:11px;color:var(--text3);margin-top:20px;}\n  </style>\n</head>\n<body>\n<div class="dot-grid"></div>\n<div class="shell">\n  <div class="wordmark">\n    <div class="wm-mark">BZ</div>\n    <div class="wm-name">Banzai<span class="admin-tag">Admin</span></div>\n  </div>\n  <div class="card">\n    <h2>Conectar con Banzai</h2>\n    <p class="sub">Ingresá con las mismas credenciales de tu workspace</p>\n    <div class="url-badge"><div class="dot"></div><span>{{ banzai_url }}</span></div>\n    <label>Email</label>\n    <input id="email" type="email" placeholder="tu@email.com" autocomplete="username">\n    <label>Contraseña</label>\n    <input id="password" type="password" placeholder="••••••••" autocomplete="current-password">\n    <button class="btn" id="btn-connect" onclick="connect()">Conectar →</button>\n    <div class="err" id="err"></div>\n    <div class="footer">Banzai Admin v1.0 · Panel de administración</div>\n  </div>\n</div>\n<script>\ndocument.getElementById(\'password\').addEventListener(\'keydown\', e=>{ if(e.key===\'Enter\') connect(); });\nasync function connect() {\n  document.getElementById(\'err\').textContent = \'\';\n  document.getElementById(\'btn-connect\').textContent = \'Entrando…\';\n  document.getElementById(\'btn-connect\').disabled = true;\n  const r = await fetch(\'/bz-admin/api/login\', {\n    method:\'POST\', headers:{\'Content-Type\':\'application/json\'},\n    body: JSON.stringify({\n      email:document.getElementById(\'email\').value,\n      password:document.getElementById(\'password\').value,\n      banzai_email:document.getElementById(\'banzai-email\')?.value||\'\',\n      banzai_password:document.getElementById(\'banzai-password\')?.value||\'\',\n    })\n  });\n  const d = await r.json();\n  if(d.ok) {\n    const status = document.getElementById(\'banzai-status\');\n    if(status) status.textContent = d.message || \'\';\n    setTimeout(() => window.location.href=\'/bz-admin/\', 800);\n  } else {\n    document.getElementById(\'err\').textContent = d.error||\'Error\';\n    document.getElementById(\'btn-connect\').textContent=\'Entrar al Admin →\';\n    document.getElementById(\'btn-connect\').disabled = false;\n  }\n}\n</script>\n</body>\n</html>\n'

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "superadmin@banzai84.com")
ADMIN_PASSWORD_ADMIN = os.environ.get("ADMIN_PASSWORD", "Banzai84#Admin2026")
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = os.environ.get("GITHUB_REPO", "twilliamsfritz-alt/banzai84")


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("bz_admin_logged_in"):
            return redirect("/bz-admin/login")
        return f(*args, **kwargs)
    return decorated


@app.get("/bz-admin/login")
def bz_admin_login_page():
    if session.get("bz_admin_logged_in"):
        return redirect("/bz-admin/")
    return render_template_string(BZ_ADMIN_LOGIN_HTML)


@app.post("/bz-admin/api/login")
def bz_admin_login():
    d = request.get_json(force=True) or {}
    if d.get("email") == ADMIN_USERNAME and d.get("password") == ADMIN_PASSWORD_ADMIN:
        session["bz_admin_logged_in"] = True
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Credenciales incorrectas"}), 401


@app.post("/bz-admin/api/logout")
def bz_admin_logout():
    session.pop("bz_admin_logged_in", None)
    return jsonify({"ok": True})


@app.get("/bz-admin/")
@admin_required
def bz_admin_panel():
    return render_template_string(BZ_ADMIN_HTML)


# Proxy all /bz-admin/api/proxy/* to internal Banzai routes
@app.route("/bz-admin/api/proxy/<path:subpath>", methods=["GET","POST","PUT","DELETE"])
@admin_required
def bz_admin_proxy(subpath):
    try:
        method = request.method
        body = request.get_json(force=True, silent=True)
        with closing(get_db()) as conn:
            # Route to internal handlers based on path
            if subpath == "dashboard":
                row = conn.execute("SELECT * FROM workspace_profiles LIMIT 1").fetchone()
                ws = conn.execute("SELECT * FROM workspaces LIMIT 1").fetchone()
                u = conn.execute("SELECT * FROM users WHERE role=\'owner\' LIMIT 1").fetchone()
                convs = conn.execute("SELECT COUNT(*) as c FROM conversations WHERE status=\'open\'").fetchone()
                hot = conn.execute("SELECT COUNT(*) as c FROM conversations WHERE ai_stage IN (\'hot_lead\',\'proposal_sent\')").fetchone()
                inc = conn.execute("SELECT COALESCE(SUM(amount),0) as s FROM ledger_entries WHERE entry_type=\'income\'").fetchone()
                exp = conn.execute("SELECT COALESCE(SUM(amount),0) as s FROM ledger_entries WHERE entry_type=\'expense\'").fetchone()
                income = float(inc["s"] if inc else 0)
                expenses = float(exp["s"] if exp else 0)
                user_data = dict(u) if u else {}
                user_data["currency"] = user_data.get("currency","USD")
                return jsonify({"ok":True,"dashboard":{"kpis":{"open_chats":convs["c"] if convs else 0,"hot_leads":hot["c"] if hot else 0,"income":income,"expenses":expenses,"net":income-expenses},"user":user_data}})
            elif subpath == "industries" and method == "GET":
                import json as _json
                merged = get_playbooks(1)
                custom_slugs = set(_load_db_industries(1).keys())
                industries_list = []
                for slug, data in merged.items():
                    industries_list.append({
                        "slug": slug,
                        "name": data.get("name", slug.replace("_"," ").title()),
                        "tactics": data.get("tactics", []),
                        "kpis": data.get("kpis", []),
                        "keywords": data.get("keywords", []),
                        "objections": data.get("objections", {}),
                        "upsell": data.get("upsell", ""),
                        "is_custom": slug in custom_slugs,
                    })
                industries_list.sort(key=lambda x: (not x["is_custom"], x["slug"]))
                return jsonify({"ok":True,"industries":industries_list})
            elif subpath == "industries" and method == "POST":
                import json as _json
                d2 = body or {}
                slug = d2.get("slug","").strip().lower().replace(" ","_")
                name = d2.get("name", slug)
                playbook_data = {
                    "name": name,
                    "tactics": d2.get("tactics", []),
                    "kpis": d2.get("kpis", []),
                    "keywords": d2.get("keywords", []),
                    "objections": d2.get("objections", {}),
                    "upsell": d2.get("upsell", ""),
                }
                now_iso = datetime.utcnow().isoformat()
                conn.execute(
                    "INSERT INTO industry_playbooks (workspace_id,slug,name,data_json,created_at,updated_at) VALUES (0,%s,%s,%s,%s,%s) ON CONFLICT(workspace_id,slug) DO UPDATE SET name=%s,data_json=%s,updated_at=%s"
                    if "postgres" in os.environ.get("DATABASE_URL","") else
                    "INSERT OR REPLACE INTO industry_playbooks (workspace_id,slug,name,data_json,created_at,updated_at) VALUES (0,?,?,?,?,?)",
                    (slug, name, _json.dumps(playbook_data), now_iso, now_iso, name, _json.dumps(playbook_data), now_iso)
                    if "postgres" in os.environ.get("DATABASE_URL","") else
                    (slug, name, _json.dumps(playbook_data), now_iso, now_iso)
                )
                conn.commit()
                return jsonify({"ok":True, "slug": slug})
            elif subpath.startswith("industries/") and method in ("PUT","DELETE"):
                import json as _json
                slug = subpath.split("/",1)[1]
                if method == "DELETE":
                    conn.execute("DELETE FROM industry_playbooks WHERE slug=%s" if "postgres" in os.environ.get("DATABASE_URL","") else "DELETE FROM industry_playbooks WHERE slug=?", (slug,))
                    conn.commit()
                    return jsonify({"ok":True})
                else:
                    d2 = body or {}
                    name = d2.get("name", slug)
                    playbook_data = {
                        "name": name,
                        "tactics": d2.get("tactics", []),
                        "kpis": d2.get("kpis", []),
                        "keywords": d2.get("keywords", []),
                        "objections": d2.get("objections", {}),
                        "upsell": d2.get("upsell", ""),
                    }
                    now_iso = datetime.utcnow().isoformat()
                    is_pg = "postgres" in os.environ.get("DATABASE_URL","")
                    if is_pg:
                        conn.execute(
                            "INSERT INTO industry_playbooks (workspace_id,slug,name,data_json,created_at,updated_at) VALUES (0,%s,%s,%s,%s,%s) ON CONFLICT(workspace_id,slug) DO UPDATE SET name=%s,data_json=%s,updated_at=%s",
                            (slug, name, _json.dumps(playbook_data), now_iso, now_iso, name, _json.dumps(playbook_data), now_iso)
                        )
                    else:
                        conn.execute(
                            "INSERT OR REPLACE INTO industry_playbooks (workspace_id,slug,name,data_json,created_at,updated_at) VALUES (0,?,?,?,?,?)",
                            (slug, name, _json.dumps(playbook_data), now_iso, now_iso)
                        )
                    conn.commit()
                    return jsonify({"ok":True})
            elif subpath == "users" and method == "GET":
                rows = conn.execute("SELECT id,name,email,role,active,last_login_at FROM users ORDER BY created_at DESC").fetchall()
                return jsonify({"ok":True,"users":[dict(r) for r in rows]})
            elif subpath == "users/quick-create" and method == "POST":
                d2 = body or {}
                username = d2.get("username","").strip().lower()
                email = username + "@banzai84.com"
                ws = conn.execute("SELECT id FROM workspaces LIMIT 1").fetchone()
                wid = ws["id"] if ws else 1
                ph = generate_password_hash(d2.get("password",""))
                conn.execute("INSERT INTO users (workspace_id,email,password_hash,name,role,active,created_at) VALUES (?,?,?,?,?,1,?)",
                    (wid,email,ph,d2.get("name",""),d2.get("role","seller"),datetime.utcnow().isoformat()))
                conn.commit()
                return jsonify({"ok":True,"email":email})
            elif subpath.startswith("users/") and method == "PUT":
                uid = subpath.split("/",1)[1]
                d2 = body or {}
                if "active" in d2:
                    conn.execute("UPDATE users SET active=? WHERE id=?", (d2["active"],uid))
                    conn.commit()
                return jsonify({"ok":True})
            elif subpath == "billing/plans" and method == "GET":
                rows = conn.execute("SELECT * FROM billing_plans ORDER BY created_at DESC").fetchall()
                return jsonify({"ok":True,"plans":[dict(r) for r in rows]})
            elif subpath == "billing/plans" and method == "POST":
                d2 = body or {}
                conn.execute("INSERT INTO billing_plans (name,description,price,currency,interval,active,created_at) VALUES (?,?,?,?,?,1,?)",
                    (d2.get("name"),d2.get("description",""),d2.get("price",0),d2.get("currency","USD"),d2.get("interval","month"),datetime.utcnow().isoformat()))
                conn.commit()
                return jsonify({"ok":True})
            elif subpath == "billing/payments" and method == "GET":
                rows = conn.execute("SELECT * FROM billing_payments ORDER BY created_at DESC LIMIT 50").fetchall()
                return jsonify({"ok":True,"payments":[dict(r) for r in rows]})
            elif subpath == "billing/payments" and method == "POST":
                d2 = body or {}
                from datetime import datetime as _dt
                now = _dt.utcnow().isoformat()
                last = conn.execute("SELECT banzai_invoice_number FROM billing_payments ORDER BY id DESC LIMIT 1").fetchone()
                if last and last["banzai_invoice_number"]:
                    try: num = int(last["banzai_invoice_number"].split("-")[1]) + 1
                    except: num = 1
                else: num = 1
                inv_num = f"FC-{num:05d}"
                conn.execute("INSERT INTO billing_payments (customer_name,customer_email,amount,currency,method,reference,status,banzai_invoice_number,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
                    (d2.get("customer_name"),d2.get("customer_email"),d2.get("amount",0),d2.get("currency","USD"),d2.get("method","transfer"),d2.get("reference",""),d2.get("status","paid"),inv_num,now))
                conn.commit()
                return jsonify({"ok":True,"invoice_number":inv_num})
            elif subpath == "surveys":
                rows = conn.execute("SELECT * FROM feedback_surveys ORDER BY sent_at DESC LIMIT 50").fetchall()
                pending = conn.execute("SELECT COUNT(*) as c FROM feedback_surveys WHERE update_status=\'requested\'").fetchone()
                return jsonify({"ok":True,"surveys":[dict(r) for r in rows],"pending_updates":pending["c"] if pending else 0})
            elif subpath == "surveys/send" and method == "POST":
                return jsonify({"ok":True,"sent":0,"message":"Sin clientes con email en esta sesión"})
            elif subpath == "releases":
                rows = conn.execute("SELECT * FROM release_names ORDER BY created_at DESC").fetchall()
                current = rows[0] if rows else {}
                ver = conn.execute("SELECT value FROM app_meta WHERE key=\'schema_version\'").fetchone()
                if current: current = dict(current)
                current["version"] = ver["value"] if ver else "5.0.0"
                return jsonify({"ok":True,"current":current,"releases":[dict(r) for r in rows]})
            elif subpath == "releases" and method == "POST":
                d2 = body or {}
                conn.execute("INSERT INTO release_names (release_name,description,created_at) VALUES (?,?,?)",
                    (d2.get("release_name"),d2.get("description",""),datetime.utcnow().isoformat()))
                conn.commit()
                return jsonify({"ok":True})
            elif subpath == "agents/events":
                rows = conn.execute("SELECT * FROM agent_events ORDER BY created_at DESC LIMIT 20").fetchall()
                return jsonify({"ok":True,"events":[dict(r) for r in rows]})
            elif subpath == "agents/audit" and method == "POST":
                return jsonify({"ok":True,"report":{"health":"ok","message":"Sistema operativo"}})
            elif subpath == "reports":
                inc2 = conn.execute("SELECT COALESCE(SUM(amount),0) as s FROM ledger_entries WHERE entry_type=\'income\'").fetchone()
                exp2 = conn.execute("SELECT COALESCE(SUM(amount),0) as s FROM ledger_entries WHERE entry_type=\'expense\'").fetchone()
                income2 = float(inc2["s"] if inc2 else 0)
                expenses2 = float(exp2["s"] if exp2 else 0)
                net2 = income2 - expenses2
                margin = round((net2/income2*100) if income2 > 0 else 0, 1)
                deals = conn.execute("SELECT COALESCE(SUM(value),0) as s FROM deals WHERE status=\'closed_won\'").fetchone()
                ws2 = conn.execute("SELECT currency FROM workspaces LIMIT 1").fetchone()
                return jsonify({"ok":True,"pl":{"income":income2,"expenses":expenses2,"net":net2,"margin_pct":margin},"deals":{"total_closed_value":float(deals["s"] if deals else 0)},"currency":ws2["currency"] if ws2 else "USD"})
            else:
                return jsonify({"ok":False,"error":f"Route not found: {subpath}"}), 404
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500


@app.post("/bz-admin/api/ai-chat")
@admin_required  
def bz_admin_ai_chat():
    from datetime import datetime as _dt
    d = request.get_json(force=True) or {}
    message = d.get("message","")
    history = d.get("history",[])
    if not os.environ.get("OPENAI_API_KEY"):
        return jsonify({"ok":False,"error":"OpenAI API key no configurada"}), 400
    try:
        with closing(get_db()) as conn:
            inds = conn.execute("SELECT slug,name FROM industry_playbooks LIMIT 20").fetchall()
            ind_list = ", ".join([r["name"] for r in inds])
        sys_msg = f"Sos la IA Experta de Banzai Admin. Tenés acceso completo al sistema. Rubros activos: {ind_list}. Respondé en español, sé conciso y accionable."
        msgs = [{"role":"system","content":sys_msg}] + history[-4:] + [{"role":"user","content":message}]
        import openai as _oai
        client = _oai.OpenAI(api_key=os.environ.get("OPENAI_API_KEY",""))
        resp = client.chat.completions.create(model="gpt-4o-mini",messages=msgs,max_tokens=800,temperature=0.7)
        return jsonify({"ok":True,"reply":resp.choices[0].message.content})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500


@app.get("/bz-admin/api/updates/status")
@admin_required
def bz_admin_updates_status():
    return jsonify({"ok":True,"github_configured":bool(GITHUB_TOKEN),"railway_configured":False,"github_repo":GITHUB_REPO})


@app.post("/bz-admin/api/updates/save-config")
@admin_required
def bz_admin_updates_save_config():
    global GITHUB_TOKEN, GITHUB_REPO
    d = request.get_json(force=True) or {}
    if d.get("github_token"): GITHUB_TOKEN=d["github_token"]; os.environ["GITHUB_TOKEN"]=d["github_token"]
    if d.get("github_repo"): GITHUB_REPO=d["github_repo"]; os.environ["GITHUB_REPO"]=d["github_repo"]
    if d.get("openai_key"): os.environ["OPENAI_API_KEY"]=d["openai_key"]
    return jsonify({"ok":True})


@app.get("/bz-admin/api/updates/file")
@admin_required
def bz_admin_updates_file():
    fp = request.args.get("path","app.py")
    tok = GITHUB_TOKEN
    repo = GITHUB_REPO
    if not tok: return jsonify({"ok":False,"error":"GITHUB_TOKEN no configurado"}), 400
    try:
        import base64, requests as _req
        r = _req.get(f"https://api.github.com/repos/{repo}/contents/{fp}",
                    headers={"Authorization":f"token {tok}","User-Agent":"banzai-admin"},timeout=10)
        if r.status_code != 200: return jsonify({"ok":False,"error":f"Error {r.status_code}"}), 500
        data = r.json()
        return jsonify({"ok":True,"content":base64.b64decode(data["content"]).decode("utf-8","replace"),"sha":data["sha"]})
    except Exception as e: return jsonify({"ok":False,"error":str(e)}), 500


@app.post("/bz-admin/api/updates/generate")
@admin_required
def bz_admin_updates_generate():
    d = request.get_json(force=True) or {}
    desc = d.get("description","").strip()
    fp = d.get("file_path","app.py")
    content = d.get("current_content","")
    if not desc: return jsonify({"ok":False,"error":"Descripcion requerida"}), 400
    key = os.environ.get("OPENAI_API_KEY", "")
    if not key: return jsonify({"ok":False,"error":"OpenAI API key no configurada"}), 400
    preview = (content[:5000]+"\n# ... [TRUNCATED] ...\n"+content[-2000:]) if len(content)>8000 else content
    lang = "Python/Flask" if fp.endswith(".py") else "JavaScript"
    try:
        import openai as _oai
        client = _oai.OpenAI(api_key=key)
        resp = client.chat.completions.create(model="gpt-4o-mini",messages=[
            {"role":"system","content":"Sos experto en "+lang+". Modifica "+fp+" segun instrucciones. Devuelve SOLO el codigo completo sin markdown."},
            {"role":"user","content":"Archivo ("+fp+"):\n"+preview+"\n\n---\nCAMBIO: "+desc}
        ],max_tokens=4000,temperature=0.1)
        code_out = resp.choices[0].message.content.strip()
        if code_out.startswith("```"): code_out="\n".join(code_out.split("\n")[1:])
        if code_out.endswith("```"): code_out="\n".join(code_out.split("\n")[:-1])
        return jsonify({"ok":True,"code":code_out,"file_path":fp})
    except Exception as e: return jsonify({"ok":False,"error":str(e)}), 500


@app.post("/bz-admin/api/updates/deploy")
@admin_required
def bz_admin_updates_deploy():
    d = request.get_json(force=True) or {}
    code_in = d.get("code","").strip()
    fp = d.get("file_path","app.py")
    msg = d.get("message","Update via Banzai Admin").strip()
    if not code_in: return jsonify({"ok":False,"error":"Codigo requerido"}), 400
    tok = GITHUB_TOKEN
    repo = GITHUB_REPO
    if not tok: return jsonify({"ok":False,"error":"GITHUB_TOKEN no configurado"}), 400
    try:
        import base64, requests as _req
        r = _req.get(f"https://api.github.com/repos/{repo}/contents/{fp}",
                    headers={"Authorization":f"token {tok}","User-Agent":"banzai-admin"},timeout=10)
        if r.status_code != 200: return jsonify({"ok":False,"error":f"No se pudo leer {fp}"}), 500
        sha = r.json()["sha"]
        encoded = base64.b64encode(code_in.encode()).decode()
        pr = _req.put(f"https://api.github.com/repos/{repo}/contents/{fp}",
                     headers={"Authorization":f"token {tok}","User-Agent":"banzai-admin"},
                     json={"message":msg,"content":encoded,"sha":sha},timeout=15)
        if pr.status_code not in (200,201): return jsonify({"ok":False,"error":f"GitHub {pr.status_code}"}), 500
        return jsonify({"ok":True,"message":"Actualizado: "+fp,"railway_triggered":False,"commit_url":pr.json().get("commit",{}).get("html_url","")})
    except Exception as e: return jsonify({"ok":False,"error":str(e)}), 500



# ── Plan Management API ───────────────────────────────────────────────────

@app.get("/api/plan/info")
def api_plan_info():
    user = session.get("user")
    if not user:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    ws_id = user.get("workspace_id")
    plan = get_workspace_plan(ws_id)
    plan_info = PLAN_FEATURES.get(plan, PLAN_FEATURES["trial"])
    all_plans = []
    for k, v in PLAN_FEATURES.items():
        if k != "trial":
            all_plans.append({"id": k, "name": v["name"], "price": v["price"], "currency": v["currency"], "features": v["features"], "max_users": v["max_users"]})
    return jsonify({"ok": True, "current_plan": plan, "plan_info": plan_info, "all_plans": all_plans, "feature_labels": FEATURE_LABELS})


@app.get("/api/plan/check/<feature>")
def api_plan_check(feature):
    user = session.get("user")
    if not user:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    ws_id = user.get("workspace_id")
    has_it = workspace_has_feature(ws_id, feature)
    return jsonify({"ok": True, "has_feature": has_it, "feature": feature})


@app.get("/api/upgrade/info")
def api_upgrade_info():
    """Show upgrade options for current plan."""
    user = session.get("user")
    if not user:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    ws_id = user.get("workspace_id")
    plan = get_workspace_plan(ws_id)
    upgrades = []
    found_current = False
    for k, v in PLAN_FEATURES.items():
        if k == "trial":
            continue
        if k == plan:
            found_current = True
            continue
        if found_current:
            upgrades.append({"id": k, "name": v["name"], "price": v["price"], "features": v["features"]})
    return jsonify({"ok": True, "current_plan": plan, "upgrades": upgrades})


@app.post("/bz-admin/api/proxy/workspaces/plan")
def admin_set_workspace_plan():
    if not session.get("bz_admin_ok"):
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    d = request.get_json(force=True) or {}
    plan = d.get("plan", "trial")
    workspace_id = d.get("workspace_id", 1)
    if plan not in PLAN_FEATURES:
        return jsonify({"ok": False, "error": "Plan invalido"}), 400
    try:
        with closing(get_db()) as conn:
            # Set trial end date for trial plan
            if plan == "trial":
                from datetime import datetime as _dt3, timedelta
                trial_end = (_dt3.utcnow() + timedelta(days=30)).isoformat()
                conn.execute("UPDATE workspaces SET plan=%s, trial_ends_at=%s WHERE id=%s", (plan, trial_end, workspace_id))
            else:
                conn.execute("UPDATE workspaces SET plan=%s, trial_ends_at=NULL WHERE id=%s", (plan, workspace_id))
            conn.commit()
        return jsonify({"ok": True, "plan": plan, "plan_name": PLAN_FEATURES[plan]["name"]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.get("/bz-admin/api/proxy/workspaces/plan")
def admin_get_workspace_plan():
    if not session.get("bz_admin_ok"):
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    workspace_id = request.args.get("workspace_id", 1)
    plan = get_workspace_plan(workspace_id)
    plan_info = PLAN_FEATURES.get(plan, PLAN_FEATURES["trial"])
    return jsonify({"ok": True, "plan": plan, "plan_info": plan_info, "all_plans": list(PLAN_FEATURES.keys()), "plan_features": PLAN_FEATURES})



@app.get("/bz-admin/plans")
def bz_admin_plans_page():
    if not session.get("bz_admin_ok"):
        return redirect("/bz-admin/login")
    plan_list = []
    for k, v in PLAN_FEATURES.items():
        plan_list.append({"id": k, "name": v["name"], "price": v["price"], "features": v["features"], "max_users": v["max_users"]})
    import json
    plans_json = json.dumps(plan_list)
    feature_labels_json = json.dumps(FEATURE_LABELS)
    return render_template_string("""<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Planes - Banzai Admin</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0;}
:root{--bg:#06080F;--surface:#0C0F1A;--surface2:#111520;--surface3:#161B2E;--border:rgba(255,255,255,.07);--border2:rgba(255,255,255,.12);--text:#F0F2FA;--text2:#9AA3C2;--text3:#4E5B7A;--brand:#6366F1;--brand-dim:rgba(99,102,241,.12);--green:#10B981;--green-dim:rgba(16,185,129,.12);--amber:#F59E0B;--amber-dim:rgba(245,158,11,.12);--red:#EF4444;}
body{font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);padding:32px;}
h1{font-size:28px;font-weight:800;color:#6366F1;margin-bottom:8px;}
.sub{color:var(--text2);font-size:14px;margin-bottom:32px;}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:16px;}
.card{background:var(--surface);border:1px solid var(--border);border-radius:16px;padding:24px;}
.card.trial{border-color:rgba(245,158,11,.3);}
.card.starter{border-color:rgba(99,102,241,.3);}
.card.basic{border-color:rgba(99,102,241,.4);}
.card.pro{border-color:rgba(99,102,241,.5);}
.card.growth{border-color:rgba(16,185,129,.4);}
.card.enterprise{border-color:rgba(16,185,129,.6);background:rgba(16,185,129,.04);}
.plan-name{font-size:20px;font-weight:700;margin-bottom:4px;}
.plan-price{font-size:32px;font-weight:800;color:#6366F1;margin-bottom:16px;}
.plan-price span{font-size:14px;color:var(--text2);font-weight:400;}
.feat{font-size:12px;color:var(--text2);margin-bottom:4px;padding-left:12px;position:relative;}
.feat:before{content:"✓";position:absolute;left:0;color:#10B981;font-weight:700;}
.set-plan{margin-top:20px;display:flex;gap:8px;align-items:center;}
select,input{background:var(--surface3);border:1px solid var(--border2);border-radius:8px;color:var(--text);font:inherit;font-size:13px;padding:8px 12px;}
button{background:#6366F1;color:#fff;border:none;border-radius:8px;padding:9px 18px;font:inherit;font-size:13px;font-weight:600;cursor:pointer;}
button:hover{background:#4F46E5;}
.back{display:inline-flex;align-items:center;gap:6px;color:var(--text2);font-size:13px;text-decoration:none;margin-bottom:24px;}
.back:hover{color:var(--text);}
.msg{margin-top:12px;font-size:12px;padding:8px 12px;border-radius:8px;display:none;}
.msg.ok{background:rgba(16,185,129,.1);color:#10B981;border:1px solid rgba(16,185,129,.2);}
.msg.err{background:rgba(239,68,68,.1);color:#EF4444;border:1px solid rgba(239,68,68,.2);}
.ws-section{background:var(--surface);border:1px solid var(--border);border-radius:16px;padding:24px;margin-bottom:32px;}
.ws-section h2{font-size:16px;font-weight:700;margin-bottom:16px;}
</style></head><body>
<a class="back" href="/bz-admin/">← Volver al Admin</a>
<h1>Planes de Banzai84</h1>
<p class="sub">Asigna planes a los workspaces de tus clientes desde aca.</p>

<div class="ws-section">
<h2>Cambiar plan del workspace activo</h2>
<div style="display:flex;gap:12px;align-items:center;flex-wrap:wrap;">
<select id="sel-plan">
</select>
<button onclick="setPlan()">Aplicar plan</button>
<div id="plan-msg" class="msg"></div>
</div>
</div>

<div class="grid" id="plans-grid"></div>

<script>
const PLANS = """ + plans_json + """;
const LABELS = """ + feature_labels_json + """;

const sel = document.getElementById('sel-plan');
PLANS.forEach(p => {
  const o = document.createElement('option');
  o.value = p.id;
  o.textContent = p.name + (p.price ? ' - $' + p.price + '/mes' : ' - GRATIS 30 dias');
  sel.appendChild(o);
});

const grid = document.getElementById('plans-grid');
PLANS.forEach(p => {
  const card = document.createElement('div');
  card.className = 'card ' + p.id;
  const feats = p.features.map(f => '<div class="feat">'+(LABELS[f]||f)+'</div>').join('');
  card.innerHTML = '<div class="plan-name">'+p.name+'</div>'
    +'<div class="plan-price">'+(p.price ? '$'+p.price : 'GRATIS')+'<span>'+(p.price ? '/mes' : ' por 30 dias')+'</span></div>'
    +'<div style="font-size:11px;color:var(--text3);margin-bottom:12px;">Hasta '+p.max_users+' usuario'+(p.max_users>1?'s':'')+'</div>'
    +feats;
  grid.appendChild(card);
});

async function setPlan() {
  const plan = document.getElementById('sel-plan').value;
  const msg = document.getElementById('plan-msg');
  try {
    const r = await fetch('/bz-admin/api/proxy/workspaces/plan', {method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({plan,workspace_id:1})});
    const d = await r.json();
    msg.className = 'msg ' + (d.ok ? 'ok' : 'err');
    msg.textContent = d.ok ? 'Plan actualizado: ' + d.plan_name : d.error;
    msg.style.display = 'block';
    setTimeout(() => { msg.style.display = 'none'; }, 3000);
  } catch(e) {
    msg.className = 'msg err'; msg.textContent = 'Error: ' + e.message; msg.style.display = 'block';
  }
}
</script>
</body></html>""")



# Twilio WhatsApp Webhook
@app.post("/webhook/whatsapp")
def twilio_whatsapp_webhook():
    """Receives inbound WhatsApp messages from Twilio and routes them to the Sales Agent."""
    try:
        from_number = request.form.get("From", "").replace("whatsapp:", "")
        to_number = request.form.get("To", "").replace("whatsapp:", "")
        body = request.form.get("Body", "").strip()
        profile_name = request.form.get("ProfileName", "Cliente WhatsApp")

        if not from_number or not body:
            return ("", 200)

        with closing(get_db()) as conn:
            # Route to the correct business by the WhatsApp number the customer wrote to.
            # Critical for multi-tenant: each business has its own number, so messages
            # must never fall back to "just pick the first workspace ever created".
            ws_row = None
            if to_number:
                ws_row = conn.execute(
                    "SELECT id, region, currency, language FROM workspaces WHERE whatsapp_number = ?",
                    (to_number,)
                ).fetchone()
            if not ws_row:
                # Fallback only for single-tenant setups where exactly one workspace exists
                # and none has a number configured yet — avoids silently misrouting when
                # multiple businesses are active.
                all_ws = conn.execute("SELECT id, region, currency, language FROM workspaces LIMIT 2").fetchall()
                if len(all_ws) == 1:
                    ws_row = all_ws[0]
                else:
                    _logger.error(f"WhatsApp webhook: no workspace matches number {to_number} and multiple workspaces exist — message dropped to avoid misrouting")
                    return ("", 200)
            if not ws_row:
                return ("", 200)
            workspace_id = ws_row["id"]
            currency = ws_row["currency"] if ws_row["currency"] else "USD"
            language = ws_row["language"] if ws_row["language"] else "es"

            conv_row = conn.execute(
                "SELECT id FROM conversations WHERE workspace_id=? AND customer_phone=? AND status='open' ORDER BY id DESC LIMIT 1",
                (workspace_id, from_number)
            ).fetchone()

            now_iso = datetime.utcnow().isoformat()
            if conv_row:
                conversation_id = conv_row["id"]
                conn.execute("UPDATE conversations SET updated_at=? WHERE id=?", (now_iso, conversation_id))
            else:
                cur = conn.execute(
                    "INSERT INTO conversations (workspace_id,customer_name,customer_phone,channel,status,country,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?)",
                    (workspace_id, profile_name, from_number, "whatsapp", "open", "AR", now_iso, now_iso)
                )
                conversation_id = cur.lastrowid
            conn.commit()

            conn.execute(
                "INSERT INTO messages (conversation_id,role,text,created_at) VALUES (?,?,?,?)",
                (conversation_id, "customer", body, now_iso)
            )
            conn.commit()

        result = orchestrator_process_message(
            workspace_id=workspace_id,
            conversation_id=conversation_id,
            customer_name=profile_name,
            customer_message=body,
            language=language,
            currency=currency,
        )

        reply_text = result.get("reply", "Gracias por tu mensaje, en breve te respondemos.")

        with closing(get_db()) as conn:
            conn.execute(
                "INSERT INTO messages (conversation_id,role,text,created_at) VALUES (?,?,?,?)",
                (conversation_id, "agent", reply_text, datetime.utcnow().isoformat())
            )
            conn.commit()

        try:
            send_whatsapp_text(from_number, reply_text)
        except Exception as send_err:
            print(f"WhatsApp send error: {send_err}")

        return ("", 200)
    except Exception as e:
        print(f"Webhook error: {e}")
        return ("", 200)


@app.get("/webhook/whatsapp")
def twilio_whatsapp_webhook_verify():
    return ("OK", 200)



# ── Vendor Invoices API ─────────────────────────────────────────────────────

@app.get("/api/vendor-invoices")
def api_vendor_invoices_list():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if not user_can(user, "see_finances"):
        return json_error("Permission denied", 403)
    status_filter = request.args.get("status", "")
    search_q = request.args.get("q", "").strip()
    vendor_filter = request.args.get("vendor", "").strip()

    query = "SELECT * FROM vendor_invoices WHERE workspace_id = ?"
    params = [user["workspace_id"]]

    if status_filter:
        query += " AND status = ?"
        params.append(status_filter)

    if vendor_filter:
        query += " AND vendor_name = ?"
        params.append(vendor_filter)

    if search_q:
        query += " AND (invoice_number LIKE ? OR vendor_name LIKE ?)"
        like_q = f"%{search_q}%"
        params.extend([like_q, like_q])

    query += " ORDER BY processed_at DESC LIMIT 300"

    with closing(get_db()) as conn:
        rows = conn.execute(query, params).fetchall()

    invoices = [dict(r) for r in rows]
    total_pending = sum(i["amount"] or 0 for i in invoices if i["status"] == "pending")
    total_paid = sum(i["amount"] or 0 for i in invoices if i["status"] == "paid")
    return jsonify({"ok": True, "invoices": invoices, "total_pending": total_pending, "total_paid": total_paid})


@app.get("/reportes/cuentas-por-pagar")
def vendor_invoices_printable_report():
    """Printable/PDF-ready Accounts Payable report."""
    try:
        user = require_auth()
    except PermissionError:
        return redirect("/")
    if not user_can(user, "see_finances"):
        return "No autorizado", 403

    with closing(get_db()) as conn:
        ws = conn.execute("SELECT name, currency FROM workspaces WHERE id = ?", (user["workspace_id"],)).fetchone()
        pending_rows = conn.execute(
            """SELECT * FROM vendor_invoices
               WHERE workspace_id = ? AND status = 'pending'
               ORDER BY vendor_name, due_date""",
            (user["workspace_id"],)
        ).fetchall()

    pending = [dict(r) for r in pending_rows]
    workspace_name = ws["name"] if ws else "Negocio"
    currency = ws["currency"] if ws else "USD"

    # Group by vendor
    by_vendor = {}
    for inv in pending:
        vname = inv.get("vendor_name") or "Proveedor sin identificar"
        by_vendor.setdefault(vname, []).append(inv)

    grand_total = sum(inv.get("amount") or 0 for inv in pending)

    def fmt_money(v):
        try:
            return f"{currency} {v:,.2f}"
        except Exception:
            return f"{currency} {v}"

    vendor_sections = []
    for vname, invs in sorted(by_vendor.items(), key=lambda kv: -sum(i.get('amount') or 0 for i in kv[1])):
        vendor_total = sum(inv.get("amount") or 0 for inv in invs)
        rows_html = "".join(
            f"<tr>"
            f"<td>{inv.get('invoice_number') or '—'}</td>"
            f"<td>{inv.get('invoice_date') or '—'}</td>"
            f"<td>{inv.get('due_date') or '—'}</td>"
            f"<td class='amt'>{fmt_money(inv.get('amount') or 0)}</td>"
            f"</tr>"
            for inv in invs
        )
        vendor_sections.append(f"""
        <div class="vendor-block">
          <div class="vendor-header">
            <span class="vendor-name">{vname}</span>
            <span class="vendor-total">{fmt_money(vendor_total)}</span>
          </div>
          <table>
            <thead><tr><th>N° Factura</th><th>Fecha</th><th>Vencimiento</th><th class="amt">Monto</th></tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
        </div>""")

    vendor_html = "".join(vendor_sections) if vendor_sections else '<p class="empty-msg">No hay facturas pendientes de pago.</p>'

    return render_template_string("""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Cuentas por Pagar — {{ workspace_name }}</title>
<style>
  * { box-sizing: border-box; }
  body { font-family: 'Segoe UI', Arial, sans-serif; color: #1F2937; margin: 0; padding: 32px; background: #fff; }
  .header { display: flex; justify-content: space-between; align-items: flex-start; border-bottom: 3px solid #6366F1; padding-bottom: 16px; margin-bottom: 24px; }
  .header h1 { font-size: 22px; margin: 0 0 4px; color: #6366F1; }
  .header .sub { font-size: 12px; color: #6B7280; }
  .header .date { font-size: 11px; color: #9CA3AF; text-align: right; }
  .summary-box { background: #EEF2FF; border-radius: 10px; padding: 18px 24px; margin-bottom: 28px; display: flex; justify-content: space-between; align-items: center; }
  .summary-box .label { font-size: 11px; text-transform: uppercase; letter-spacing: .06em; color: #6366F1; font-weight: 700; }
  .summary-box .value { font-size: 28px; font-weight: 800; color: #4F46E5; }
  .vendor-block { margin-bottom: 22px; page-break-inside: avoid; }
  .vendor-header { display: flex; justify-content: space-between; align-items: center; background: #F3F4F6; padding: 8px 12px; border-radius: 6px 6px 0 0; }
  .vendor-name { font-weight: 700; font-size: 13px; }
  .vendor-total { font-weight: 700; font-size: 13px; color: #F59E0B; }
  table { width: 100%; border-collapse: collapse; font-size: 12px; }
  th { text-align: left; padding: 6px 12px; border-bottom: 2px solid #E5E7EB; color: #6B7280; font-weight: 600; font-size: 11px; text-transform: uppercase; }
  td { padding: 7px 12px; border-bottom: 1px solid #F3F4F6; }
  .amt { text-align: right; }
  .empty-msg { color: #6B7280; font-style: italic; padding: 20px; text-align: center; }
  .footer { margin-top: 32px; padding-top: 12px; border-top: 1px solid #E5E7EB; font-size: 10px; color: #9CA3AF; text-align: center; }
  .print-btn { position: fixed; top: 20px; right: 20px; background: #6366F1; color: #fff; border: none; padding: 10px 18px; border-radius: 8px; font-size: 13px; font-weight: 600; cursor: pointer; }
  @media print {
    .print-btn { display: none; }
    body { padding: 0; }
  }
</style>
</head>
<body>
  <button class="print-btn" onclick="window.print()">🖨 Imprimir / Guardar PDF</button>
  <div class="header">
    <div>
      <h1>Cuentas por Pagar</h1>
      <div class="sub">{{ workspace_name }}</div>
    </div>
    <div class="date">Generado el {{ today }}</div>
  </div>

  <div class="summary-box">
    <div>
      <div class="label">Total adeudado a proveedores</div>
    </div>
    <div class="value">{{ grand_total_fmt }}</div>
  </div>

  {{ vendor_html | safe }}

  <div class="footer">Banzai84 — Reporte generado automáticamente</div>
</body>
</html>""",
        workspace_name=workspace_name,
        today=datetime.utcnow().strftime("%d/%m/%Y %H:%M"),
        grand_total_fmt=fmt_money(grand_total),
        vendor_html=vendor_html,
    )


@app.get("/api/vendor-invoices/by-vendor")
def api_vendor_invoices_by_vendor():
    """Group invoices by vendor for the vendor-centric view."""
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if not user_can(user, "see_finances"):
        return json_error("Permission denied", 403)
    with closing(get_db()) as conn:
        rows = conn.execute(
            """SELECT vendor_name,
                      COUNT(*) as invoice_count,
                      SUM(CASE WHEN status='pending' THEN amount ELSE 0 END) as pending_total,
                      SUM(CASE WHEN status='paid' THEN amount ELSE 0 END) as paid_total,
                      MAX(processed_at) as last_invoice_at
               FROM vendor_invoices
               WHERE workspace_id = ? AND vendor_name IS NOT NULL
               GROUP BY vendor_name
               ORDER BY last_invoice_at DESC""",
            (user["workspace_id"],)
        ).fetchall()
    vendors = [dict(r) for r in rows]
    return jsonify({"ok": True, "vendors": vendors})


@app.get("/api/vendor-invoices/<int:invoice_id>")
def api_vendor_invoice_detail(invoice_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if not user_can(user, "see_finances"):
        return json_error("Permission denied", 403)
    with closing(get_db()) as conn:
        inv = conn.execute(
            "SELECT * FROM vendor_invoices WHERE id = ? AND workspace_id = ?",
            (invoice_id, user["workspace_id"])
        ).fetchone()
        if not inv:
            return json_error("Invoice not found", 404)
        inv = dict(inv)
        has_file = bool(inv.get("file_base64"))
        inv.pop("file_base64", None)  # don't ship the huge base64 blob in the JSON detail
        inv["has_file"] = has_file
        ledger_entry = None
        if inv.get("ledger_entry_id"):
            led = conn.execute(
                "SELECT * FROM ledger_entries WHERE id = ?", (inv["ledger_entry_id"],)
            ).fetchone()
            if led:
                ledger_entry = dict(led)
    return jsonify({"ok": True, "invoice": inv, "ledger_entry": ledger_entry})


@app.get("/api/vendor-invoices/<int:invoice_id>/file")
def api_vendor_invoice_download_file(invoice_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if not user_can(user, "see_finances"):
        return json_error("Permission denied", 403)
    with closing(get_db()) as conn:
        inv = conn.execute(
            "SELECT file_base64, file_mime_type, file_name FROM vendor_invoices WHERE id = ? AND workspace_id = ?",
            (invoice_id, user["workspace_id"])
        ).fetchone()
    if not inv or not inv["file_base64"]:
        return json_error("Archivo no disponible para esta factura", 404)
    import base64 as _b64mod
    import io as _iomod
    file_bytes = _b64mod.b64decode(inv["file_base64"])
    return send_file(
        _iomod.BytesIO(file_bytes),
        mimetype=inv["file_mime_type"] or "application/pdf",
        as_attachment=False,
        download_name=inv["file_name"] or f"factura_{invoice_id}.pdf",
    )


@app.post("/api/vendor-invoices/<int:invoice_id>/mark-paid")
def api_vendor_invoice_mark_paid(invoice_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if not user_can(user, "see_finances"):
        return json_error("Permission denied", 403)
    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        inv = conn.execute("SELECT * FROM vendor_invoices WHERE id = ? AND workspace_id = ?",
                            (invoice_id, user["workspace_id"])).fetchone()
        if not inv:
            return json_error("Invoice not found", 404)
        inv = dict(inv)
        if inv["status"] == "paid":
            return jsonify({"ok": True, "message": "Ya estaba marcada como pagada"})

        ledger_cur = conn.execute(
            """INSERT INTO ledger_entries
               (workspace_id, entry_type, concept, category, amount, currency, state, due_date, created_at)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (user["workspace_id"], "Expense",
             f"Factura proveedor {inv.get('vendor_name','')} — {inv.get('invoice_number','')}",
             "Proveedores", inv["amount"], inv["currency"], "Paid", now[:10], now)
        )
        conn.execute("UPDATE vendor_invoices SET status = 'paid', paid_at = ?, ledger_entry_id = ? WHERE id = ?",
                     (now, ledger_cur.lastrowid, invoice_id))
        conn.commit()
    return jsonify({"ok": True})


@app.put("/api/vendor-invoices/<int:invoice_id>")
def api_vendor_invoice_update(invoice_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if not user_can(user, "see_finances"):
        return json_error("Permission denied", 403)
    payload = request.get_json(force=True)
    if "amount" in payload and not is_positive_number(payload.get("amount"), allow_zero=False):
        return json_error("El monto de la factura debe ser mayor a cero")
    fields = {}
    for k in ("vendor_name", "invoice_number", "invoice_date", "due_date", "amount", "currency"):
        if k in payload:
            fields[k] = sanitize_text(payload[k]) if isinstance(payload[k], str) else payload[k]
    if not fields:
        return json_error("Nothing to update")
    set_clause = ", ".join(f"{k} = ?" for k in fields)
    values = list(fields.values()) + [invoice_id, user["workspace_id"]]
    with closing(get_db()) as conn:
        conn.execute(f"UPDATE vendor_invoices SET {set_clause}, needs_review = 0 WHERE id = ? AND workspace_id = ?", values)
        conn.commit()
    return jsonify({"ok": True})


@app.delete("/api/vendor-invoices/<int:invoice_id>")
def api_vendor_invoice_delete(invoice_id: int):
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if not user_can(user, "see_finances"):
        return json_error("Permission denied", 403)
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM vendor_invoices WHERE id = ? AND workspace_id = ?", (invoice_id, user["workspace_id"]))
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/email-inbox/config")
def api_email_inbox_get_config():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if not user_can(user, "manage_settings"):
        return json_error("Permission denied", 403)
    with closing(get_db()) as conn:
        cfg = conn.execute("SELECT id, imap_host, imap_port, email_address, last_checked_at, active FROM email_inbox_config WHERE workspace_id = ?",
                            (user["workspace_id"],)).fetchone()
    return jsonify({"ok": True, "config": dict(cfg) if cfg else None})


@app.post("/api/email-inbox/config")
def api_email_inbox_save_config():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if not user_can(user, "manage_settings"):
        return json_error("Permission denied", 403)
    payload = request.get_json(force=True)
    imap_host = payload.get("imap_host", "").strip()
    email_address = payload.get("email_address", "").strip()
    email_password = payload.get("email_password", "").strip()
    imap_port = int(payload.get("imap_port", 993))

    common_hosts = {
        "gmail.com": "imap.gmail.com",
        "outlook.com": "outlook.office365.com",
        "hotmail.com": "outlook.office365.com",
        "yahoo.com": "imap.mail.yahoo.com",
    }
    if not imap_host and "@" in email_address:
        domain = email_address.split("@")[1].lower()
        imap_host = common_hosts.get(domain, f"imap.{domain}")

    if not email_address or not email_password or not imap_host:
        return json_error("Faltan datos: email_address, email_password, imap_host")

    now = datetime.utcnow().isoformat()
    with closing(get_db()) as conn:
        existing = conn.execute("SELECT id FROM email_inbox_config WHERE workspace_id = ?", (user["workspace_id"],)).fetchone()
        if existing:
            conn.execute(
                "UPDATE email_inbox_config SET imap_host=?, imap_port=?, email_address=?, email_password=?, active=1 WHERE workspace_id=?",
                (imap_host, imap_port, email_address, email_password, user["workspace_id"])
            )
        else:
            conn.execute(
                "INSERT INTO email_inbox_config (workspace_id, imap_host, imap_port, email_address, email_password, active, created_at) VALUES (?,?,?,?,?,1,?)",
                (user["workspace_id"], imap_host, imap_port, email_address, email_password, now)
            )
        conn.commit()
    return jsonify({"ok": True, "imap_host": imap_host})


@app.post("/api/email-inbox/check-now")
def api_email_inbox_check_now():
    try:
        user = require_auth()
    except PermissionError:
        return json_error("Not authenticated", 401)
    if not user_can(user, "manage_settings"):
        return json_error("Permission denied", 403)
    result = check_vendor_invoice_inbox(user["workspace_id"])
    return jsonify(result)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV", "production") == "development"
    app.run(host="0.0.0.0", port=port, debug=debug)
